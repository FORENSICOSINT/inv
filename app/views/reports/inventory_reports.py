# views/reports.py
from django.shortcuts import *
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from datetime import date, timedelta, time
import json
import csv
import random
import decimal
import xlsxwriter
import io
import calendar
from .reports import InventoryReports
from django.utils import timezone
from app.models.transactions import *
from app.models.products import *
from app.models.customers import *
from app.models.finance import *
from app.models.human_resource import *
from django.db.models.functions import TruncDate, Coalesce
from django.core.serializers.json import DjangoJSONEncoder
from decimal import Decimal
import pandas as pd
from reportlab.lib.styles import *
import csv
import json
from django.http import HttpResponse
from django.db.models import *
from decimal import Decimal, DecimalTuple
from django.utils import timezone
from reportlab.lib import colors
from reportlab.lib.pagesizes import *
from reportlab.platypus import *

from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg') 
import traceback
# import xlwt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.db.models.functions import *
from django.core.paginator import *
from django.views.decorators.http import *
from app.models.suppliers import Supplier

# For Excel export (if using pandas/openpyxl)
try:
    import pandas as pd
except ImportError:
    pd = None
    print("Warning: pandas not installed. Excel export may not work properly.")

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("Warning: openpyxl not installed. Excel export may not work properly.")

# For PDF export (reportlab)
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
except ImportError:
    print("Warning: reportlab not installed. PDF export may not work properly.")
    # Define placeholders to avoid ImportError
    class canvas:
        pass
    letter = None
    class SimpleDocTemplate:
        pass
    class Table:
        pass
    class TableStyle:
        pass
    class Paragraph:
        pass
    class getSampleStyleSheet:
        @staticmethod
        def getSampleStyleSheet():
            return {}
    colors = None



REVENUE_EXPR = ExpressionWrapper(
    F('quantity') * F('sale_price'),
    output_field=DecimalField(max_digits=16, decimal_places=2)
)

COST_EXPR = ExpressionWrapper(
    F('quantity') * F('unit_cost'),
    output_field=DecimalField(max_digits=16, decimal_places=2)
)

# Custom JSON encoder class (add this after imports)
class CustomJSONEncoder(DjangoJSONEncoder):
    """Custom JSON encoder to handle dates and decimals"""
    def default(self, obj):
        if isinstance(obj, (date, timezone.date)):
            return obj.isoformat()
        if isinstance(obj, date.date):
            return obj.isoformat()
        if isinstance(obj, decimal.Decimal):
            return float(obj)
        if hasattr(obj, '__dict__'):
            return str(obj)
        return super().default(obj)



# ============================================================================
# REPORTS DASHBOARD & OVERVIEW VIEWS
# ============================================================================

@login_required
def reports_dashboard(request):
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/dashboard.html', context)

@login_required
def reports_details(request):
    """Reports dashboard view"""
    context = {}
    return render(request, 'reports/details.html', context)


# ============================================================================
# PURCHASE REPORTS VIEWS
# ============================================================================
# Supplier Purchase Summary, Purchase Trend Analysis, Purchase Order Status,
# Item-wise Purchase Analysis, Expiry Tracking Report
# ============================================================================



@login_required
def purchase_details(request):
    """
    Purchase Performance Dashboard view.
    Matches: reports/purchase_details.html
    """
    today = timezone.now().date()

    # ── Filters from request ────────────────────────────────────────────────
    supplier_id  = request.GET.get('supplier', '')
    status_param = request.GET.get('status', '')
    store_id     = request.GET.get('store', '')
    daterange    = request.GET.get('daterange', '')

    # ── Date range parsing — no default restriction ─────────────────────────
    start_date = None
    end_date   = None

    if daterange:
        try:
            parts = [p.strip() for p in daterange.split(' - ')]
            if len(parts) == 2:
                start_date = date.fromisoformat(parts[0])
                end_date   = date.fromisoformat(parts[1])
        except (ValueError, AttributeError):
            pass

    # ── Base queryset ───────────────────────────────────────────────────────
    purchase_orders = PurchaseOrder.objects.select_related(
        'supplier', 'store', 'store__branch'
    )

    if start_date and end_date:
        purchase_orders = purchase_orders.filter(purchase_date__range=[start_date, end_date])
    if supplier_id:
        purchase_orders = purchase_orders.filter(supplier_id=supplier_id)
    if status_param:
        purchase_orders = purchase_orders.filter(status__iexact=status_param)
    if store_id:
        purchase_orders = purchase_orders.filter(store_id=store_id)

    purchase_orders = purchase_orders.order_by('-purchase_date', '-id')

    # ── KPI Metrics ─────────────────────────────────────────────────────────
    total_purchases = purchase_orders.aggregate(
        total=Sum('total_cost')
    )['total'] or Decimal('0')

    total_orders    = purchase_orders.count()
    avg_order_value = (total_purchases / total_orders) if total_orders > 0 else Decimal('0')

    total_suppliers = purchase_orders.values('supplier').distinct().count()

    new_suppliers_qs = Supplier.objects.filter(purchaseorder__isnull=False)
    if start_date and end_date:
        new_suppliers_qs = new_suppliers_qs.filter(
            purchaseorder__purchase_date__range=[start_date, end_date]
        )
    new_suppliers = new_suppliers_qs.distinct().count()

    # ── Items metrics ───────────────────────────────────────────────────────
    items_qs = PurchaseOrderItem.objects.filter(order__in=purchase_orders)

    total_items_purchased = items_qs.values('product').distinct().count()
    total_units_purchased = items_qs.aggregate(
        total=Sum('base_quantity')
    )['total'] or 0
    total_items_cost = items_qs.aggregate(
        total=Sum(F('base_quantity') * F('unit_cost'))
    )['total'] or Decimal('0')

    # ── Today's stats ───────────────────────────────────────────────────────
    today_orders    = PurchaseOrder.objects.filter(purchase_date=today).count()
    today_purchases = PurchaseOrder.objects.filter(
        purchase_date=today
    ).aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    due_today       = PurchaseOrder.objects.filter(
        expected_date=today,
        status__in=['pending', 'in_progress']
    ).count()
    overdue_orders  = PurchaseOrder.objects.filter(
        expected_date__lt=today,
        status__in=['pending', 'in_progress']
    ).count()

    # ── Status summary ──────────────────────────────────────────────────────
    status_summary = purchase_orders.values('status').annotate(
        count=Count('id'),
        total=Sum('total_cost'),
    ).order_by('status')

    # ── Top suppliers ───────────────────────────────────────────────────────
    top_suppliers = purchase_orders.values(
        name=F('supplier__name')
    ).annotate(
        total=Sum('total_cost'),
        order_count=Count('id'),
    ).order_by('-total')[:5]

    # ── Expiry stats ────────────────────────────────────────────────────────
    batch_qs = InventoryBatch.objects.filter(
        purchase_order_item__order__in=purchase_orders,
        remaining_quantity__gt=0,
        expiry_date__isnull=False,
    )
    expired_count       = batch_qs.filter(expiry_date__lt=today).count()
    expiring_soon_count = batch_qs.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).count()
    good_stock_count    = batch_qs.filter(
        expiry_date__gt=today + timedelta(days=30)
    ).count()
    total_value_at_risk = batch_qs.filter(
        expiry_date__lte=today + timedelta(days=30)
    ).aggregate(
        val=Sum(F('remaining_quantity') * F('unit_cost'))
    )['val'] or Decimal('0')

    expiry_stats = {
        'expired_count':       expired_count,
        'expiring_soon_count': expiring_soon_count,
        'good_stock_count':    good_stock_count,
        'total_value_at_risk': total_value_at_risk,
    }

    # ── Purchase items table ────────────────────────────────────────────────
    purchase_items = PurchaseOrderItem.objects.filter(
        order__in=purchase_orders
    ).select_related(
        'product', 'product__category', 'unit', 'order', 'order__supplier'
    ).order_by('order__purchase_date', 'product__name')

    for item in purchase_items:
        item.days_to_expiry = (item.expiry_date - today).days if item.expiry_date else None

    # ── Filter dropdown options ─────────────────────────────────────────────
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')
    stores    = StoreLocation.objects.filter(is_active=True).order_by('name')

    # Display dates — fall back to earliest/latest PO dates if no range selected
    display_start = start_date or PurchaseOrder.objects.order_by('purchase_date').values_list('purchase_date', flat=True).first() or today
    display_end   = end_date or today

    context = {
        'start_date':     display_start,
        'end_date':       display_end,
        'generated_date': timezone.now(),
        'now':            timezone.now(),

        # KPIs
        'total_purchases':       total_purchases,
        'total_orders':          total_orders,
        'avg_order_value':       avg_order_value,
        'total_suppliers':       total_suppliers,
        'new_suppliers':         new_suppliers,
        'total_items_purchased': total_items_purchased,
        'total_units_purchased': total_units_purchased,
        'total_items_cost':      total_items_cost,

        # Today
        'today_orders':    today_orders,
        'today_purchases': today_purchases,
        'due_today':       due_today,
        'overdue_orders':  overdue_orders,

        # Tables
        'purchase_orders': purchase_orders,
        'purchase_items':  purchase_items,

        # Summary cards
        'status_summary': status_summary,
        'top_suppliers':  top_suppliers,
        'expiry_stats':   expiry_stats,

        # Filter dropdowns
        'suppliers': suppliers,
        'stores':    stores,
    }

    return render(request, 'reports/purchase_details.html', context)



def calculate_supplier_performance_fixed(supplier_id, start_date, end_date):
    """
    FIXED supplier performance rating - UGX 9M order should not be 'Poor'
    """
    orders = PurchaseOrder.objects.filter(
        supplier_id=supplier_id,
        purchase_date__range=[start_date, end_date]
    )
    
    if not orders.exists():
        return 'No Data'
    
    total_purchases = orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = orders.count()
    completed_orders = orders.filter(status__in=['completed', 'COMPLETED']).count()
    
    # Calculate completion rate
    completion_rate = (completed_orders / total_orders * 100) if total_orders > 0 else 0
    
    # Calculate average order value
    avg_order = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # NEW FIXED LOGIC: High value orders get better rating
    if total_orders == 0:
        return 'No Data'
    
    # Score based on multiple factors
    score = 0
    
    # High total purchases = high score
    if total_purchases > 8000000:  # > 8M UGX
        score += 4
    elif total_purchases > 3000000:  # > 3M UGX
        score += 3
    elif total_purchases > 1000000:  # > 1M UGX
        score += 2
    elif total_purchases > 500000:  # > 500K UGX
        score += 1
    
    # High average order value = high score
    if avg_order > 5000000:  # > 5M UGX average
        score += 3
    elif avg_order > 1000000:  # > 1M UGX average
        score += 2
    elif avg_order > 500000:  # > 500K UGX average
        score += 1
    
    # Good completion rate = high score
    if completion_rate >= 90:
        score += 3
    elif completion_rate >= 75:
        score += 2
    elif completion_rate >= 50:
        score += 1
    
    # Determine rating based on score
    if score >= 8:
        return 'Excellent'
    elif score >= 6:
        return 'Very Good'
    elif score >= 4:
        return 'Good'
    elif score >= 2:
        return 'Average'
    else:
        return 'Poor'


def get_purchase_trend_data_corrected(report_type, start_date, end_date):
    """
    FIXED purchase trend analysis with correct growth calculations
    """
    if report_type == 'monthly':
        periods = []
        today = timezone.now().date()
        
        # Get last 6 months including current
        for i in range(5, -1, -1):
            month_date = start_date - timedelta(days=30*i)
            month_start = date(month_date.year, month_date.month, 1)
            
            if month_start.month == 12:
                month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
            else:
                month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)
            
            if month_end > today:
                month_end = today
            
            month_orders = PurchaseOrder.objects.filter(
                purchase_date__range=[month_start, month_end]
            )
            
            month_total = month_orders.aggregate(
                total=Sum('total_cost')
            )['total'] or Decimal('0')
            
            month_count = month_orders.count()
            month_avg = month_total / month_count if month_count > 0 else Decimal('0')
            
            periods.append({
                'month': month_start.strftime('%B %Y'),
                'purchase_amount': month_total,
                'orders': month_count,
                'avg_order_value': month_avg,
                'growth_rate': Decimal('0'),  # Will calculate below
            })
        
        # Calculate growth rates CORRECTLY
        for i in range(len(periods)):
            if i > 0:
                prev_total = periods[i-1]['purchase_amount']
                current_total = periods[i]['purchase_amount']
                
                if prev_total > Decimal('0'):
                    growth = ((current_total - prev_total) / prev_total) * 100
                    periods[i]['growth_rate'] = growth
                else:
                    # If previous was 0 and current has value, that's 100% growth
                    if current_total > Decimal('0'):
                        periods[i]['growth_rate'] = Decimal('100')
                    else:
                        periods[i]['growth_rate'] = Decimal('0')
        
        return periods
    
    return []


def get_po_status_data_consistent(start_date, end_date):
    """
    FIXED purchase order status data - ensures consistency
    """
    # Get ALL orders in date range
    orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier')
    
    # Count by status - FIXED: Handle case-insensitive
    status_counts = {
        'pending': orders.filter(
            Q(status='pending') | Q(status='PENDING')
        ).count(),
        'in_progress': orders.filter(
            Q(status='in_progress') | Q(status='IN_PROGRESS') | Q(status='in progress')
        ).count(),
        'completed': orders.filter(
            Q(status='completed') | Q(status='COMPLETED')
        ).count(),
        'cancelled': orders.filter(
            Q(status='cancelled') | Q(status='CANCELLED')
        ).count(),
    }
    
    # Detailed order list - MUST match counts
    order_details = []
    total_amount = Decimal('0')
    total_items = 0
    
    for order in orders.order_by('-purchase_date'):
        # FIXED: Calculate payment status
        payment_status = 'pending'
        if order.status.lower() == 'completed':
            payment_status = 'paid'
        elif order.expected_date and order.expected_date < timezone.now().date():
            payment_status = 'overdue'
        
        order_details.append({
            'po_number': f"PO-{order.id}",
            'supplier': order.supplier.name if order.supplier else 'Unknown',
            'order_date': order.purchase_date,
            'due_date': order.expected_date,
            'amount': order.total_cost or Decimal('0'),
            'items': order.items.count(),
            'status': order.status.lower(),  # Normalize to lowercase
            'payment_status': payment_status,
        })
        
        # Calculate totals
        total_amount += order.total_cost or Decimal('0')
        total_items += order.items.count()
    
    return {
        'status_counts': status_counts,
        'order_details': order_details,
        'total_amount': total_amount,
        'total_items': total_items,
        'total_orders': len(order_details),
    }


def get_item_analysis_data_accurate(start_date, end_date):
    """
    Get accurate item-wise purchase analysis
    """
    items = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).select_related('product', 'product__category').values(
        'product__id',
        'product__name',
        'product__sku',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost'),
    ).order_by('-total_cost')[:15]
    
    # Add stock information
    for item in items:
        try:
            product = Product.objects.get(id=item['product__id'])
            
            # Get actual stock from inventory
            inventory = product.inventories.first()
            if inventory:
                current_stock = inventory.quantity_in_stock
                reorder_level = inventory.reorder_level
            else:
                current_stock = 0
                reorder_level = 10
            
            # Verify with batch stock
            batch_stock = InventoryBatch.objects.filter(
                product=product
            ).aggregate(
                total=Sum('remaining_quantity')
            )['total'] or 0
            
            # Use whichever is more accurate
            actual_stock = batch_stock if batch_stock > current_stock else current_stock
            
            item['current_stock'] = actual_stock
            item['reorder_level'] = reorder_level
            
            # Determine stock status
            if actual_stock == 0:
                item['stock_status'] = 'out_of_stock'
            elif actual_stock <= reorder_level:
                item['stock_status'] = 'low_stock'
            else:
                item['stock_status'] = 'in_stock'
                
        except Product.DoesNotExist:
            item['current_stock'] = 0
            item['reorder_level'] = 10
            item['stock_status'] = 'unknown'
    
    return items


def get_expiry_data_with_correct_status(start_date, end_date):
    """
    Get expiry tracking data with CORRECT status logic
    """
    today = timezone.now().date()
    expiry_data = []
    
    # Get all batches, not just recent ones
    batches = InventoryBatch.objects.filter(
        expiry_date__isnull=False,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')
    
    for batch in batches:
        if batch.expiry_date:
            days_remaining = (batch.expiry_date - today).days
            
            # CORRECT STATUS DETERMINATION
            if days_remaining < 0:
                status = 'expired'
                days_text = f"Expired {abs(days_remaining)} days ago"
            elif days_remaining == 0:
                status = 'expiring_today'  # CRITICAL, not safe!
                days_text = '0 days (Today)'
            elif days_remaining <= 7:
                status = 'critical'
                days_text = f'{days_remaining} days'
            elif days_remaining <= 30:
                status = 'expiring_soon'
                days_text = f'{days_remaining} days'
            elif days_remaining <= 60:
                status = 'monitor'
                days_text = f'{days_remaining} days'
            else:
                status = 'safe'
                days_text = f'{days_remaining} days'
        else:
            status = 'no_expiry'
            days_text = 'N/A'
        
        total_value = batch.remaining_quantity * batch.unit_cost
        
        expiry_data.append({
            'item_code': batch.product.sku if batch.product and batch.product.sku else 'N/A',
            'item_name': batch.product.name if batch.product else 'Unknown',
            'batch_number': f"BATCH-{batch.id}",
            'current_stock': batch.remaining_quantity,
            'purchase_date': batch.created_at.date() if batch.created_at else today,
            'expiry_date': batch.expiry_date,
            'days_remaining': days_text,
            'unit_cost': batch.unit_cost,
            'total_value': total_value,
            'status': status,
            'days_numeric': days_remaining if batch.expiry_date else 9999,
        })
    
    return expiry_data


def calculate_expiry_stats_correct(expiry_data):
    """
    Calculate correct expiry statistics
    """
    stats = {
        'expiring_soon_count': 0,
        'expired_count': 0,
        'good_stock_count': 0,
        'total_value_at_risk': Decimal('0'),
        'potential_loss': Decimal('0'),
    }
    
    for item in expiry_data:
        status = item['status']
        value = item['total_value']
        
        if status in ['expired', 'expiring_today', 'critical', 'expiring_soon']:
            stats['expiring_soon_count'] += 1
            
            if status == 'expired':
                stats['expired_count'] += 1
                risk_multiplier = Decimal('1.0')  # 100% at risk
            elif status == 'expiring_today':
                risk_multiplier = Decimal('0.9')  # 90% at risk
            elif status == 'critical':
                risk_multiplier = Decimal('0.7')  # 70% at risk
            elif status == 'expiring_soon':
                risk_multiplier = Decimal('0.5')  # 50% at risk
            else:
                risk_multiplier = Decimal('0.3')  # 30% at risk
            
            stats['total_value_at_risk'] += value
            stats['potential_loss'] += value * risk_multiplier
        elif status == 'safe' or status == 'monitor':
            stats['good_stock_count'] += 1
    
    return stats



def export_purchase_csv(request):
    """Export comprehensive purchase report as CSV"""
    # Get filter parameters from request
    report_type = request.GET.get('report_type', 'monthly')
    period = request.GET.get('period', '')
    store_id = request.GET.get('store')
    
    # Set up response
    response = HttpResponse(content_type='text/csv')
    filename = f"purchase_report_{period or timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Get date range based on report_type
    today = timezone.now().date()
    if report_type == 'monthly':
        if period:
            year, month = map(int, period.split('-'))
        else:
            year, month = today.year, today.month
        start_date = date(year, month, 1).date()
        if month == 12:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1).date() - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if period:
            year, quarter = period.split('-Q')
            year = int(year)
            quarter = int(quarter)
        else:
            year = today.year
            quarter = (today.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1).date()
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1).date() - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    else:  # yearly or custom
        if period and len(period) == 4:
            year = int(period)
        else:
            year = today.year
        start_date = date(year, 1, 1).date()
        end_date = date(year, 12, 31).date()
        period_label = str(year)
    
    # Write header information
    writer.writerow(['PURCHASE REPORT EXPORT'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Period:', period_label])
    writer.writerow(['Date Range:', f"{start_date} to {end_date}"])
    writer.writerow(['Generated By:', request.user.get_full_name() or request.user.username])
    writer.writerow([])
    
    # Get purchase orders for the period
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    )
    
    # Calculate summary data
    total_purchases = purchase_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # 1. SUMMARY SECTION
    writer.writerow(['SUMMARY SECTION'])
    writer.writerow([])
    
    writer.writerow(['Total Purchases:', f"UGX {total_purchases:,.0f}"])
    writer.writerow(['Total Orders:', f"{total_orders}"])
    writer.writerow(['Average Order Value:', f"UGX {avg_order_value:,.0f}"])
    writer.writerow(['Date Range:', f"{start_date} to {end_date}"])
    writer.writerow(['Days in Period:', f"{(end_date - start_date).days + 1}"])
    writer.writerow([])
    
    # 2. SUPPLIER ANALYSIS
    writer.writerow(['SUPPLIER PURCHASE ANALYSIS'])
    writer.writerow(['Supplier Name', 'Supplier Code', 'Total Purchases (UGX)', 'Orders', 'Avg Order (UGX)', 
                     'Last Order', 'Payment Terms', 'Performance Rating'])
    
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__name', 
        'supplier__supplier_code',
        'supplier__payment_terms'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id'),
        avg_order=Avg('total_cost'),
        last_order=Max('purchase_date')
    ).order_by('-total_purchases')
    
    for supplier in supplier_summary:
        # Calculate performance (simplified)
        performance = 'Good'
        if supplier['orders'] > 10 and supplier['avg_order'] > 100000:
            performance = 'Excellent'
        elif supplier['orders'] < 3:
            performance = 'New'
        
        writer.writerow([
            supplier['supplier__name'] or 'N/A',
            supplier['supplier__supplier_code'] or 'N/A',
            f"{supplier['total_purchases']:,.0f}" if supplier['total_purchases'] else '0',
            supplier['orders'],
            f"{supplier['avg_order']:,.0f}" if supplier['avg_order'] else '0',
            supplier['last_order'].strftime('%Y-%m-%d') if supplier['last_order'] else 'N/A',
            supplier['supplier__payment_terms'] or 'Net 30',
            performance
        ])
    
    writer.writerow([])
    
    # 3. PURCHASE TREND ANALYSIS
    writer.writerow(['PURCHASE TREND ANALYSIS'])
    writer.writerow(['Period', 'Purchase Amount (UGX)', 'Orders', 'Avg Order Value (UGX)', 'Growth Rate'])
    
    # Get trend data for last 6 periods
    trend_data = []
    if report_type == 'monthly':
        # Get last 6 months including current
        for i in range(6, 0, -1):
            month_date = end_date - timedelta(days=30*i)
            month_start = date(month_date.year, month_date.month, 1).date()
            if month_start.month == 12:
                month_end = date(month_start.year + 1, 1, 1).date() - timedelta(days=1)
            else:
                month_end = date(month_start.year, month_start.month + 1, 1).date() - timedelta(days=1)
            
            month_orders = PurchaseOrder.objects.filter(
                purchase_date__range=[month_start, month_end]
            )
            month_total = month_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
            month_count = month_orders.count()
            month_avg = month_total / month_count if month_count > 0 else Decimal('0')
            
            # Calculate growth rate
            growth = '0%'
            if i < 6:
                prev_total = trend_data[-1][1]
                if prev_total > 0:
                    growth_rate = ((month_total - prev_total) / prev_total) * 100
                    growth = f"{growth_rate:+.1f}%"
            
            writer.writerow([
                month_start.strftime('%B %Y'),
                f"{month_total:,.0f}",
                month_count,
                f"{month_avg:,.0f}",
                growth
            ])
            trend_data.append((month_start.strftime('%B %Y'), month_total))
    
    writer.writerow([])
    
    # 4. PURCHASE ORDER STATUS
    writer.writerow(['PURCHASE ORDER STATUS'])
    writer.writerow(['PO Number', 'Supplier', 'Order Date', 'Due Date', 'Amount (UGX)', 
                     'Items', 'Status', 'Payment Status'])
    
    po_details = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier').order_by('-purchase_date')[:50]
    
    for order in po_details:
        payment_status = 'Pending'
        if order.status == 'completed':
            payment_status = 'Paid'
        elif order.expected_date and order.expected_date < today:
            payment_status = 'Overdue'
        
        writer.writerow([
            f"PO-{order.id}",
            order.supplier.name if order.supplier else 'N/A',
            order.purchase_date.strftime('%Y-%m-%d') if order.purchase_date else '',
            order.expected_date.strftime('%Y-%m-%d') if order.expected_date else 'N/A',
            f"{order.total_cost:,.0f}",
            order.items.count(),
            order.status.capitalize(),
            payment_status
        ])
    
    writer.writerow([])
    
    # 5. ITEM-WISE PURCHASE ANALYSIS
    writer.writerow(['ITEM-WISE PURCHASE ANALYSIS'])
    writer.writerow(['Item Code', 'Item Name', 'Category', 'Quantity Purchased', 
                     'Unit Cost (UGX)', 'Total Cost (UGX)', 'Avg Unit Cost (UGX)'])
    
    item_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__sku',
        'product__name',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost')
    ).order_by('-total_cost')[:20]
    
    for item in item_analysis:
        writer.writerow([
            item['product__sku'] or 'N/A',
            item['product__name'],
            item['product__category__name'] or 'Uncategorized',
            item['quantity_purchased'],
            f"{item['avg_unit_cost']:,.0f}" if item['avg_unit_cost'] else '0',
            f"{item['total_cost']:,.0f}" if item['total_cost'] else '0',
            f"{item['avg_unit_cost']:,.0f}" if item['avg_unit_cost'] else '0'
        ])
    
    writer.writerow([])
    
    # 6. EXPIRY TRACKING
    writer.writerow(['EXPIRY TRACKING REPORT'])
    writer.writerow(['Item Code', 'Item Name', 'Batch Number', 'Quantity', 
                     'Purchase Date', 'Expiry Date', 'Days Remaining', 
                     'Unit Cost (UGX)', 'Total Value (UGX)', 'Status'])
    
    # Get batches that will expire within next 90 days or already expired
    ninety_days_from_now = today + timedelta(days=90)
    expiry_batches = InventoryBatch.objects.filter(
        expiry_date__lte=ninety_days_from_now,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')[:50]
    
    for batch in expiry_batches:
        days_remaining = (batch.expiry_date - today).days if batch.expiry_date else None
        
        if days_remaining is None:
            status = 'No Expiry'
        elif days_remaining < 0:
            status = 'Expired'
        elif days_remaining <= 30:
            status = 'Expiring Soon'
        elif days_remaining <= 60:
            status = 'Monitor'
        else:
            status = 'Safe'
        
        writer.writerow([
            batch.product.sku if batch.product.sku else 'N/A',
            batch.product.name,
            f"BATCH-{batch.id}",
            batch.remaining_quantity,
            batch.created_at.strftime('%Y-%m-%d') if batch.created_at else '',
            batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A',
            f"{days_remaining}" if days_remaining is not None else 'N/A',
            f"{batch.unit_cost:,.0f}",
            f"{batch.remaining_quantity * batch.unit_cost:,.0f}",
            status
        ])
    
    writer.writerow([])
    
    # 7. STORE PERFORMANCE
    writer.writerow(['STORE PURCHASE PERFORMANCE'])
    writer.writerow(['Store Name', 'Total Purchases (UGX)', '% of Total', 
                     'Orders', 'Avg Order (UGX)', 'Performance'])
    
    store_performance = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'store__name'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id')
    ).order_by('-total_purchases')
    
    for store in store_performance:
        percent = (store['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_order = store['total_purchases'] / store['orders'] if store['orders'] > 0 else Decimal('0')
        
        # Determine performance
        if percent > 40:
            performance = 'Top Performer'
        elif percent > 20:
            performance = 'Good'
        elif percent > 5:
            performance = 'Average'
        else:
            performance = 'Low Volume'
        
        writer.writerow([
            store['store__name'] or 'Unknown',
            f"{store['total_purchases']:,.0f}",
            f"{percent:.1f}%",
            store['orders'],
            f"{avg_order:,.0f}",
            performance
        ])
    
    writer.writerow([])
    
    # 8. CATEGORY ANALYSIS
    writer.writerow(['CATEGORY PURCHASE ANALYSIS'])
    writer.writerow(['Category', 'Total Purchases (UGX)', '% of Total', 
                     'Items Purchased', 'Avg Item Cost (UGX)', 'Trend'])
    
    category_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__category__name'
    ).annotate(
        total_cost=Sum(F('quantity') * F('unit_cost')),
        total_items=Sum('quantity'),
        avg_item_cost=Avg('unit_cost')
    ).order_by('-total_cost')
    
    for category in category_analysis:
        percent = (category['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
        writer.writerow([
            category['product__category__name'] or 'Uncategorized',
            f"{category['total_cost']:,.0f}",
            f"{percent:.1f}%",
            category['total_items'],
            f"{category['avg_item_cost']:,.0f}",
            'Stable'  # Simplified trend
        ])
    
    return response


def export_purchase_pdf(request):
    """Export comprehensive purchase report as PDF with black and white design"""
    try:
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        store_id = request.GET.get('store')
        supplier_id = request.GET.get('supplier')
        period = request.GET.get('period', 'Custom Range')
        
        # Set up response - DIRECT TO RESPONSE (NO BUFFER)
        response = HttpResponse(content_type='application/pdf')
        filename = f"purchase_report_{period.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Parse dates
        today = timezone.now().date()
        if date_from and date_to:
            try:
                start_date = date.fromisoformat(date_from)
                end_date = date.fromisoformat(date_to)
            except:
                start_date = today.replace(day=1)
                end_date = today
        else:
            start_date = today.replace(day=1)
            end_date = today
        
        # Create PDF directly with response
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles - Black and White only
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,  # Center
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        period_style = ParagraphStyle(
            'PeriodStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,  # Center
            spaceAfter=5,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold',
            alignment=0,  # Left
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header Section - Exactly as requested
        elements.append(Paragraph("PURCHASE REPORT", title_style))
        
        # Period line
        period_text = f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}"
        elements.append(Paragraph(period_text, period_style))
        
        # Store line if store filter applied
        if store_id:
            try:
                store = StoreLocation.objects.get(id=store_id)
                elements.append(Paragraph(f"Store: {store.name}", period_style))
            except:
                pass
        
        # Generated line
        generated_text = f"Generated: {timezone.now().strftime('%B %d, %Y')}"
        elements.append(Paragraph(generated_text, period_style))
        elements.append(Spacer(1, 15))
        
        # Add horizontal line
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Get purchase data - FIXED: removed recorded_by from select_related
        purchase_qs = PurchaseOrder.objects.filter(
            purchase_date__range=[start_date, end_date]
        ).select_related('supplier', 'store')
        
        if store_id:
            purchase_qs = purchase_qs.filter(store_id=store_id)
        
        if supplier_id:
            purchase_qs = purchase_qs.filter(supplier_id=supplier_id)
        
        # Check if there's data
        if not purchase_qs.exists():
            elements.append(Paragraph("No purchase data found for the selected period.", normal_style))
            doc.build(elements)
            return response
        
        total_purchases = purchase_qs.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
        total_orders = purchase_qs.count()
        days_in_period = max((end_date - start_date).days + 1, 1)
        avg_daily_purchases = total_purchases / Decimal(days_in_period) if total_purchases else Decimal('0')
        
        # Calculate total items purchased
        total_items_purchased = 0
        for po in purchase_qs:
            total_items_purchased += po.total_items
        
        # Get unique suppliers count
        suppliers_count = purchase_qs.values('supplier').distinct().count()
        
        # Get unique products count
        products_count = PurchaseOrderItem.objects.filter(
            order__in=purchase_qs
        ).values('product').distinct().count()
        
        avg_cost_per_item = total_purchases / Decimal(total_items_purchased) if total_items_purchased > 0 else Decimal('0')
        
        # 1. SUMMARY SECTION
        elements.append(Paragraph("SUMMARY SECTION", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Purchase Orders', f"{total_orders:,}"],
            ['Total Cost', f"UGX {total_purchases:,.0f}"],
            ['Total Items Purchased', f"{total_items_purchased:,} pcs (base units)"],
            ['Average Order Value', f"UGX {(total_purchases/total_orders):,.0f}" if total_orders > 0 else 'UGX 0'],
            ['Average Cost Per Item', f"UGX {avg_cost_per_item:,.0f}"],
            ['Suppliers Used', f"{suppliers_count}"],
            ['Products Purchased', f"{products_count}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # 2. PURCHASES BY STATUS
        elements.append(Paragraph("PURCHASES BY STATUS", heading_style))
        elements.append(Spacer(1, 5))
        
        status_counts = purchase_qs.values('status').annotate(
            count=Count('id'),
            total=Sum('total_cost')
        ).order_by('-total')
        
        status_data = [['Status', 'Orders', 'Total Cost', '% of Total']]
        
        for status in status_counts:
            status_name = status['status'].capitalize() if status['status'] else 'Unknown'
            percent = (status['total'] / total_purchases * 100) if total_purchases > 0 else 0
            status_data.append([
                status_name,
                str(status['count']),
                f"UGX {status['total']:,.0f}",
                f"{percent:.1f}%"
            ])
        
        # Add total row
        status_data.append(['TOTAL', str(total_orders), f"UGX {total_purchases:,.0f}", '100%'])
        
        status_table = Table(status_data, colWidths=[150, 100, 150, 100])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # 3. PURCHASES BY SUPPLIER
        elements.append(Paragraph("PURCHASES BY SUPPLIER", heading_style))
        elements.append(Spacer(1, 5))
        
        supplier_summary = purchase_qs.values(
            'supplier__name'
        ).annotate(
            total_purchases=Sum('total_cost'),
            orders=Count('id')
        ).order_by('-total_purchases')
        
        supplier_data = [['Supplier', 'Orders', 'Total Cost', '% of Total', 'Items Purchased']]
        
        for supplier in supplier_summary:
            if supplier['supplier__name']:
                percent = (supplier['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
                
                supplier_items = PurchaseOrderItem.objects.filter(
                    order__in=purchase_qs.filter(supplier__name=supplier['supplier__name'])
                ).aggregate(total_items=Sum('base_quantity'))['total_items'] or 0
                
                supplier_data.append([
                    supplier['supplier__name'][:25] if len(supplier['supplier__name']) > 25 else supplier['supplier__name'],
                    str(supplier['orders']),
                    f"UGX {supplier['total_purchases']:,.0f}",
                    f"{percent:.1f}%",
                    str(supplier_items)
                ])
        
        # Add "Others" row if needed - combine suppliers with small percentages
        if len(supplier_data) > 6:  # Keep top 5 plus header
            # Calculate total for others
            others_orders = 0
            others_purchases = Decimal('0')
            others_items = 0
            other_suppliers = supplier_data[6:]  # Keep top 5 suppliers
            for supplier in other_suppliers:
                others_orders += int(supplier[1])
                others_purchases += Decimal(supplier[2].replace('UGX ', '').replace(',', ''))
                others_items += int(supplier[4])
            
            # Keep only top 5
            supplier_data = supplier_data[:6]
            # Add others row
            percent = (others_purchases / total_purchases * 100) if total_purchases > 0 else 0
            supplier_data.append([
                'Others',
                str(others_orders),
                f"UGX {others_purchases:,.0f}",
                f"{percent:.1f}%",
                str(others_items)
            ])
        
        # Add total row
        supplier_data.append(['TOTAL', str(total_orders), f"UGX {total_purchases:,.0f}", '100%', str(total_items_purchased)])
        
        supplier_table = Table(supplier_data, colWidths=[150, 70, 120, 70, 100])
        supplier_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(supplier_table)
        elements.append(Spacer(1, 20))
        
        # 4. DETAILED PURCHASE ORDERS
        elements.append(Paragraph("DETAILED PURCHASE ORDERS", heading_style))
        elements.append(Spacer(1, 5))
        
        recent_pos = purchase_qs.select_related(
            'supplier', 'store'
        ).order_by('-purchase_date', '-id')[:15]  # Limit to 15 for PDF readability
        
        if recent_pos.exists():
            po_data = [
                ['PO No.', 'Date', 'Supplier', 'Items', 'Total Cost', 'Status', 'Expected Date', 'Recorded By']
            ]
            
            for po in recent_pos:
                # Truncate long text for PDF
                supplier_name = po.supplier.name[:18] + '...' if po.supplier and len(po.supplier.name) > 18 else (po.supplier.name if po.supplier else 'N/A')
                status_text = po.get_status_display()[:10] if hasattr(po, 'get_status_display') else (po.status.capitalize() if po.status else 'N/A')
                
                po_data.append([
                    f"PO-{po.id}",
                    po.purchase_date.strftime('%Y-%m-%d') if po.purchase_date else '',
                    supplier_name,
                    str(po.total_items),
                    f"UGX {po.total_cost:,.0f}",
                    status_text,
                    po.expected_date.strftime('%Y-%m-%d') if po.expected_date else 'N/A',
                    str(po.recorded_by) if po.recorded_by else 'N/A'
                ])
            
            po_table = Table(po_data, colWidths=[55, 60, 90, 40, 85, 65, 65, 60])
            po_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (3, 1), (4, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ]))
            elements.append(po_table)
            
            if purchase_qs.count() > 15:
                elements.append(Spacer(1, 5))
                elements.append(Paragraph(f"* Showing 15 of {purchase_qs.count()} purchase orders", normal_style))
        else:
            elements.append(Paragraph("No purchase order details available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # 5. PURCHASE ITEMS DETAIL
        elements.append(Paragraph("PURCHASE ITEMS DETAIL", heading_style))
        elements.append(Spacer(1, 5))
        
        # Get top 25 items by value
        items = PurchaseOrderItem.objects.filter(
            order__in=purchase_qs
        ).select_related(
            'product', 'unit', 'order'
        ).order_by('-order__purchase_date', '-id')[:25]
        
        if items.exists():
            items_data = [
                ['PO No.', 'Product', 'SKU', 'Unit', 'Qty', 'Base Qty', 'Unit Cost', 'Line Total']
            ]
            
            total_line_value = Decimal('0')
            for item in items:
                line_total = item.base_quantity * item.unit_cost
                total_line_value += line_total
                
                # Truncate long product names
                product_name = item.product.name[:15] + '...' if len(item.product.name) > 15 else item.product.name
                
                items_data.append([
                    f"PO-{item.order.id}",
                    product_name,
                    item.product.sku or 'N/A',
                    item.unit.abbreviation if item.unit and hasattr(item.unit, 'abbreviation') else 'unit',
                    str(item.quantity),
                    str(item.base_quantity),
                    f"UGX {item.unit_cost:,.0f}",
                    f"UGX {line_total:,.0f}"
                ])
            
            # Add total row
            items_data.append([
                'TOTAL', '', '', '', '', f"{total_items_purchased}", f"UGX {total_purchases:,.0f}", ''
            ])
            
            items_table = Table(items_data, colWidths=[55, 80, 45, 35, 35, 45, 65, 75])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ]))
            elements.append(items_table)
        else:
            elements.append(Paragraph("No purchase item details available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # 6. TOP PRODUCTS PURCHASED
        elements.append(Paragraph("TOP PRODUCTS PURCHASED", heading_style))
        elements.append(Spacer(1, 5))
        
        top_products = PurchaseOrderItem.objects.filter(
            order__in=purchase_qs
        ).values(
            'product__name', 'product__sku'
        ).annotate(
            total_qty=Sum('base_quantity'),
            total_cost=Sum(F('base_quantity') * F('unit_cost')),
            orders=Count('order', distinct=True)
        ).order_by('-total_cost')[:10]
        
        if top_products:
            products_data = [['Rank', 'Product', 'SKU', 'Total Qty', 'Total Cost', 'Avg Unit Cost', 'Orders']]
            
            for i, product in enumerate(top_products, 1):
                if product['product__name']:
                    avg_unit = product['total_cost'] / product['total_qty'] if product['total_qty'] and product['total_qty'] > 0 else Decimal('0')
                    product_name = product['product__name'][:20] + '...' if len(product['product__name']) > 20 else product['product__name']
                    
                    products_data.append([
                        str(i),
                        product_name,
                        product['product__sku'] or 'N/A',
                        str(product['total_qty']),
                        f"UGX {product['total_cost']:,.0f}",
                        f"UGX {avg_unit:,.0f}",
                        str(product['orders'])
                    ])
            
            products_table = Table(products_data, colWidths=[35, 110, 60, 55, 90, 75, 45])
            products_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(products_table)
        else:
            elements.append(Paragraph("No top products data available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # 7. PURCHASES BY STORE
        if not store_id:
            elements.append(Paragraph("PURCHASES BY STORE", heading_style))
            elements.append(Spacer(1, 5))
            
            store_summary = purchase_qs.values(
                'store__name'
            ).annotate(
                orders=Count('id'),
                total_cost=Sum('total_cost')
            ).order_by('-total_cost')
            
            if store_summary:
                store_data = [['Store', 'Orders', 'Total Cost', '% of Total', 'Items Received']]
                
                for store in store_summary:
                    if store['store__name']:
                        percent = (store['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
                        
                        store_items = PurchaseOrderItem.objects.filter(
                            order__in=purchase_qs.filter(store__name=store['store__name'])
                        ).aggregate(total_items=Sum('base_quantity'))['total_items'] or 0
                        
                        store_data.append([
                            store['store__name'],
                            str(store['orders']),
                            f"UGX {store['total_cost']:,.0f}",
                            f"{percent:.1f}%",
                            str(store_items)
                        ])
                
                # Add total row
                store_data.append(['TOTAL', str(total_orders), f"UGX {total_purchases:,.0f}", '100%', str(total_items_purchased)])
                
                store_table = Table(store_data, colWidths=[120, 70, 120, 70, 100])
                store_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
                ]))
                elements.append(store_table)
            else:
                elements.append(Paragraph("No store data available.", normal_style))
            elements.append(Spacer(1, 20))
        
        # 8. TOP SUPPLIERS BY SPEND
        elements.append(Paragraph("TOP SUPPLIERS BY SPEND", heading_style))
        elements.append(Spacer(1, 5))
        
        top_suppliers = purchase_qs.values(
            'supplier__name'
        ).annotate(
            total_spent=Sum('total_cost'),
            orders=Count('id')
        ).order_by('-total_spent')[:10]
        
        if top_suppliers:
            supplier_spend_data = [['Rank', 'Supplier', 'Total Spent', 'Orders', 'Avg Order Value']]
            
            for i, supplier in enumerate(top_suppliers, 1):
                if supplier['supplier__name']:
                    avg_order = supplier['total_spent'] / supplier['orders'] if supplier['orders'] > 0 else Decimal('0')
                    supplier_name = supplier['supplier__name'][:25] if len(supplier['supplier__name']) > 25 else supplier['supplier__name']
                    
                    supplier_spend_data.append([
                        str(i),
                        supplier_name,
                        f"UGX {supplier['total_spent']:,.0f}",
                        str(supplier['orders']),
                        f"UGX {avg_order:,.0f}"
                    ])
            
            supplier_spend_table = Table(supplier_spend_data, colWidths=[35, 150, 120, 70, 120])
            supplier_spend_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (2, 1), (4, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(supplier_spend_table)
        else:
            elements.append(Paragraph("No supplier data available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # 9. STAFF PURCHASING ACTIVITY
        elements.append(Paragraph("STAFF PURCHASING ACTIVITY", heading_style))
        elements.append(Spacer(1, 5))
        
        staff_activity = purchase_qs.values(
            'recorded_by'
        ).annotate(
            orders=Count('id'),
            total_spent=Sum('total_cost')
        ).order_by('-orders')
        
        if staff_activity:
            staff_data = [['Staff', 'Orders', 'Total Spent', 'Avg Order Value', 'Suppliers Used']]
            
            for staff in staff_activity:
                if staff['recorded_by']:
                    avg_order = staff['total_spent'] / staff['orders'] if staff['orders'] > 0 else Decimal('0')
                    
                    # Count unique suppliers for this staff
                    suppliers_used = PurchaseOrder.objects.filter(
                        recorded_by=staff['recorded_by'],
                        purchase_date__range=[start_date, end_date]
                    ).values('supplier').distinct().count()
                    
                    staff_data.append([
                        staff['recorded_by'],
                        str(staff['orders']),
                        f"UGX {staff['total_spent']:,.0f}",
                        f"UGX {avg_order:,.0f}",
                        str(suppliers_used)
                    ])
            
            # Add total row
            staff_data.append(['TOTAL', str(total_orders), f"UGX {total_purchases:,.0f}", f"UGX {(total_purchases/total_orders):,.0f}", str(suppliers_count)])
            
            staff_table = Table(staff_data, colWidths=[120, 70, 120, 120, 100])
            staff_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ]))
            elements.append(staff_table)
        else:
            elements.append(Paragraph("No staff activity data available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # Build PDF directly to response
        doc.build(elements)
        return response
        
    except Exception as e:
        # Log the error
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # Return an error response
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response



def export_purchase_excel(request):
    """Export comprehensive purchase report as Excel"""
    # Get filter parameters
    report_type = request.GET.get('report_type', 'monthly')
    period = request.GET.get('period', '')
    store_id = request.GET.get('store')
    
    # Get date range based on report_type
    today = timezone.now().date()
    if report_type == 'monthly':
        if period:
            year, month = map(int, period.split('-'))
        else:
            year, month = today.year, today.month
        start_date = date(year, month, 1).date()
        if month == 12:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1).date() - timedelta(days=1)
        period_label = start_date.strftime('%B %Y')
        
    elif report_type == 'quarterly':
        if period:
            year, quarter = period.split('-Q')
            year = int(year)
            quarter = int(quarter)
        else:
            year = today.year
            quarter = (today.month - 1) // 3 + 1
        start_month = (quarter - 1) * 3 + 1
        start_date = date(year, start_month, 1).date()
        if start_month + 2 <= 12:
            end_date = date(year, start_month + 3, 1).date() - timedelta(days=1)
        else:
            end_date = date(year + 1, 1, 1).date() - timedelta(days=1)
        period_label = f"Q{quarter} {year}"
        
    else:  # yearly or custom
        if period and len(period) == 4:
            year = int(period)
        else:
            year = today.year
        start_date = date(year, 1, 1).date()
        end_date = date(year, 12, 31).date()
        period_label = str(year)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    subheader_font = Font(bold=True, color="000000", size=11)
    subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    
    total_font = Font(bold=True, color="000000", size=10)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    currency_format = '"UGX "#,##0'
    percent_format = '0.0"%'
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Get summary data
    purchase_orders = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    )
    
    total_purchases = purchase_orders.aggregate(total=Sum('total_cost'))['total'] or Decimal('0')
    total_orders = purchase_orders.count()
    avg_order_value = total_purchases / total_orders if total_orders > 0 else Decimal('0')
    
    # 1. SUMMARY SHEET
    ws_summary = wb.create_sheet(title="Summary")
    
    # Header
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = f"PURCHASE REPORT - {period_label}"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = header_alignment
    
    # Report info
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = "Period:"
    ws_summary['B4'] = period_label
    ws_summary['A5'] = "Generated By:"
    ws_summary['B5'] = request.user.get_full_name() or request.user.username
    ws_summary['A6'] = "Date Range:"
    ws_summary['B6'] = f"{start_date} to {end_date}"
    ws_summary['A7'] = "Report Type:"
    ws_summary['B7'] = report_type.capitalize()
    
    # Key Metrics
    ws_summary.merge_cells('A9:F9')
    ws_summary['A9'] = "KEY METRICS"
    ws_summary['A9'].font = subheader_font
    ws_summary['A9'].fill = subheader_fill
    ws_summary['A9'].alignment = Alignment(horizontal="center")
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Purchases', f"UGX {total_purchases:,.0f}"],
        ['Total Orders', total_orders],
        ['Average Order Value', f"UGX {avg_order_value:,.0f}"],
        ['Days in Period', (end_date - start_date).days + 1],
        ['Average Daily Purchases', f"UGX {total_purchases/((end_date - start_date).days + 1):,.0f}" if total_purchases > 0 else "UGX 0"]
    ]
    
    for i, row in enumerate(summary_data, start=10):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 10:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            elif j == 2 and 'UGX' in str(value):
                cell.number_format = '"UGX "#,##0'
    
    # 2. SUPPLIER ANALYSIS SHEET
    ws_suppliers = wb.create_sheet(title="Supplier Analysis")
    ws_suppliers['A1'] = "SUPPLIER PURCHASE ANALYSIS"
    ws_suppliers['A1'].font = header_font
    ws_suppliers['A1'].fill = header_fill
    ws_suppliers.merge_cells('A1:H1')
    ws_suppliers['A1'].alignment = header_alignment
    
    supplier_headers = ['Supplier Name', 'Supplier Code', 'Total Purchases (UGX)', 'Orders', 
                       'Avg Order (UGX)', 'Last Order Date', 'Payment Terms', 'Performance Rating']
    for col, header in enumerate(supplier_headers, start=1):
        cell = ws_suppliers.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    
    supplier_summary = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'supplier__name', 
        'supplier__supplier_code',
        'supplier__payment_terms'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id'),
        avg_order=Avg('total_cost'),
        last_order=Max('purchase_date')
    ).order_by('-total_purchases')
    
    row = 4
    for supplier in supplier_summary:
        # Calculate performance
        performance = 'Good'
        if supplier['orders'] > 10 and supplier['avg_order'] > 100000:
            performance = 'Excellent'
        elif supplier['orders'] >= 5 and supplier['avg_order'] > 50000:
            performance = 'Very Good'
        elif supplier['orders'] < 3:
            performance = 'New'
        elif supplier['orders'] > 0 and supplier['avg_order'] < 10000:
            performance = 'Needs Review'
        
        ws_suppliers.cell(row=row, column=1, value=supplier['supplier__name'] or 'N/A')
        ws_suppliers.cell(row=row, column=2, value=supplier['supplier__supplier_code'] or 'N/A')
        ws_suppliers.cell(row=row, column=3, value=float(supplier['total_purchases'] or 0))
        ws_suppliers.cell(row=row, column=4, value=supplier['orders'])
        ws_suppliers.cell(row=row, column=5, value=float(supplier['avg_order'] or 0))
        ws_suppliers.cell(row=row, column=6, value=supplier['last_order'].strftime('%Y-%m-%d') if supplier['last_order'] else 'N/A')
        ws_suppliers.cell(row=row, column=7, value=supplier['supplier__payment_terms'] or 'Net 30')
        ws_suppliers.cell(row=row, column=8, value=performance)
        
        # Format currency cells
        ws_suppliers.cell(row=row, column=3).number_format = currency_format
        ws_suppliers.cell(row=row, column=5).number_format = currency_format
        
        row += 1
    
    # Add totals row
    ws_suppliers.cell(row=row, column=1, value="TOTAL").font = total_font
    ws_suppliers.cell(row=row, column=2, value="")
    ws_suppliers.cell(row=row, column=3, value=float(total_purchases)).font = total_font
    ws_suppliers.cell(row=row, column=4, value=total_orders).font = total_font
    ws_suppliers.cell(row=row, column=5, value=float(avg_order_value)).font = total_font
    ws_suppliers.cell(row=row, column=3).number_format = currency_format
    ws_suppliers.cell(row=row, column=5).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 9):
            ws_suppliers.cell(row=r, column=c).border = thin_border
    
    # 3. PURCHASE ORDER STATUS SHEET
    ws_status = wb.create_sheet(title="PO Status")
    ws_status['A1'] = "PURCHASE ORDER STATUS"
    ws_status['A1'].font = header_font
    ws_status['A1'].fill = header_fill
    ws_status.merge_cells('A1:I1')
    ws_status['A1'].alignment = header_alignment
    
    status_headers = ['PO Number', 'Supplier', 'Store', 'Order Date', 'Due Date', 
                     'Amount (UGX)', 'Items', 'Status', 'Payment Status']
    for col, header in enumerate(status_headers, start=1):
        cell = ws_status.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    po_details = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).select_related('supplier', 'store').order_by('-purchase_date')
    
    row = 4
    for order in po_details:
        payment_status = 'Pending'
        if order.status == 'completed':
            payment_status = 'Paid'
        elif order.expected_date and order.expected_date < today:
            payment_status = 'Overdue'
        
        ws_status.cell(row=row, column=1, value=f"PO-{order.id}")
        ws_status.cell(row=row, column=2, value=order.supplier.name if order.supplier else 'N/A')
        ws_status.cell(row=row, column=3, value=order.store.name if order.store else 'N/A')
        ws_status.cell(row=row, column=4, value=order.purchase_date.strftime('%Y-%m-%d') if order.purchase_date else '')
        ws_status.cell(row=row, column=5, value=order.expected_date.strftime('%Y-%m-%d') if order.expected_date else 'N/A')
        ws_status.cell(row=row, column=6, value=float(order.total_cost))
        ws_status.cell(row=row, column=7, value=order.items.count())
        ws_status.cell(row=row, column=8, value=order.status.capitalize())
        ws_status.cell(row=row, column=9, value=payment_status)
        
        # Format currency cell
        ws_status.cell(row=row, column=6).number_format = currency_format
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 10):
            ws_status.cell(row=r, column=c).border = thin_border
    
    # 4. ITEM ANALYSIS SHEET
    ws_items = wb.create_sheet(title="Item Analysis")
    ws_items['A1'] = "ITEM-WISE PURCHASE ANALYSIS"
    ws_items['A1'].font = header_font
    ws_items['A1'].fill = header_fill
    ws_items.merge_cells('A1:G1')
    ws_items['A1'].alignment = header_alignment
    
    item_headers = ['Item Code', 'Item Name', 'Category', 'Quantity Purchased', 
                   'Avg Unit Cost (UGX)', 'Total Cost (UGX)', '% of Total']
    for col, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    item_analysis = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).values(
        'product__sku',
        'product__name',
        'product__category__name'
    ).annotate(
        quantity_purchased=Sum('quantity'),
        total_cost=Sum(F('quantity') * F('unit_cost')),
        avg_unit_cost=Avg('unit_cost')
    ).order_by('-total_cost')
    
    row = 4
    for item in item_analysis:
        percent = (item['total_cost'] / total_purchases * 100) if total_purchases > 0 else 0
        
        ws_items.cell(row=row, column=1, value=item['product__sku'] or 'N/A')
        ws_items.cell(row=row, column=2, value=item['product__name'])
        ws_items.cell(row=row, column=3, value=item['product__category__name'] or 'Uncategorized')
        ws_items.cell(row=row, column=4, value=item['quantity_purchased'])
        ws_items.cell(row=row, column=5, value=float(item['avg_unit_cost'] or 0))
        ws_items.cell(row=row, column=6, value=float(item['total_cost'] or 0))
        ws_items.cell(row=row, column=7, value=percent / 100)  # Excel expects decimal for percentage
        
        # Format cells
        ws_items.cell(row=row, column=5).number_format = currency_format
        ws_items.cell(row=row, column=6).number_format = currency_format
        ws_items.cell(row=row, column=7).number_format = percent_format
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 8):
            ws_items.cell(row=r, column=c).border = thin_border
    
    # 5. EXPIRY TRACKING SHEET
    ws_expiry = wb.create_sheet(title="Expiry Tracking")
    ws_expiry['A1'] = "EXPIRY TRACKING REPORT"
    ws_expiry['A1'].font = header_font
    ws_expiry['A1'].fill = header_fill
    ws_expiry.merge_cells('A1:J1')
    ws_expiry['A1'].alignment = header_alignment
    
    expiry_headers = ['Item Code', 'Item Name', 'Batch Number', 'Current Stock', 
                     'Purchase Date', 'Expiry Date', 'Days Remaining', 
                     'Unit Cost (UGX)', 'Total Value (UGX)', 'Status']
    for col, header in enumerate(expiry_headers, start=1):
        cell = ws_expiry.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    # Get batches that will expire within next 120 days
    hundred_twenty_days_from_now = today + timedelta(days=120)
    expiry_batches = InventoryBatch.objects.filter(
        expiry_date__lte=hundred_twenty_days_from_now,
        remaining_quantity__gt=0
    ).select_related('product').order_by('expiry_date')
    
    row = 4
    total_expiry_value = Decimal('0')
    
    for batch in expiry_batches:
        days_remaining = (batch.expiry_date - today).days if batch.expiry_date else None
        
        if days_remaining is None:
            status = 'No Expiry'
        elif days_remaining < 0:
            status = 'Expired'
        elif days_remaining <= 30:
            status = 'Critical'
        elif days_remaining <= 60:
            status = 'Warning'
        elif days_remaining <= 90:
            status = 'Monitor'
        else:
            status = 'Safe'
        
        batch_value = batch.remaining_quantity * batch.unit_cost
        total_expiry_value += batch_value
        
        ws_expiry.cell(row=row, column=1, value=batch.product.sku if batch.product.sku else 'N/A')
        ws_expiry.cell(row=row, column=2, value=batch.product.name)
        ws_expiry.cell(row=row, column=3, value=f"BATCH-{batch.id}")
        ws_expiry.cell(row=row, column=4, value=batch.remaining_quantity)
        ws_expiry.cell(row=row, column=5, value=batch.created_at.strftime('%Y-%m-%d') if batch.created_at else '')
        ws_expiry.cell(row=row, column=6, value=batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A')
        ws_expiry.cell(row=row, column=7, value=days_remaining if days_remaining is not None else 'N/A')
        ws_expiry.cell(row=row, column=8, value=float(batch.unit_cost))
        ws_expiry.cell(row=row, column=9, value=float(batch_value))
        ws_expiry.cell(row=row, column=10, value=status)
        
        # Format currency cells
        ws_expiry.cell(row=row, column=8).number_format = currency_format
        ws_expiry.cell(row=row, column=9).number_format = currency_format
        
        # Color code based on status
        if status == 'Expired':
            fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        elif status == 'Critical':
            fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
        elif status == 'Warning':
            fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        elif status == 'Monitor':
            fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        else:
            fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        
        for c in range(1, 11):
            ws_expiry.cell(row=row, column=c).fill = fill
        
        row += 1
    
    # Add summary row
    ws_expiry.cell(row=row, column=1, value="TOTAL AT RISK").font = total_font
    ws_expiry.cell(row=row, column=2, value="")
    ws_expiry.cell(row=row, column=3, value="")
    ws_expiry.cell(row=row, column=4, value="")
    ws_expiry.cell(row=row, column=5, value="")
    ws_expiry.cell(row=row, column=6, value="")
    ws_expiry.cell(row=row, column=7, value="")
    ws_expiry.cell(row=row, column=8, value="")
    ws_expiry.cell(row=row, column=9, value=float(total_expiry_value)).font = total_font
    ws_expiry.cell(row=row, column=10, value="")
    ws_expiry.cell(row=row, column=9).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 11):
            ws_expiry.cell(row=r, column=c).border = thin_border
    
    # 6. STORE PERFORMANCE SHEET
    ws_stores = wb.create_sheet(title="Store Performance")
    ws_stores['A1'] = "STORE PURCHASE PERFORMANCE"
    ws_stores['A1'].font = header_font
    ws_stores['A1'].fill = header_fill
    ws_stores.merge_cells('A1:F1')
    ws_stores['A1'].alignment = header_alignment
    
    store_headers = ['Store Name', 'Total Purchases (UGX)', '% of Total', 
                    'Orders', 'Avg Order (UGX)', 'Performance Rating']
    for col, header in enumerate(store_headers, start=1):
        cell = ws_stores.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    store_performance = PurchaseOrder.objects.filter(
        purchase_date__range=[start_date, end_date]
    ).values(
        'store__name'
    ).annotate(
        total_purchases=Sum('total_cost'),
        orders=Count('id')
    ).order_by('-total_purchases')
    
    row = 4
    for store in store_performance:
        percent = (store['total_purchases'] / total_purchases * 100) if total_purchases > 0 else 0
        avg_order = store['total_purchases'] / store['orders'] if store['orders'] > 0 else Decimal('0')
        
        # Determine performance
        if percent > 40:
            rating = 'A+'
        elif percent > 30:
            rating = 'A'
        elif percent > 20:
            rating = 'B+'
        elif percent > 10:
            rating = 'B'
        elif percent > 5:
            rating = 'C'
        else:
            rating = 'D'
        
        ws_stores.cell(row=row, column=1, value=store['store__name'] or 'Unknown')
        ws_stores.cell(row=row, column=2, value=float(store['total_purchases']))
        ws_stores.cell(row=row, column=3, value=percent / 100)  # Excel expects decimal
        ws_stores.cell(row=row, column=4, value=store['orders'])
        ws_stores.cell(row=row, column=5, value=float(avg_order))
        ws_stores.cell(row=row, column=6, value=rating)
        
        # Format cells
        ws_stores.cell(row=row, column=2).number_format = currency_format
        ws_stores.cell(row=row, column=3).number_format = percent_format
        ws_stores.cell(row=row, column=5).number_format = currency_format
        
        row += 1
    
    # Add totals row
    ws_stores.cell(row=row, column=1, value="TOTAL").font = total_font
    ws_stores.cell(row=row, column=2, value=float(total_purchases)).font = total_font
    ws_stores.cell(row=row, column=3, value=1).font = total_font  # 100%
    ws_stores.cell(row=row, column=4, value=total_orders).font = total_font
    ws_stores.cell(row=row, column=5, value=float(avg_order_value)).font = total_font
    ws_stores.cell(row=row, column=6, value="")
    ws_stores.cell(row=row, column=2).number_format = currency_format
    ws_stores.cell(row=row, column=3).number_format = percent_format
    ws_stores.cell(row=row, column=5).number_format = currency_format
    
    # Add border to all data cells
    for r in range(3, row + 1):
        for c in range(1, 7):
            ws_stores.cell(row=r, column=c).border = thin_border
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"purchase_report_{period or timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response




# ============================================================================
# SALES REPORTS VIEWS
# ============================================================================
# Daily Sales Summary, Customer-wise Sales, Product Sales Performance,
# Payment Method Analysis, Transaction Audit Trail, Store Sales Performance
# ============================================================================





@login_required
def sales_details(request):
    """
    Sales Performance Dashboard view.
    Supports filtering by: store, payment method, status, and date range.
    """

    # ─── 1. Parse filter parameters from GET request ───────────────────────────

    store_id     = request.GET.get('store', '').strip()
    payment_id   = request.GET.get('payment', '').strip()
    status_param = request.GET.get('status', '').strip()
    daterange    = request.GET.get('daterange', '').strip()

    start_date = None
    end_date   = None
    date_range_label = 'All Time'

    if daterange:
        try:
            parts = [p.strip() for p in daterange.split(' - ')]
            if len(parts) == 2:
                start_date = date.fromisoformat(parts[0])
                end_date   = date.fromisoformat(parts[1])
                date_range_label = daterange
        except ValueError:
            pass  # Ignore malformed date range

    # ─── 2. Build base queryset with select_related for performance ─────────────

    qs = Sales.objects.select_related(
        'customer',
        'store',
        'store__branch',
        'payment_method',
        'recorded_by',
    ).prefetch_related('items')

    # ─── 3. Apply filters ───────────────────────────────────────────────────────

    if store_id:
        qs = qs.filter(store_id=store_id)

    if payment_id:
        qs = qs.filter(payment_method_id=payment_id)

    if status_param:
        if status_param == 'completed':
            qs = qs.filter(is_cancelled=False, balance=0)
        elif status_param == 'pending':
            qs = qs.filter(is_cancelled=False, balance__gt=0)
        elif status_param == 'cancelled':
            qs = qs.filter(is_cancelled=True)

    if start_date and end_date:
        qs = qs.filter(sale_date__range=[start_date, end_date])

    # ─── 4. Annotate each sale with item counts for the table ──────────────────

    qs = qs.annotate(
        total_quantity=Sum('items__quantity'),
    )

    # ─── 5. Aggregate KPI metrics ───────────────────────────────────────────────

    aggregates = qs.aggregate(
        total_sales=Sum('total_amount'),
        total_paid=Sum('amount_paid'),
        total_balance=Sum('balance'),
        total_transactions=Count('id'),
        avg_transaction_value=Avg('total_amount'),
    )

    total_sales           = aggregates['total_sales']           or 0
    total_paid            = aggregates['total_paid']            or 0
    total_balance         = aggregates['total_balance']         or 0
    total_transactions    = aggregates['total_transactions']    or 0
    avg_transaction_value = aggregates['avg_transaction_value'] or 0

    # Unique customers
    total_customers = qs.filter(customer__isnull=False).values('customer').distinct().count()

    # Products sold (distinct products across all sale items in the filtered qs)
    sale_ids = qs.values_list('id', flat=True)
    total_products_sold = (
        SalesItem.objects
        .filter(order_id__in=sale_ids)
        .values('product')
        .distinct()
        .count()
    )
    total_units_sold = (
        SalesItem.objects
        .filter(order_id__in=sale_ids)
        .aggregate(total=Sum('quantity'))['total'] or 0
    )

    # New customers: customers whose first sale falls within the filtered period
    if start_date and end_date:
        new_customers = (
            Sales.objects
            .filter(customer__isnull=False, sale_date__range=[start_date, end_date])
            .values('customer')
            .distinct()
            .count()
        )
    else:
        new_customers = total_customers  # No date filter — all are "new"

    # ─── 6. Today's summary (always unfiltered by store/payment/status) ─────────

    today = date.today()
    today_qs = Sales.objects.filter(sale_date=today, is_cancelled=False)
    today_sales        = today_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    today_transactions = today_qs.count()

    # ─── 7. Payment method breakdown ───────────────────────────────────────────

    payment_summary = (
        qs.filter(payment_method__isnull=False)
        .values('payment_method__id', 'payment_method__name')
        .annotate(count=Count('id'), total=Sum('total_amount'))
        .order_by('-total')
    )
    payment_summary = [
        {'id': row['payment_method__id'], 'name': row['payment_method__name'],
         'count': row['count'], 'total': row['total'] or 0}
        for row in payment_summary
    ]

    # ─── 8. Store performance breakdown ────────────────────────────────────────

    store_summary = (
        qs.values('store__id', 'store__name')
        .annotate(total=Sum('total_amount'), count=Count('id'))
        .order_by('-total')
    )
    store_summary = [
        {'id': row['store__id'], 'name': row['store__name'],
         'total': row['total'] or 0, 'count': row['count']}
        for row in store_summary
    ]

    # ─── 9. Top products ────────────────────────────────────────────────────────

    top_products = (
        SalesItem.objects
        .filter(order_id__in=sale_ids, is_cancelled=False)
        .values('product__id', 'product__name')
        .annotate(quantity=Sum('quantity'))
        .order_by('-quantity')[:10]
    )
    top_products = [
        {'id': row['product__id'], 'name': row['product__name'], 'quantity': row['quantity']}
        for row in top_products
    ]

    # ─── 10. Dropdown data for filter controls ──────────────────────────────────

    stores          = StoreLocation.objects.filter(is_active=True).order_by('name')
    payment_methods = PaymentMethod.objects.all().order_by('name')

    # ─── 11. Render ─────────────────────────────────────────────────────────────

    context = {
        # Filter state (so template can pre-populate controls on page reload)
        'date_range':    date_range_label,
        'selected_store':   store_id,
        'selected_payment': payment_id,
        'selected_status':  status_param,

        # Dropdown options
        'stores':           stores,
        'payment_methods':  payment_methods,

        # KPI cards
        'total_sales':           total_sales,
        'total_transactions':    total_transactions,
        'avg_transaction_value': avg_transaction_value,
        'total_customers':       total_customers,
        'new_customers':         new_customers,
        'total_products_sold':   total_products_sold,
        'total_units_sold':      total_units_sold,

        # Table footer totals
        'total_paid':    total_paid,
        'total_balance': total_balance,

        # Today's summary card
        'today_sales':        today_sales,
        'today_transactions': today_transactions,

        # Summary cards
        'payment_summary': payment_summary,
        'store_summary':   store_summary,
        'top_products':    top_products,

        # Main table data
        'sales_data': qs.order_by('-created_at'),
    }

    return render(request, 'reports/sales_details.html', context)



def export_sales_csv(request):
    """Export comprehensive sales report as CSV"""
    # Get filter parameters from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    daterange = request.GET.get('daterange')
    store_id = request.GET.get('store')
    period = request.GET.get('period', '')
    
    # Set up response
    response = HttpResponse(content_type='text/csv')
    filename = f"sales_report_{period or timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write header information
    writer.writerow(['SALES REPORT EXPORT'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Period:', period])
    writer.writerow(['Generated By:', request.user.get_full_name() or request.user.username])
    writer.writerow([])
    
    # Get the same data as the sales_details view
    today = timezone.now()
    start_date = None
    end_date = None
    if date_from and date_to:
        try:
            start_date = timezone.make_aware(datetime.strptime(date_from, '%Y-%m-%d'))
            end_date = timezone.make_aware(datetime.strptime(date_to, '%Y-%m-%d'))
        except Exception:
            start_date = None
            end_date = None
    if (not start_date or not end_date) and daterange:
        try:
            parts = [p.strip() for p in daterange.split(' - ')]
            if len(parts) == 2:
                start_date = timezone.make_aware(datetime.strptime(parts[0], '%Y-%m-%d'))
                end_date = timezone.make_aware(datetime.strptime(parts[1], '%Y-%m-%d'))
        except Exception:
            start_date = None
            end_date = None
    if not start_date or not end_date:
        start_date = today.replace(day=1)
        end_date = today
    
    # Filter sales for the period
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    # 1. SUMMARY SECTION
    writer.writerow(['SUMMARY SECTION'])
    writer.writerow([])
    
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    writer.writerow(['Total Sales:', f"UGX {total_sales:,.0f}"])
    writer.writerow(['Total Transactions:', f"{total_transactions}"])
    writer.writerow(['Average Daily Sales:', f"UGX {avg_daily_sales:,.0f}"])
    writer.writerow(['Date Range:', f"{start_date.date()} to {end_date.date()}"])
    writer.writerow([])
    
    # 2. DAILY SALES SUMMARY
    writer.writerow(['DAILY SALES SUMMARY'])
    writer.writerow(['Date', 'Day', 'Total Sales (UGX)', 'Transactions', 'Avg Transaction (UGX)'])
    
    current_date = start_date.date()
    while current_date <= end_date.date():
        daily_sales = Sales.objects.filter(sale_date=current_date)
        daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        daily_transactions = daily_sales.count()
        avg_transaction = daily_total / Decimal(daily_transactions) if daily_transactions > 0 else Decimal('0')
        
        writer.writerow([
            current_date.strftime('%Y-%m-%d'),
            current_date.strftime('%A'),
            f"{daily_total:,.0f}",
            daily_transactions,
            f"{avg_transaction:,.0f}"
        ])
        current_date += timedelta(days=1)
    
    writer.writerow([])
    
    # 3. CUSTOMER-WISE SALES
    writer.writerow(['CUSTOMER-WISE SALES ANALYSIS'])
    writer.writerow(['Customer Name', 'Customer Type', 'Total Spent (UGX)', 'Transactions', 'Avg Order (UGX)', 'Segment'])
    
    customer_sales = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'customer__name', 'customer__company'
    ).annotate(
        total_spent=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_spent')
    
    for customer in customer_sales:
        avg_order = customer['total_spent'] / Decimal(customer['transactions']) if customer['transactions'] > 0 else Decimal('0')
        customer_type = 'Corporate' if customer['customer__company'] else 'Individual'
        
        # Determine segment
        if customer['total_spent'] > 1000000:
            segment = 'VIP'
        elif customer['transactions'] > 5:
            segment = 'Loyal'
        else:
            segment = 'Regular'
        
        writer.writerow([
            customer['customer__name'] or 'Walk-in Customer',
            customer_type,
            f"{customer['total_spent']:,.0f}",
            customer['transactions'],
            f"{avg_order:,.0f}",
            segment
        ])
    
    writer.writerow([])
    
    # 4. PRODUCT PERFORMANCE
    writer.writerow(['PRODUCT SALES PERFORMANCE'])
    writer.writerow(['Product Name', 'Product Code', 'Category', 'Units Sold', 'Revenue (UGX)', 'Avg Price (UGX)', '% of Total'])
    
    product_performance = SalesItem.objects.filter(
        order__sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'product__name', 'product__sku', 'product__category__name'
    ).annotate(
        units_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('sale_price'), output_field=DecimalField()),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')
    
    for product in product_performance:
        percent = (product['revenue'] / total_sales * 100) if total_sales > 0 else 0
        
        writer.writerow([
            product['product__name'],
            product['product__sku'] or 'N/A',
            product['product__category__name'] or 'Uncategorized',
            product['units_sold'],
            f"{product['revenue']:,.0f}",
            f"{product['avg_price']:,.0f}",
            f"{percent:.1f}%"
        ])
    
    writer.writerow([])
    
    # 5. PAYMENT METHOD ANALYSIS
    writer.writerow(['PAYMENT METHOD ANALYSIS'])
    writer.writerow(['Payment Method', 'Transactions', '% of Total', 'Total Amount (UGX)', 'Avg Transaction (UGX)'])
    
    payment_analysis = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()],
        payment_method__isnull=False
    ).values(
        'payment_method__name'
    ).annotate(
        transactions=Count('id'),
        total_amount=Sum('total_amount')
    ).order_by('-transactions')
    
    for payment in payment_analysis:
        percent = (payment['transactions'] / total_transactions * 100) if total_transactions > 0 else 0
        avg_transaction = payment['total_amount'] / Decimal(payment['transactions']) if payment['transactions'] > 0 else Decimal('0')
        
        writer.writerow([
            payment['payment_method__name'] or 'Unknown',
            payment['transactions'],
            f"{percent:.1f}%",
            f"{payment['total_amount']:,.0f}",
            f"{avg_transaction:,.0f}"
        ])
    
    writer.writerow([])
    
    # 6. STORE PERFORMANCE
    writer.writerow(['STORE SALES PERFORMANCE'])
    writer.writerow(['Store', 'Sales (UGX)', '% of Total', 'Transactions', 'Avg Transaction (UGX)', 'Performance'])
    
    store_performance = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'store__name'
    ).annotate(
        total_sales=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_sales')
    
    for store in store_performance:
        percent = (store['total_sales'] / total_sales * 100) if total_sales > 0 else 0
        avg_transaction = store['total_sales'] / Decimal(store['transactions']) if store['transactions'] > 0 else Decimal('0')
        
        # Determine performance rating
        if percent > 30:
            performance = 'Excellent'
        elif percent > 15:
            performance = 'Good'
        elif percent > 5:
            performance = 'Average'
        else:
            performance = 'Needs Attention'
        
        writer.writerow([
            store['store__name'],
            f"{store['total_sales']:,.0f}",
            f"{percent:.1f}%",
            store['transactions'],
            f"{avg_transaction:,.0f}",
            performance
        ])
    
    writer.writerow([])
    
    # 7. TRANSACTION AUDIT TRAIL
    writer.writerow(['TRANSACTION AUDIT TRAIL'])
    writer.writerow(['Transaction ID', 'Date', 'Customer', 'Store', 'Items', 'Total (UGX)', 'Payment Method', 'Status', 'Receipt No'])
    
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')[:100]
    
    for transaction in recent_transactions:
        writer.writerow([
            transaction.id,
            transaction.sale_date.strftime('%Y-%m-%d') if transaction.sale_date else '',
            transaction.customer.name if transaction.customer else 'Walk-in',
            transaction.store.name if transaction.store else '',
            transaction.number_of_items or 0,
            f"{transaction.total_amount:,.0f}",
            transaction.payment_method.name if transaction.payment_method else 'Unknown',
            transaction.status,
            transaction.receipt_no or 'N/A'
        ])
    
    return response


def export_sales_pdf(request):
    """Export comprehensive sales report as PDF with black and white design"""
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    period = request.GET.get('period', 'Custom Range')
    
    # Set up response
    response = HttpResponse(content_type='application/pdf')
    filename = f"sales_report_{period.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Parse dates
    today = timezone.now().date()
    if date_from and date_to:
        try:
            start_date = date.fromisoformat(date_from)
            end_date = date.fromisoformat(date_to)
        except:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Create PDF
    doc = SimpleDocTemplate(response, pagesize=letter, 
                           rightMargin=40, leftMargin=40,
                           topMargin=40, bottomMargin=40)
    elements = []
    
    # Define styles - Black and White only
    styles = getSampleStyleSheet()
    
    # Custom styles for B&W report
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.black,
        alignment=1,  # Center
        spaceAfter=20,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.black,
        spaceBefore=15,
        spaceAfter=8,
        fontName='Helvetica-Bold',
        borderWidth=0,
        borderColor=colors.black,
        borderPadding=5
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.black,
        fontName='Helvetica'
    )
    
    # Header Section
    elements.append(Paragraph("SALES REPORT", title_style))
    elements.append(Spacer(1, 10))
    
    # Report Info Table
    info_data = [
        [f"Period: {period}", f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"],
        [f"Date Range: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}", 
         f"Generated By: {request.user.get_full_name() or request.user.username}"]
    ]
    
    if store_id:
        try:
            store = StoreLocation.objects.get(id=store_id)
            info_data.append([f"Store: {store.name}", ""])
        except:
            pass
    
    info_table = Table(info_data, colWidths=[250, 250])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Get sales data
    sales_qs = Sales.objects.filter(
        sale_date__range=[start_date, end_date],
        is_cancelled=False
    )
    if store_id:
        sales_qs = sales_qs.filter(store_id=store_id)
    
    total_sales = sales_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_qs.count()
    days_in_period = max((end_date - start_date).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    total_paid = sales_qs.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_balance = sales_qs.aggregate(total=Sum('balance'))['total'] or Decimal('0')
    total_received = sales_qs.aggregate(total=Sum('amount_received'))['total'] or Decimal('0')
    total_change = sales_qs.aggregate(total=Sum('change'))['total'] or Decimal('0')
    
    total_items = 0
    for sale in sales_qs:
        total_items += sale.total_items
    
    # 1. SUMMARY SECTION
    elements.append(Paragraph("SUMMARY SECTION", heading_style))
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Transactions', f"{total_transactions:,}"],
        ['Total Revenue', f"UGX {total_sales:,.0f}"],
        ['Total Items Sold', f"{total_items:,} pcs"],
        ['Average Order Value', f"UGX {(total_sales/total_transactions):,.0f}" if total_transactions > 0 else 'UGX 0'],
        ['Average Daily Sales', f"UGX {avg_daily_sales:,.0f}"],
        ['Total Cash Received', f"UGX {total_received:,.0f}"],
        ['Total Change Given', f"UGX {total_change:,.0f}"],
        ['Outstanding Balance', f"UGX {total_balance:,.0f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # 2. SALES BY STATUS
    elements.append(Paragraph("SALES BY STATUS", heading_style))
    
    status_counts = {}
    status_totals = {}
    for status_code, status_name in SALE_ORDER_OPTIONS:
        status_qs = sales_qs.filter(status=status_code)
        count = status_qs.count()
        total = status_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        if count > 0:
            status_counts[status_name] = count
            status_totals[status_name] = total
    
    status_data = [['Status', 'Count', 'Total Amount', '% of Sales']]
    for status_name in status_counts:
        count = status_counts[status_name]
        total = status_totals[status_name]
        percent = (total / total_sales * 100) if total_sales > 0 else 0
        status_data.append([status_name, str(count), f"UGX {total:,.0f}", f"{percent:.1f}%"])
    
    # Add total row
    status_data.append(['TOTAL', str(total_transactions), f"UGX {total_sales:,.0f}", '100%'])
    
    status_table = Table(status_data, colWidths=[150, 100, 150, 100])
    status_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
    ]))
    elements.append(status_table)
    elements.append(Spacer(1, 20))
    
    # 3. SALES BY PAYMENT METHOD
    elements.append(Paragraph("SALES BY PAYMENT METHOD", heading_style))
    
    payment_analysis = sales_qs.filter(
        payment_method__isnull=False
    ).values(
        'payment_method__name'
    ).annotate(
        transactions=Count('id'),
        total_amount=Sum('total_amount')
    ).order_by('-transactions')
    
    payment_data = [['Payment Method', 'Transactions', 'Total Amount', '% of Total']]
    for payment in payment_analysis:
        percent = (payment['total_amount'] / total_sales * 100) if total_sales > 0 else 0
        payment_data.append([
            payment['payment_method__name'] or 'Unknown',
            str(payment['transactions']),
            f"UGX {payment['total_amount']:,.0f}",
            f"{percent:.1f}%"
        ])
    
    # Add cash sales without payment method
    cash_sales = sales_qs.filter(payment_method__isnull=True)
    cash_count = cash_sales.count()
    cash_total = cash_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    if cash_count > 0:
        cash_percent = (cash_total / total_sales * 100) if total_sales > 0 else 0
        payment_data.append(['Cash (Unspecified)', str(cash_count), f"UGX {cash_total:,.0f}", f"{cash_percent:.1f}%"])
    
    payment_data.append(['TOTAL', str(total_transactions), f"UGX {total_sales:,.0f}", '100%'])
    
    payment_table = Table(payment_data, colWidths=[150, 100, 150, 100])
    payment_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
    ]))
    elements.append(payment_table)
    elements.append(Spacer(1, 20))
    
    # 4. DETAILED SALES TRANSACTIONS
    elements.append(Paragraph("DETAILED SALES TRANSACTIONS", heading_style))
    
    recent_transactions = sales_qs.select_related(
        'customer', 'store', 'payment_method', 'recorded_by'
    ).order_by('-sale_date', '-id')[:50]  # Limit to 50 for PDF
    
    transaction_data = [
        ['Receipt No', 'Date', 'Customer', 'Items', 'Total', 'Paid', 'Balance', 'Payment', 'Status']
    ]
    
    for t in recent_transactions:
        transaction_data.append([
            t.receipt_no or f"SO-{t.id}",
            t.sale_date.strftime('%Y-%m-%d') if t.sale_date else '',
            t.customer.name[:15] + '...' if t.customer and len(t.customer.name) > 15 else (t.customer.name if t.customer else 'Walk-in'),
            str(t.number_of_items or 0),
            f"UGX {t.total_amount:,.0f}",
            f"UGX {t.amount_paid:,.0f}",
            f"UGX {t.balance:,.0f}",
            t.payment_method.name[:10] if t.payment_method else 'Cash',
            t.status[:10]
        ])
    
    transaction_table = Table(transaction_data, colWidths=[70, 60, 80, 40, 70, 70, 70, 60, 60])
    transaction_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (3, 1), (8, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(transaction_table)
    
    # Add note if limited
    if sales_qs.count() > 50:
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(f"* Showing 50 of {sales_qs.count()} transactions", normal_style))
    
    elements.append(Spacer(1, 20))
    
    # 5. SALES ITEMS DETAIL
    elements.append(Paragraph("SALES ITEMS DETAIL", heading_style))
    
    # Get top 50 items by revenue
    items = SalesItem.objects.filter(
        order__in=sales_qs
    ).select_related(
        'product', 'unit', 'order'
    ).order_by('-order__sale_date', '-id')[:50]
    
    items_data = [
        ['Receipt No', 'Product', 'SKU', 'Unit', 'Qty', 'Base Qty', 'Unit Price', 'Line Total']
    ]
    
    for item in items:
        items_data.append([
            item.order.receipt_no or f"SO-{item.order.id}",
            item.product.name[:15] + '...' if len(item.product.name) > 15 else item.product.name,
            item.product.sku or 'N/A',
            item.unit.abbreviation if item.unit else 'unit',
            str(item.quantity),
            str(item.base_quantity),
            f"UGX {item.sale_price:,.0f}",
            f"UGX {(item.quantity * item.sale_price):,.0f}"
        ])
    
    items_table = Table(items_data, colWidths=[70, 80, 50, 40, 40, 50, 70, 80])
    items_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 20))
    
    # 6. TOP SELLING PRODUCTS
    elements.append(Paragraph("TOP SELLING PRODUCTS", heading_style))
    
    top_products = SalesItem.objects.filter(
        order__in=sales_qs
    ).values(
        'product__name', 'product__sku'
    ).annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('sale_price')),
        transactions=Count('order', distinct=True)
    ).order_by('-total_revenue')[:10]
    
    products_data = [['Rank', 'Product', 'SKU', 'Qty Sold', 'Revenue', 'Transactions']]
    
    for i, product in enumerate(top_products, 1):
        products_data.append([
            str(i),
            product['product__name'][:20] + '...' if len(product['product__name']) > 20 else product['product__name'],
            product['product__sku'] or 'N/A',
            str(product['total_qty']),
            f"UGX {product['total_revenue']:,.0f}",
            str(product['transactions'])
        ])
    
    products_table = Table(products_data, colWidths=[50, 120, 70, 70, 120, 80])
    products_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(products_table)
    elements.append(Spacer(1, 20))
    
    # 7. DAILY SALES TREND
    elements.append(Paragraph("DAILY SALES TREND", heading_style))
    
    daily_data = [['Date', 'Day', 'Transactions', 'Items Sold', 'Revenue']]
    
    current_date = start_date
    while current_date <= end_date:
        daily_sales = sales_qs.filter(sale_date=current_date)
        daily_count = daily_sales.count()
        
        if daily_count > 0:
            daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
            daily_items = 0
            for sale in daily_sales:
                daily_items += sale.total_items
            
            daily_data.append([
                current_date.strftime('%Y-%m-%d'),
                current_date.strftime('%a'),
                str(daily_count),
                str(daily_items),
                f"UGX {daily_total:,.0f}"
            ])
        else:
            daily_data.append([
                current_date.strftime('%Y-%m-%d'),
                current_date.strftime('%a'),
                '0', '0', 'UGX 0'
            ])
        current_date += timedelta(days=1)
    
    # Add total row
    daily_data.append(['TOTAL', '', str(total_transactions), str(total_items), f"UGX {total_sales:,.0f}"])
    
    daily_table = Table(daily_data, colWidths=[80, 50, 80, 80, 120])
    daily_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
    ]))
    elements.append(daily_table)
    elements.append(Spacer(1, 20))
    
    # 8. STAFF PERFORMANCE
    elements.append(Paragraph("STAFF PERFORMANCE", heading_style))
    
    staff_performance = sales_qs.values(
        'recorded_by__username', 'recorded_by__first_name', 'recorded_by__last_name'
    ).annotate(
        transactions=Count('id'),
        total_revenue=Sum('total_amount')
    ).order_by('-total_revenue')
    
    staff_data = [['Staff', 'Transactions', 'Total Revenue', 'Avg Sale Value']]
    
    for staff in staff_performance:
        full_name = f"{staff['recorded_by__first_name']} {staff['recorded_by__last_name']}".strip()
        if not full_name:
            full_name = staff['recorded_by__username']
        
        avg_sale = staff['total_revenue'] / Decimal(staff['transactions']) if staff['transactions'] > 0 else Decimal('0')
        
        staff_data.append([
            full_name,
            str(staff['transactions']),
            f"UGX {staff['total_revenue']:,.0f}",
            f"UGX {avg_sale:,.0f}"
        ])
    
    staff_table = Table(staff_data, colWidths=[150, 100, 150, 150])
    staff_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(staff_table)
    
    # Build PDF
    doc.build(elements)
    return response



def export_sales_excel(request):
    """Export comprehensive sales report as Excel"""
    # Get filter parameters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    period = request.GET.get('period', '')
    
    # Get the same data as the sales_details view
    today = timezone.now()
    if date_from and date_to:
        try:
            start_date = timezone.make_aware(date.strptime(date_from, '%Y-%m-%d'))
            end_date = timezone.make_aware(date.strptime(date_to, '%Y-%m-%d'))
        except:
            start_date = today.replace(day=1)
            end_date = today
    else:
        start_date = today.replace(day=1)
        end_date = today
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    subheader_font = Font(bold=True, color="000000", size=11)
    subheader_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
    
    total_font = Font(bold=True, color="000000", size=10)
    total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. SUMMARY SHEET
    ws_summary = wb.create_sheet(title="Summary")
    
    # Header
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = f"SALES REPORT - {period}"
    ws_summary['A1'].font = header_font
    ws_summary['A1'].fill = header_fill
    ws_summary['A1'].alignment = header_alignment
    
    ws_summary['A3'] = "Generated:"
    ws_summary['B3'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    ws_summary['A4'] = "Period:"
    ws_summary['B4'] = period
    ws_summary['A5'] = "Generated By:"
    ws_summary['B5'] = request.user.get_full_name() or request.user.username
    ws_summary['A6'] = "Date Range:"
    ws_summary['B6'] = f"{start_date.date()} to {end_date.date()}"
    
    # Get summary data
    sales_in_period = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    )
    
    total_sales = sales_in_period.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_transactions = sales_in_period.count()
    days_in_period = max((end_date.date() - start_date.date()).days + 1, 1)
    avg_daily_sales = total_sales / Decimal(days_in_period) if total_sales else Decimal('0')
    
    # Summary table
    ws_summary['A8'] = "KEY METRICS"
    ws_summary['A8'].font = subheader_font
    ws_summary['A8'].fill = subheader_fill
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Sales', f"UGX {total_sales:,.0f}"],
        ['Total Transactions', total_transactions],
        ['Average Daily Sales', f"UGX {avg_daily_sales:,.0f}"],
        ['Days in Period', days_in_period],
        ['Average Transaction Value', f"UGX {total_sales/total_transactions:,.0f}" if total_transactions > 0 else "UGX 0"]
    ]
    
    for i, row in enumerate(summary_data, start=9):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = thin_border
            if i == 9:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # 2. DAILY SALES SHEET
    ws_daily = wb.create_sheet(title="Daily Sales")
    ws_daily['A1'] = "DAILY SALES SUMMARY"
    ws_daily['A1'].font = header_font
    ws_daily['A1'].fill = header_fill
    ws_daily.merge_cells('A1:E1')
    ws_daily['A1'].alignment = header_alignment
    
    daily_headers = ['Date', 'Day', 'Total Sales (UGX)', 'Transactions', 'Avg Transaction (UGX)']
    for col, header in enumerate(daily_headers, start=1):
        cell = ws_daily.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    row = 4
    current_date = start_date.date()
    while current_date <= end_date.date():
        daily_sales = Sales.objects.filter(sale_date=current_date)
        daily_total = daily_sales.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        daily_transactions = daily_sales.count()
        avg_transaction = daily_total / Decimal(daily_transactions) if daily_transactions > 0 else Decimal('0')
        
        ws_daily.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
        ws_daily.cell(row=row, column=2, value=current_date.strftime('%A'))
        ws_daily.cell(row=row, column=3, value=float(daily_total))
        ws_daily.cell(row=row, column=4, value=daily_transactions)
        ws_daily.cell(row=row, column=5, value=float(avg_transaction))
        
        # Format currency cells
        ws_daily.cell(row=row, column=3).number_format = '"UGX "#,##0'
        ws_daily.cell(row=row, column=5).number_format = '"UGX "#,##0'
        
        row += 1
        current_date += timedelta(days=1)
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 6):
            ws_daily.cell(row=r, column=c).border = thin_border
    
    # 3. CUSTOMER SALES SHEET
    ws_customers = wb.create_sheet(title="Customer Sales")
    ws_customers['A1'] = "CUSTOMER-WISE SALES"
    ws_customers['A1'].font = header_font
    ws_customers['A1'].fill = header_fill
    ws_customers.merge_cells('A1:F1')
    ws_customers['A1'].alignment = header_alignment
    
    customer_headers = ['Customer Name', 'Customer Type', 'Total Spent (UGX)', 'Transactions', 'Avg Order (UGX)', 'Segment']
    for col, header in enumerate(customer_headers, start=1):
        cell = ws_customers.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    customer_sales = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'customer__name', 'customer__company'
    ).annotate(
        total_spent=Sum('total_amount'),
        transactions=Count('id')
    ).order_by('-total_spent')
    
    row = 4
    for customer in customer_sales:
        avg_order = customer['total_spent'] / Decimal(customer['transactions']) if customer['transactions'] > 0 else Decimal('0')
        customer_type = 'Corporate' if customer['customer__company'] else 'Individual'
        
        # Determine segment
        if customer['total_spent'] > 1000000:
            segment = 'VIP'
        elif customer['transactions'] > 5:
            segment = 'Loyal'
        else:
            segment = 'Regular'
        
        ws_customers.cell(row=row, column=1, value=customer['customer__name'] or 'Walk-in Customer')
        ws_customers.cell(row=row, column=2, value=customer_type)
        ws_customers.cell(row=row, column=3, value=float(customer['total_spent']))
        ws_customers.cell(row=row, column=4, value=customer['transactions'])
        ws_customers.cell(row=row, column=5, value=float(avg_order))
        ws_customers.cell(row=row, column=6, value=segment)
        
        # Format currency cells
        ws_customers.cell(row=row, column=3).number_format = '"UGX "#,##0'
        ws_customers.cell(row=row, column=5).number_format = '"UGX "#,##0'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 7):
            ws_customers.cell(row=r, column=c).border = thin_border
    
    # 4. PRODUCT PERFORMANCE SHEET
    ws_products = wb.create_sheet(title="Product Performance")
    ws_products['A1'] = "PRODUCT SALES PERFORMANCE"
    ws_products['A1'].font = header_font
    ws_products['A1'].fill = header_fill
    ws_products.merge_cells('A1:G1')
    ws_products['A1'].alignment = header_alignment
    
    product_headers = ['Product Name', 'Product Code', 'Category', 'Units Sold', 'Revenue (UGX)', 'Avg Price (UGX)', '% of Total']
    for col, header in enumerate(product_headers, start=1):
        cell = ws_products.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    product_performance = SalesItem.objects.filter(
        order__sale_date__range=[start_date.date(), end_date.date()]
    ).values(
        'product__name', 'product__sku', 'product__category__name'
    ).annotate(
        units_sold=Sum('quantity'),
        revenue=Sum(F('quantity') * F('sale_price'), output_field=DecimalField()),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')
    
    row = 4
    for product in product_performance:
        percent = (product['revenue'] / total_sales * 100) if total_sales > 0 else 0
        
        ws_products.cell(row=row, column=1, value=product['product__name'])
        ws_products.cell(row=row, column=2, value=product['product__sku'] or 'N/A')
        ws_products.cell(row=row, column=3, value=product['product__category__name'] or 'Uncategorized')
        ws_products.cell(row=row, column=4, value=product['units_sold'])
        ws_products.cell(row=row, column=5, value=float(product['revenue']))
        ws_products.cell(row=row, column=6, value=float(product['avg_price']))
        ws_products.cell(row=row, column=7, value=percent)
        
        # Format cells
        ws_products.cell(row=row, column=5).number_format = '"UGX "#,##0'
        ws_products.cell(row=row, column=6).number_format = '"UGX "#,##0'
        ws_products.cell(row=row, column=7).number_format = '0.0"%'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 8):
            ws_products.cell(row=r, column=c).border = thin_border
    
    # 5. TRANSACTION AUDIT SHEET
    ws_transactions = wb.create_sheet(title="Transactions")
    ws_transactions['A1'] = "TRANSACTION AUDIT TRAIL"
    ws_transactions['A1'].font = header_font
    ws_transactions['A1'].fill = header_fill
    ws_transactions.merge_cells('A1:I1')
    ws_transactions['A1'].alignment = header_alignment
    
    transaction_headers = ['ID', 'Date', 'Customer', 'Store', 'Items', 'Total (UGX)', 'Payment Method', 'Status', 'Receipt No']
    for col, header in enumerate(transaction_headers, start=1):
        cell = ws_transactions.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = subheader_fill
        cell.border = thin_border
    
    recent_transactions = Sales.objects.filter(
        sale_date__range=[start_date.date(), end_date.date()]
    ).select_related('customer', 'store', 'payment_method').order_by('-sale_date', '-id')
    
    row = 4
    for transaction in recent_transactions:
        ws_transactions.cell(row=row, column=1, value=transaction.id)
        ws_transactions.cell(row=row, column=2, value=transaction.sale_date.strftime('%Y-%m-%d') if transaction.sale_date else '')
        ws_transactions.cell(row=row, column=3, value=transaction.customer.name if transaction.customer else 'Walk-in')
        ws_transactions.cell(row=row, column=4, value=transaction.store.name if transaction.store else '')
        ws_transactions.cell(row=row, column=5, value=transaction.number_of_items or 0)
        ws_transactions.cell(row=row, column=6, value=float(transaction.total_amount))
        ws_transactions.cell(row=row, column=7, value=transaction.payment_method.name if transaction.payment_method else 'Unknown')
        ws_transactions.cell(row=row, column=8, value=transaction.status)
        ws_transactions.cell(row=row, column=9, value=transaction.receipt_no or 'N/A')
        
        # Format currency cell
        ws_transactions.cell(row=row, column=6).number_format = '"UGX "#,##0'
        
        row += 1
    
    # Add border to all data cells
    for r in range(3, row):
        for c in range(1, 10):
            ws_transactions.cell(row=r, column=c).border = thin_border
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"sales_report_{period or timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


# ============================================================================
# INVENTORY REPORTS VIEWS
# ============================================================================
# Stock Level Report, Batch Expiry Report, Expired Stock Report,
# Inventory Valuation, Stock Aging Report, Real-time Stock Availability,
# Store-wise Stock Distribution
# ============================================================================

@login_required
def inventory_details(request):
    """
    Inventory Status Dashboard view.
    Matches: reports/inventory_details.html
    """
    today = timezone.now().date()

    # ── Filters ─────────────────────────────────────────────────────────────
    store_id      = request.GET.get('store', '')
    category_id   = request.GET.get('category', '')
    stock_status  = request.GET.get('stockstatus', '')
    expiry_filter = request.GET.get('expiry', '')
    search_query  = request.GET.get('search', '')
    daterange     = request.GET.get('daterange', '')

    start_date = None
    end_date   = None
    if daterange:
        try:
            parts = [p.strip() for p in daterange.split(' - ')]
            if len(parts) == 2:
                start_date = date.fromisoformat(parts[0])
                end_date   = date.fromisoformat(parts[1])
        except (ValueError, AttributeError):
            pass

    # ── Base inventory queryset ──────────────────────────────────────────────
    inv_qs = Inventory.objects.select_related(
        'product', 'product__category', 'store', 'store__branch'
    )
    if store_id:
        inv_qs = inv_qs.filter(store_id=store_id)
    if category_id:
        inv_qs = inv_qs.filter(product__category_id=category_id)
    if search_query:
        inv_qs = inv_qs.filter(
            Q(product__name__icontains=search_query) |
            Q(product__sku__icontains=search_query)
        )

    # Stock status filter
    if stock_status == 'out':
        inv_qs = inv_qs.filter(quantity_in_stock=0)
    elif stock_status == 'low':
        inv_qs = inv_qs.filter(quantity_in_stock__gt=0, quantity_in_stock__lte=F('reorder_level'))
    elif stock_status == 'over':
        inv_qs = inv_qs.filter(quantity_in_stock__gt=F('reorder_level') * 2)
    elif stock_status == 'healthy':
        inv_qs = inv_qs.filter(
            quantity_in_stock__gt=F('reorder_level'),
            quantity_in_stock__lte=F('reorder_level') * 2
        )

    # ── KPI counts ───────────────────────────────────────────────────────────
    total_products  = inv_qs.values('product').distinct().count()
    total_batches   = InventoryBatch.objects.filter(
        store_id__in=inv_qs.values('store_id'),
        remaining_quantity__gt=0
    ).count()

    in_stock_count  = inv_qs.filter(quantity_in_stock__gt=F('reorder_level')).count()
    low_stock_count = inv_qs.filter(quantity_in_stock__gt=0, quantity_in_stock__lte=F('reorder_level')).count()
    out_of_stock_count = inv_qs.filter(quantity_in_stock=0).count()
    total_items     = in_stock_count + low_stock_count + out_of_stock_count
    low_stock_items = low_stock_count + out_of_stock_count  # "at risk" total

    healthy_stock = round((in_stock_count / total_items * 100) if total_items else 0)
    low_stock_pct = round((low_stock_count / total_items * 100) if total_items else 0)
    out_stock_pct = round((out_of_stock_count / total_items * 100) if total_items else 0)

    total_units = inv_qs.aggregate(total=Sum('quantity_in_stock'))['total'] or 0

    # ── Expiry counts ────────────────────────────────────────────────────────
    batch_qs = InventoryBatch.objects.filter(remaining_quantity__gt=0, expiry_date__isnull=False)
    if store_id:
        batch_qs = batch_qs.filter(store_id=store_id)
    if category_id:
        batch_qs = batch_qs.filter(product__category_id=category_id)

    expired_count      = batch_qs.filter(expiry_date__lt=today).count()
    critical_count     = batch_qs.filter(expiry_date__gte=today, expiry_date__lte=today + timedelta(days=30)).count()
    warning_count      = batch_qs.filter(expiry_date__gt=today + timedelta(days=30), expiry_date__lte=today + timedelta(days=60)).count()
    monitor_count      = batch_qs.filter(expiry_date__gt=today + timedelta(days=60), expiry_date__lte=today + timedelta(days=90)).count()
    expiring_soon_count = critical_count  # for KPI card

    total_expiry = critical_count + warning_count + monitor_count or 1
    critical_percentage = round(critical_count / total_expiry * 100)
    warning_percentage  = round(warning_count  / total_expiry * 100)
    monitor_percentage  = round(monitor_count  / total_expiry * 100)

    # ── Total inventory value ────────────────────────────────────────────────
    raw_value = InventoryBatch.objects.filter(remaining_quantity__gt=0)
    if store_id:
        raw_value = raw_value.filter(store_id=store_id)
    if category_id:
        raw_value = raw_value.filter(product__category_id=category_id)
    total_inventory_value = raw_value.aggregate(
        val=Sum(F('remaining_quantity') * F('unit_cost'))
    )['val'] or Decimal('0')

    # ── Build inventory_data rows ────────────────────────────────────────────
    # Apply expiry filter at the row level after we pull batch data
    inventory_data = []

    for inv in inv_qs.order_by('product__name', 'store__name'):
        product = inv.product
        store   = inv.store

        # Batches for this product+store
        p_batches = InventoryBatch.objects.filter(
            product=product, store=store, remaining_quantity__gt=0
        ).order_by('expiry_date')

        # Committed stock (pending/in-transit transfers out)
        committed = StockTransferItem.objects.filter(
            product=product,
            stock_transfer__from_store=store,
            stock_transfer__status__in=['pending', 'in_transit']
        ).aggregate(c=Sum('base_quantity'))['c'] or 0

        # Unit cost & total value from batches
        avg_cost = p_batches.aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
        total_value = float(avg_cost * inv.quantity_in_stock)

        # Stock status
        if inv.quantity_in_stock == 0:
            stock_status_text = 'Out of Stock'
            status_class      = 'critical'
            row_class         = 'table-danger'
        elif inv.quantity_in_stock <= inv.reorder_level:
            stock_status_text = 'Low Stock'
            status_class      = 'warning'
            row_class         = 'table-warning'
        elif inv.quantity_in_stock > inv.reorder_level * 2:
            stock_status_text = 'Overstock'
            status_class      = 'info'
            row_class         = ''
        else:
            stock_status_text = 'Healthy'
            status_class      = 'good'
            row_class         = ''

        days_of_stock = int(inv.quantity_in_stock / 10 * 30)  # rough estimate

        # Nearest expiry from batches
        nearest_batch = p_batches.filter(expiry_date__isnull=False).first()
        days_to_expiry   = None
        expiry_status    = None
        expiry_class     = None
        expiry_progress  = None
        expiring_value   = None

        if nearest_batch and nearest_batch.expiry_date:
            days_to_expiry = (nearest_batch.expiry_date - today).days
            if days_to_expiry < 0:
                expiry_status = 'Expired'
                expiry_class  = 'critical'
            elif days_to_expiry <= 30:
                expiry_status = 'Critical'
                expiry_class  = 'critical'
            elif days_to_expiry <= 60:
                expiry_status = 'Warning'
                expiry_class  = 'warning'
            elif days_to_expiry <= 90:
                expiry_status = 'Monitor'
                expiry_class  = 'info'
            else:
                expiry_status = 'Safe'
                expiry_class  = 'good'

            expiry_progress = min(100, max(0, int(days_to_expiry / 365 * 100)))

            at_risk_batches = p_batches.filter(expiry_date__lte=today + timedelta(days=30))
            expiring_value = float(
                at_risk_batches.aggregate(
                    v=Sum(F('remaining_quantity') * F('unit_cost'))
                )['v'] or 0
            )

        # Apply expiry filter
        if expiry_filter:
            if expiry_filter == 'expired'  and (days_to_expiry is None or days_to_expiry >= 0):
                continue
            elif expiry_filter == 'critical' and (days_to_expiry is None or days_to_expiry < 0 or days_to_expiry > 30):
                continue
            elif expiry_filter == 'warning' and (days_to_expiry is None or days_to_expiry <= 30 or days_to_expiry > 60):
                continue
            elif expiry_filter == 'monitor' and (days_to_expiry is None or days_to_expiry <= 60 or days_to_expiry > 90):
                continue
            elif expiry_filter == 'safe'   and (days_to_expiry is None or days_to_expiry <= 90):
                continue

        # Last movement
        last_movement = StockMovement.objects.filter(
            product=product, store=store
        ).order_by('-timestamp').first()

        # Batch info list for table column
        batch_info = [
            {
                'batch_no':   b.id,
                'quantity':   b.remaining_quantity,
                'unit_cost':  float(b.unit_cost),
                'expiry_date': b.expiry_date,
            }
            for b in p_batches[:3]  # show max 3 batches inline
        ]

        # Value as % of total
        value_percentage = round(
            (total_value / float(total_inventory_value) * 100)
            if total_inventory_value else 0, 1
        )

        inventory_data.append({
            'sku':                product.sku,
            'product_name':       product.name,
            'brand':              product.brand,
            'category':           product.category.name if product.category else 'Uncategorized',
            'store_name':         store.name,
            'branch_name':        store.branch.name if store.branch else None,
            'current_stock':      inv.quantity_in_stock,
            'reorder_level':      inv.reorder_level,
            'committed_stock':    committed,
            'stock_status':       stock_status_text,
            'status_class':       status_class,
            'row_class':          row_class,
            'days_of_stock':      days_of_stock,
            'unit_cost':          float(avg_cost),
            'total_value':        total_value,
            'value_percentage':   value_percentage,
            'batch_count':        p_batches.count(),
            'has_multiple_batches': p_batches.count() > 1,
            'batch_info':         batch_info,
            'expiry_status':      expiry_status,
            'expiry_class':       expiry_class,
            'days_to_expiry':     days_to_expiry,
            'expiry_progress':    expiry_progress,
            'expiring_value':     expiring_value,
            'last_movement_date': last_movement.timestamp.date() if last_movement else None,
            'last_movement_type': last_movement.transaction_type if last_movement else None,
        })

    # ── Category summary card ────────────────────────────────────────────────
    categories = Category.objects.all()
    total_inv_count = inv_qs.count() or 1
    category_summary = []
    for cat in categories:
        count = inv_qs.filter(product__category=cat).count()
        if count:
            category_summary.append({
                'name':       cat.name,
                'count':      count,
                'percentage': round(count / total_inv_count * 100),
            })
    category_summary.sort(key=lambda x: x['count'], reverse=True)

    # ── Store summary card ───────────────────────────────────────────────────
    stores_qs = StoreLocation.objects.filter(is_active=True)
    if store_id:
        stores_qs = stores_qs.filter(id=store_id)

    store_summary = []
    for s in stores_qs:
        val = InventoryBatch.objects.filter(store=s, remaining_quantity__gt=0).aggregate(
            v=Sum(F('remaining_quantity') * F('unit_cost'))
        )['v'] or Decimal('0')
        store_summary.append({
            'name':       s.name,
            'value':      float(val),
            'percentage': round(float(val) / float(total_inventory_value) * 100
                                if total_inventory_value else 0),
        })
    store_summary.sort(key=lambda x: x['value'], reverse=True)

    # ── Recent movements ─────────────────────────────────────────────────────
    movements_qs = StockMovement.objects.select_related('product', 'store').order_by('-timestamp')
    if store_id:
        movements_qs = movements_qs.filter(store_id=store_id)
    recent_movements = movements_qs[:20]

    # ── Filter dropdowns ─────────────────────────────────────────────────────
    stores     = StoreLocation.objects.filter(is_active=True).select_related('branch').order_by('name')
    categories = Category.objects.order_by('name')

    context = {
        # KPI cards
        'total_inventory_value': total_inventory_value,
        'total_products':        total_products,
        'total_batches':         total_batches,
        'total_units':           total_units,
        'healthy_stock':         healthy_stock,
        'low_stock_pct':         low_stock_pct,
        'out_stock_pct':         out_stock_pct,
        'low_stock_items':       low_stock_items,
        'low_stock_count':       low_stock_count,
        'out_of_stock_count':    out_of_stock_count,
        'expiring_soon_count':   expiring_soon_count,
        'critical_count':        critical_count,
        'warning_count':         warning_count,
        'monitor_count':         monitor_count,

        # Expiry progress bars
        'critical_percentage':   critical_percentage,
        'warning_percentage':    warning_percentage,
        'monitor_percentage':    monitor_percentage,

        # Main table
        'inventory_data':   inventory_data,

        # Summary cards
        'category_summary': category_summary,
        'store_summary':    store_summary,

        # Recent movements
        'recent_movements': recent_movements,

        # Filter dropdowns
        'stores':     stores,
        'categories': categories,
    }

    return render(request, 'reports/inventory_details.html', context)



@login_required
def export_inventory_pdf(request):
    """Export comprehensive inventory report as PDF"""
    try:
        # Get filter parameters
        store_id = request.GET.get('store')
        category_id = request.GET.get('category')
        include_expiry = request.GET.get('include_expiry', 'true')
        include_movements = request.GET.get('include_movements', 'true')
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"inventory_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Get current date for report
        today = timezone.now().date()
        
        # Create PDF directly with response
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,  # Center
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,  # Center
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold',
            alignment=0,  # Left
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        small_style = ParagraphStyle(
            'SmallStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header Section
        elements.append(Paragraph("INVENTORY STATUS REPORT", title_style))
        
        # Store info
        store_text = "All Stores"
        if store_id:
            try:
                store = StoreLocation.objects.get(id=store_id)
                store_text = f"Store: {store.name}"
            except:
                pass
        elements.append(Paragraph(store_text, subtitle_style))
        
        # Date and generation info
        elements.append(Paragraph(f"As of: {today.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        
        # Horizontal line
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        # Base querysets
        inventories_qs = Inventory.objects.select_related('product', 'product__category', 'store')
        if store_id:
            inventories_qs = inventories_qs.filter(store_id=store_id)
        if category_id:
            inventories_qs = inventories_qs.filter(product__category_id=category_id)
        
        # Calculate metrics
        total_products = inventories_qs.values('product').distinct().count()
        total_items = inventories_qs.aggregate(total=Sum('quantity_in_stock'))['total'] or 0
        
        # Calculate total inventory value
        total_value = Decimal('0')
        product_values = {}
        for inv in inventories_qs:
            # Get average cost from batches
            avg_cost = InventoryBatch.objects.filter(
                product=inv.product,
                store=inv.store
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            
            item_value = inv.quantity_in_stock * avg_cost
            total_value += item_value
            product_values[inv.product.id] = {
                'quantity': inv.quantity_in_stock,
                'avg_cost': avg_cost,
                'value': item_value
            }
        
        # Store counts
        stores_count = inventories_qs.values('store').distinct().count()
        categories_count = inventories_qs.values('product__category').distinct().count()
        
        # Expiry calculations
        thirty_days_from_now = today + timedelta(days=30)
        expiring_soon_count = InventoryBatch.objects.filter(
            expiry_date__lte=thirty_days_from_now,
            expiry_date__gte=today,
            remaining_quantity__gt=0
        ).values('product').distinct().count()
        
        expired_count = InventoryBatch.objects.filter(
            expiry_date__lt=today,
            remaining_quantity__gt=0
        ).values('product').distinct().count()
        
        # Stock status
        low_stock_count = inventories_qs.filter(
            quantity_in_stock__lte=F('reorder_level'),
            quantity_in_stock__gt=0
        ).count()
        
        out_of_stock_count = inventories_qs.filter(quantity_in_stock=0).count()
        
        avg_cost_per_item = total_value / Decimal(total_items) if total_items > 0 else Decimal('0')
        
        # Summary table
        summary_data = [
            ['Metric', 'Value'],
            ['Total Products in Stock', f"{total_products:,}"],
            ['Total Items in Stock (Base Units)', f"{total_items:,} pcs"],
            ['Total Inventory Value', f"UGX {total_value:,.0f}"],
            ['Average Cost Per Item', f"UGX {avg_cost_per_item:,.0f}"],
            ['Stores with Inventory', f"{stores_count}"],
            ['Categories Represented', f"{categories_count}"],
            ['Expiring in 30 Days', f"{expiring_soon_count} products"],
            ['Expired Items', f"{expired_count} products"],
            ['Low Stock Items', f"{low_stock_count} products"],
            ['Out of Stock Items', f"{out_of_stock_count} products"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== INVENTORY BY STORE ==========
        elements.append(Paragraph("INVENTORY BY STORE", heading_style))
        elements.append(Spacer(1, 5))
        
        store_summary = inventories_qs.values(
            'store__id', 'store__name'
        ).annotate(
            products=Count('product', distinct=True),
            total_qty=Sum('quantity_in_stock')
        ).order_by('-total_qty')
        
        store_data = [['Store', 'Products', 'Total Quantity', 'Total Value', '% of Total']]
        
        for store in store_summary:
            # Calculate store value
            store_value = Decimal('0')
            store_inventories = inventories_qs.filter(store_id=store['store__id'])
            for inv in store_inventories:
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store_id=store['store__id']
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                store_value += inv.quantity_in_stock * avg_cost
            
            percent = (store_value / total_value * 100) if total_value > 0 else 0
            
            store_data.append([
                store['store__name'],
                str(store['products']),
                f"{store['total_qty']:,}",
                f"UGX {store_value:,.0f}",
                f"{percent:.1f}%"
            ])
        
        # Add total row
        store_data.append([
            'TOTAL',
            str(total_products),
            f"{total_items:,}",
            f"UGX {total_value:,.0f}",
            '100%'
        ])
        
        store_table = Table(store_data, colWidths=[120, 70, 100, 120, 70])
        store_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(store_table)
        elements.append(Spacer(1, 20))
        
        # ========== INVENTORY BY CATEGORY ==========
        elements.append(Paragraph("INVENTORY BY CATEGORY", heading_style))
        elements.append(Spacer(1, 5))
        
        category_summary = inventories_qs.values(
            'product__category__id', 'product__category__name'
        ).annotate(
            products=Count('product', distinct=True),
            total_qty=Sum('quantity_in_stock')
        ).order_by('-total_qty')
        
        category_data = [['Category', 'Products', 'Total Quantity', 'Total Value', '% of Value']]
        
        for cat in category_summary:
            cat_name = cat['product__category__name'] or 'Uncategorized'
            
            # Calculate category value
            cat_value = Decimal('0')
            cat_inventories = inventories_qs.filter(product__category_id=cat['product__category__id'])
            for inv in cat_inventories:
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                cat_value += inv.quantity_in_stock * avg_cost
            
            percent = (cat_value / total_value * 100) if total_value > 0 else 0
            
            category_data.append([
                cat_name,
                str(cat['products']),
                f"{cat['total_qty']:,}",
                f"UGX {cat_value:,.0f}",
                f"{percent:.1f}%"
            ])
        
        # Add total row
        category_data.append([
            'TOTAL',
            str(total_products),
            f"{total_items:,}",
            f"UGX {total_value:,.0f}",
            '100%'
        ])
        
        category_table = Table(category_data, colWidths=[120, 70, 100, 120, 70])
        category_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(category_table)
        elements.append(Spacer(1, 20))
        
        # ========== CURRENT INVENTORY LEVELS ==========
        elements.append(Paragraph("CURRENT INVENTORY LEVELS", heading_style))
        elements.append(Spacer(1, 5))
        
        inventory_data = [
            ['Product', 'SKU', 'Category', 'Store', 'In Stock', 'Unit Cost', 'Total Value', 'Status']
        ]
        
        # Get top 25 items by value
        top_inventories = []
        for inv in inventories_qs.select_related('product', 'product__category', 'store')[:25]:
            avg_cost = InventoryBatch.objects.filter(
                product=inv.product,
                store=inv.store
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            
            total_value = inv.quantity_in_stock * avg_cost
            
            # Determine status
            if inv.quantity_in_stock == 0:
                status = "Out of Stock 🔴"
            elif inv.quantity_in_stock <= inv.reorder_level:
                status = "Low Stock ⚠️"
            else:
                # Check for expiry
                expiring_batches = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store,
                    expiry_date__lte=today + timedelta(days=7),
                    remaining_quantity__gt=0
                ).exists()
                
                if expiring_batches:
                    status = "Expiring Soon ⚠️"
                else:
                    status = "Normal"
            
            inventory_data.append([
                inv.product.name[:20] + '...' if len(inv.product.name) > 20 else inv.product.name,
                inv.product.sku or 'N/A',
                inv.product.category.name[:12] if inv.product.category else 'Uncat',
                inv.store.name[:10] if inv.store else 'N/A',
                f"{inv.quantity_in_stock:,}",
                f"UGX {avg_cost:,.0f}",
                f"UGX {total_value:,.0f}",
                status
            ])
        
        # Add total row
        inventory_data.append([
            'TOTAL', '', '', '', f"{total_items:,}", f"UGX {avg_cost_per_item:,.0f}", f"UGX {total_value:,.0f}", ''
        ])
        
        inventory_table = Table(inventory_data, colWidths=[80, 45, 60, 50, 50, 60, 80, 60])
        inventory_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (4, 1), (6, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(inventory_table)
        elements.append(Spacer(1, 20))
        
        # ========== LOW STOCK ALERTS ==========
        elements.append(Paragraph("LOW STOCK ALERTS (Below Reorder Level)", heading_style))
        elements.append(Spacer(1, 5))
        
        low_stock_inventories = inventories_qs.filter(
            quantity_in_stock__lte=F('reorder_level'),
            quantity_in_stock__gt=0
        ).select_related('product', 'store')[:15]
        
        if low_stock_inventories:
            low_stock_data = [
                ['Product', 'SKU', 'Store', 'Current Stock', 'Reorder Level', 'Shortage', 'Unit Cost', 'Value at Risk']
            ]
            
            total_risk_value = Decimal('0')
            for inv in low_stock_inventories:
                shortage = inv.reorder_level - inv.quantity_in_stock
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                
                risk_value = shortage * avg_cost
                total_risk_value += risk_value
                
                low_stock_data.append([
                    inv.product.name[:15] + '...' if len(inv.product.name) > 15 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(inv.quantity_in_stock),
                    str(inv.reorder_level),
                    str(shortage),
                    f"UGX {avg_cost:,.0f}",
                    f"UGX {risk_value:,.0f}"
                ])
            
            # Add total row
            low_stock_data.append([
                'TOTAL', '', '', '', '', '', '', f"UGX {total_risk_value:,.0f}"
            ])
            
            low_stock_table = Table(low_stock_data, colWidths=[80, 45, 45, 45, 45, 40, 60, 70])
            low_stock_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(low_stock_table)
        else:
            elements.append(Paragraph("No low stock items found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== OUT OF STOCK ITEMS ==========
        elements.append(Paragraph("OUT OF STOCK ITEMS", heading_style))
        elements.append(Spacer(1, 5))
        
        out_of_stock = inventories_qs.filter(
            quantity_in_stock=0
        ).select_related('product', 'product__category', 'store')[:15]
        
        if out_of_stock:
            oos_data = [
                ['Product', 'SKU', 'Category', 'Store', 'Last Received', 'Days Out']
            ]
            
            for inv in out_of_stock:
                last_batch = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).order_by('-received_date').first()
                
                last_received = last_batch.received_date.date() if last_batch else 'Never'
                days_out = (today - last_received).days if last_batch and last_received != 'Never' else 'N/A'
                
                oos_data.append([
                    inv.product.name[:15] + '...' if len(inv.product.name) > 15 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.product.category.name[:12] if inv.product.category else 'Uncat',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(last_received) if last_received != 'Never' else 'Never',
                    str(days_out) if days_out != 'N/A' else 'N/A'
                ])
            
            oos_table = Table(oos_data, colWidths=[80, 50, 60, 50, 70, 50])
            oos_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(oos_table)
        else:
            elements.append(Paragraph("No out of stock items found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== EXPIRY TRACKING ==========
        if include_expiry == 'true':
            elements.append(Paragraph("EXPIRY TRACKING", heading_style))
            elements.append(Spacer(1, 5))
            
            # Expired Items
            elements.append(Paragraph("Expired Items (Immediate Action Required)", normal_style))
            elements.append(Spacer(1, 3))
            
            expired_batches = InventoryBatch.objects.filter(
                expiry_date__lt=today,
                remaining_quantity__gt=0
            ).select_related('product', 'store')[:10]
            
            if expired_batches:
                expired_data = [
                    ['Product', 'SKU', 'Store', 'Batch', 'Quantity', 'Expiry Date', 'Days Expired', 'Value Lost']
                ]
                
                total_expired_value = Decimal('0')
                for batch in expired_batches:
                    days_expired = (today - batch.expiry_date).days
                    value_lost = batch.remaining_quantity * batch.unit_cost
                    total_expired_value += value_lost
                    
                    expired_data.append([
                        batch.product.name[:15] + '...' if len(batch.product.name) > 15 else batch.product.name,
                        batch.product.sku or 'N/A',
                        batch.store.name[:8] if batch.store else 'N/A',
                        f"BATCH-{batch.id}",
                        str(batch.remaining_quantity),
                        batch.expiry_date.strftime('%Y-%m-%d'),
                        str(days_expired),
                        f"UGX {value_lost:,.0f}"
                    ])
                
                # Add total row
                expired_data.append([
                    'TOTAL', '', '', '', '', '', '', f"UGX {total_expired_value:,.0f}"
                ])
                
                expired_table = Table(expired_data, colWidths=[70, 40, 40, 45, 40, 60, 45, 60])
                expired_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ]))
                elements.append(expired_table)
            else:
                elements.append(Paragraph("No expired items found.", small_style))
            
            elements.append(Spacer(1, 10))
            
            # Expiring in Next 30 Days
            elements.append(Paragraph("Expiring in Next 30 Days", normal_style))
            elements.append(Spacer(1, 3))
            
            expiring_batches = InventoryBatch.objects.filter(
                expiry_date__lte=today + timedelta(days=30),
                expiry_date__gte=today,
                remaining_quantity__gt=0
            ).select_related('product', 'store').order_by('expiry_date')[:10]
            
            if expiring_batches:
                expiring_data = [
                    ['Product', 'SKU', 'Store', 'Batch', 'Quantity', 'Expiry Date', 'Days Left', 'Value at Risk']
                ]
                
                total_risk_value = Decimal('0')
                for batch in expiring_batches:
                    days_left = (batch.expiry_date - today).days
                    value_at_risk = batch.remaining_quantity * batch.unit_cost
                    total_risk_value += value_at_risk
                    
                    expiring_data.append([
                        batch.product.name[:15] + '...' if len(batch.product.name) > 15 else batch.product.name,
                        batch.product.sku or 'N/A',
                        batch.store.name[:8] if batch.store else 'N/A',
                        f"BATCH-{batch.id}",
                        str(batch.remaining_quantity),
                        batch.expiry_date.strftime('%Y-%m-%d'),
                        str(days_left),
                        f"UGX {value_at_risk:,.0f}"
                    ])
                
                # Add total row
                expiring_data.append([
                    'TOTAL', '', '', '', '', '', '', f"UGX {total_risk_value:,.0f}"
                ])
                
                expiring_table = Table(expiring_data, colWidths=[70, 40, 40, 45, 40, 60, 40, 60])
                expiring_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
                ]))
                elements.append(expiring_table)
            else:
                elements.append(Paragraph("No items expiring in next 30 days.", small_style))
            
            elements.append(Spacer(1, 20))
        
        # ========== BATCH-WISE INVENTORY DETAIL ==========
        elements.append(Paragraph("BATCH-WISE INVENTORY DETAIL", heading_style))
        elements.append(Spacer(1, 5))
        
        recent_batches = InventoryBatch.objects.filter(
            remaining_quantity__gt=0
        ).select_related('product', 'store').order_by('-received_date')[:20]
        
        if recent_batches:
            batch_data = [
                ['Batch ID', 'Product', 'Store', 'Received', 'Qty', 'Remaining', 'Unit Cost', 'Total Value', 'Expiry Date']
            ]
            
            total_batch_value = Decimal('0')
            for batch in recent_batches:
                total_value = batch.remaining_quantity * batch.unit_cost
                total_batch_value += total_value
                
                batch_data.append([
                    f"BATCH-{batch.id}",
                    batch.product.name[:12] + '...' if len(batch.product.name) > 12 else batch.product.name,
                    batch.store.name[:8] if batch.store else 'N/A',
                    batch.received_date.strftime('%Y-%m-%d') if batch.received_date else 'N/A',
                    str(batch.quantity),
                    str(batch.remaining_quantity),
                    f"UGX {batch.unit_cost:,.0f}",
                    f"UGX {total_value:,.0f}",
                    batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A'
                ])
            
            # Add total row
            batch_data.append([
                'TOTAL', '', '', '', '', f"{total_items}", f"UGX {avg_cost_per_item:,.0f}", f"UGX {total_batch_value:,.0f}", ''
            ])
            
            batch_table = Table(batch_data, colWidths=[55, 70, 40, 60, 35, 45, 55, 70, 60])
            batch_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(batch_table)
        else:
            elements.append(Paragraph("No batch data available.", normal_style))
        
        elements.append(Spacer(1, 20))
        
        # ========== RECOMMENDATIONS ==========
        elements.append(Paragraph("RECOMMENDATIONS", heading_style))
        elements.append(Spacer(1, 5))
        
        recommendations = [
            ['Priority', 'Action', 'Affected Items', 'Estimated Impact'],
            ['High', 'Reorder out-of-stock items', f"{out_of_stock_count} products", 'UGX 2,500,000 potential sales'],
            ['High', 'Mark down expiring items', f"{expiring_soon_count} products", f"Prevent UGX {total_risk_value:,.0f} loss" if 'total_risk_value' in locals() else 'Prevent losses'],
            ['Medium', 'Review reorder levels', f"{low_stock_count} products", 'Prevent stockouts'],
            ['Medium', 'Investigate slow-moving stock', 'Review inventory turnover', 'Optimize working capital'],
            ['Low', 'Category mix optimization', 'All categories', 'Improve turnover'],
        ]
        
        rec_table = Table(recommendations, colWidths=[80, 150, 100, 150])
        rec_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(rec_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response



# ============================================================================
# TRANSFER & MOVEMENT REPORTS VIEWS
# ============================================================================
# Inter-store Transfer Summary, Transfer Request Status, Transfer History Audit,
# Department-wise Transfers
# ============================================================================


@login_required
def transfer_details(request):
    """
    Transfer & Movement Reports view.
    Matches: reports/transfer_details.html
    """
    today = timezone.now().date()

    # ── Filters ──────────────────────────────────────────────────────────────
    from_store_id = request.GET.get('from_store', '')
    to_store_id   = request.GET.get('to_store', '')
    status_param  = request.GET.get('status', '')
    date_range    = request.GET.get('date_range', '')

    # Default: no date restriction (show all). Only restrict when user picks a range.
    start_date = None
    end_date   = None

    if date_range:
        try:
            parts = [p.strip() for p in date_range.split(' - ')]
            if len(parts) == 2:
                start_date = date.fromisoformat(parts[0])
                end_date   = date.fromisoformat(parts[1])
        except (ValueError, AttributeError):
            pass

    # ── StockTransfer queryset ────────────────────────────────────────────────
    st_qs = StockTransfer.objects.select_related(
        'from_store', 'to_store', 'created_by',
        'transfer_request', 'transfer_request__requested_by',
        'transfer_request__approved_by', 'transfer_request__department',
    )

    if start_date and end_date:
        st_qs = st_qs.filter(transfer_date__range=[start_date, end_date])
    if from_store_id:
        st_qs = st_qs.filter(from_store_id=from_store_id)
    if to_store_id:
        st_qs = st_qs.filter(to_store_id=to_store_id)
    if status_param:
        st_qs = st_qs.filter(status__iexact=status_param)

    st_qs = st_qs.order_by('-created_at')

    # ── TransferRequest queryset (standalone requests not yet executed) ────────
    req_qs = TransferRequest.objects.select_related(
        'from_store', 'to_store', 'requested_by', 'approved_by', 'department'
    )

    if start_date and end_date:
        req_qs = req_qs.filter(request_date__date__range=[start_date, end_date])
    if from_store_id:
        req_qs = req_qs.filter(from_store_id=from_store_id)
    if to_store_id:
        req_qs = req_qs.filter(to_store_id=to_store_id)
    if status_param:
        req_qs = req_qs.filter(status__iexact=status_param)

    req_qs = req_qs.order_by('-request_date')

    # ── Build unified transfers_data list ─────────────────────────────────────
    st_request_ids = set(
        st_qs.exclude(transfer_request=None)
             .values_list('transfer_request_id', flat=True)
    )
    standalone_requests = req_qs.exclude(id__in=st_request_ids)
    transfers_data = list(st_qs) + list(standalone_requests)

    # ── KPI metrics ───────────────────────────────────────────────────────────
    total_transfers  = st_qs.count()
    active_transfers = st_qs.filter(status__in=['pending', 'in_transit']).count()

    total_items_moved = StockTransferItem.objects.filter(
        stock_transfer__in=st_qs
    ).aggregate(total=Sum('base_quantity'))['total'] or 0

    all_st_items = StockTransferItem.objects.filter(
        stock_transfer__in=st_qs
    ).select_related('product', 'units')
    total_value = sum(item.total_value for item in all_st_items)
    avg_transfer_value = (total_value / total_transfers) if total_transfers else 0

    pending_requests = TransferRequest.objects.filter(status='pending').count()
    urgent_requests  = TransferRequest.objects.filter(
        status='pending', priority='urgent'
    ).count()

    # ── Status counts ─────────────────────────────────────────────────────────
    status_counts = {
        'pending':    st_qs.filter(status='pending').count()
                      + req_qs.filter(status='pending').count(),
        'approved':   req_qs.filter(status='approved').count(),
        'in_transit': st_qs.filter(status='in_transit').count(),
        'completed':  st_qs.filter(status='completed').count(),
    }

    # ── Priority counts ───────────────────────────────────────────────────────
    priority_counts = {
        'urgent': req_qs.filter(priority='urgent').count(),
        'high':   req_qs.filter(priority='high').count(),
        'normal': req_qs.filter(priority='normal').count(),
    }

    # ── Top routes ────────────────────────────────────────────────────────────
    top_routes = [
        {
            'from_store': r['from_store__name'],
            'to_store':   r['to_store__name'],
            'count':      r['count'],
        }
        for r in st_qs
            .values('from_store__name', 'to_store__name')
            .annotate(count=Count('id'))
            .order_by('-count')[:5]
    ]

    # ── Today's activity ──────────────────────────────────────────────────────
    today_transfers = StockTransfer.objects.filter(transfer_date=today).count()
    today_items     = StockTransferItem.objects.filter(
        stock_transfer__transfer_date=today
    ).aggregate(total=Sum('base_quantity'))['total'] or 0

    completed_today = StockTransfer.objects.filter(
        transfer_date=today, status='completed'
    ).count()
    completion_rate = round(
        (completed_today / today_transfers * 100) if today_transfers else 0
    )

    # ── Filter dropdowns ──────────────────────────────────────────────────────
    stores = StoreLocation.objects.filter(is_active=True).order_by('name')

    # For the date display in the template — show actual range or sensible defaults
    display_start = start_date or StockTransfer.objects.order_by('transfer_date').values_list('transfer_date', flat=True).first() or today
    display_end   = end_date or today

    context = {
        'start_date': display_start,
        'end_date':   display_end,

        # KPI cards
        'total_transfers':    total_transfers,
        'active_transfers':   active_transfers,
        'total_items_moved':  total_items_moved,
        'total_value':        total_value,
        'avg_transfer_value': avg_transfer_value,
        'pending_requests':   pending_requests,
        'urgent_requests':    urgent_requests,

        # Main table
        'transfers_data': transfers_data,

        # Summary cards
        'status_counts':   status_counts,
        'priority_counts': priority_counts,
        'top_routes':      top_routes,

        # Today
        'today_transfers': today_transfers,
        'today_items':     today_items,
        'completion_rate': completion_rate,

        # Filter dropdowns
        'stores': stores,
    }

    return render(request, 'reports/transfer_details.html', context)


@login_required
def export_transfer_pdf(request):
    """Export comprehensive stock transfer report as PDF"""
    try:
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        from_store_id = request.GET.get('from_store')
        to_store_id = request.GET.get('to_store')
        status = request.GET.get('status')
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"transfer_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Parse dates
        today = timezone.now().date()
        if date_from and date_to:
            try:
                start_date = date.fromisoformat(date_from)
                end_date = date.fromisoformat(date_to)
            except:
                start_date = today.replace(day=1)
                end_date = today
        else:
            start_date = today.replace(day=1)
            end_date = today
        
        # Create PDF directly with response
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        small_style = ParagraphStyle(
            'SmallStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header Section
        elements.append(Paragraph("STOCK TRANSFER REPORT", title_style))
        elements.append(Paragraph(f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Base queryset
        transfers_qs = StockTransfer.objects.filter(
            transfer_date__range=[start_date, end_date]
        ).select_related('from_store', 'to_store', 'created_by')
        
        if from_store_id:
            transfers_qs = transfers_qs.filter(from_store_id=from_store_id)
        if to_store_id:
            transfers_qs = transfers_qs.filter(to_store_id=to_store_id)
        if status:
            transfers_qs = transfers_qs.filter(status=status)
        
        # Transfer requests queryset
        requests_qs = TransferRequest.objects.filter(
            request_date__date__range=[start_date, end_date]
        ).select_related('from_store', 'to_store', 'requested_by', 'department')
        
        # Calculate metrics manually instead of using aggregates with F expressions
        total_transfers = transfers_qs.count()
        total_items = 0
        total_value = Decimal('0')
        
        # Get all transfer items and calculate values
        transfer_items = StockTransferItem.objects.filter(
            stock_transfer__in=transfers_qs
        ).select_related('product', 'stock_transfer')
        
        # Calculate totals by iterating
        for item in transfer_items:
            total_items += item.base_quantity
            # Get unit cost from product's purchase history or default
            avg_cost = InventoryBatch.objects.filter(
                product=item.product
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            total_value += item.base_quantity * avg_cost
        
        avg_transfer_size = total_items / total_transfers if total_transfers > 0 else 0
        stores_involved = transfers_qs.values('from_store').distinct().count() + \
                         transfers_qs.values('to_store').distinct().count()
        products_transferred = transfer_items.values('product').distinct().count()
        
        # Status counts
        status_counts = {
            'completed': transfers_qs.filter(status='completed').count(),
            'in_transit': transfers_qs.filter(status='in_transit').count(),
            'pending': transfers_qs.filter(status='pending').count(),
            'cancelled': transfers_qs.filter(status='cancelled').count(),
        }
        
        # Calculate status items and values manually
        status_items = {}
        status_values = {}
        
        for status_name in ['completed', 'in_transit', 'pending', 'cancelled']:
            status_transfers = transfers_qs.filter(status=status_name)
            status_items[status_name] = 0
            status_values[status_name] = Decimal('0')
            
            status_transfer_items = StockTransferItem.objects.filter(
                stock_transfer__in=status_transfers
            )
            
            for item in status_transfer_items:
                status_items[status_name] += item.base_quantity
                avg_cost = InventoryBatch.objects.filter(
                    product=item.product
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                status_values[status_name] += item.base_quantity * avg_cost
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Transfers', f"{total_transfers}"],
            ['Total Items Transferred', f"{total_items:,} pcs (base units)"],
            ['Total Transfer Value', f"UGX {total_value:,.0f}"],
            ['Average Transfer Size', f"{avg_transfer_size:.0f} pcs"],
            ['Stores Involved', f"{stores_involved}"],
            ['Products Transferred', f"{products_transferred}"],
            ['Pending Transfers', f"{status_counts['pending']}"],
            ['Completed Transfers', f"{status_counts['completed']}"],
            ['In Transit', f"{status_counts['in_transit']}"],
            ['Cancelled', f"{status_counts['cancelled']}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== TRANSFERS BY STATUS ==========
        elements.append(Paragraph("TRANSFERS BY STATUS", heading_style))
        elements.append(Spacer(1, 5))
        
        status_data = [['Status', 'Count', 'Items Transferred', 'Total Value', '% of Total']]
        
        for status_name, count in status_counts.items():
            if count > 0:
                percent = (status_values[status_name] / total_value * 100) if total_value > 0 else 0
                status_data.append([
                    status_name.capitalize().replace('_', ' '),
                    str(count),
                    f"{status_items[status_name]:,}",
                    f"UGX {status_values[status_name]:,.0f}",
                    f"{percent:.1f}%"
                ])
        
        status_data.append(['TOTAL', str(total_transfers), f"{total_items:,}", f"UGX {total_value:,.0f}", '100%'])
        
        status_table = Table(status_data, colWidths=[100, 70, 100, 120, 70])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # ========== TRANSFERS BY STORE ROUTE ==========
        elements.append(Paragraph("TRANSFERS BY STORE", heading_style))
        elements.append(Spacer(1, 5))
        
        # Calculate route summaries manually
        route_data_dict = {}
        for transfer in transfers_qs:
            from_store = transfer.from_store.name if transfer.from_store else 'Unknown'
            to_store = transfer.to_store.name if transfer.to_store else 'Unknown'
            key = f"{from_store}→{to_store}"
            
            if key not in route_data_dict:
                route_data_dict[key] = {
                    'from': from_store,
                    'to': to_store,
                    'transfers': 0,
                    'items': 0,
                    'value': Decimal('0')
                }
            
            route_data_dict[key]['transfers'] += 1
            
            # Get items for this transfer
            transfer_items_for_transfer = StockTransferItem.objects.filter(stock_transfer=transfer)
            for item in transfer_items_for_transfer:
                route_data_dict[key]['items'] += item.base_quantity
                avg_cost = InventoryBatch.objects.filter(
                    product=item.product
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                route_data_dict[key]['value'] += item.base_quantity * avg_cost
        
        route_data = [['From Store', 'To Store', 'Transfers', 'Items', 'Value']]
        
        for route in route_data_dict.values():
            route_data.append([
                route['from'],
                route['to'],
                str(route['transfers']),
                f"{route['items']:,}",
                f"UGX {route['value']:,.0f}"
            ])
        
        route_data.append(['TOTAL', '', str(total_transfers), f"{total_items:,}", f"UGX {total_value:,.0f}"])
        
        route_table = Table(route_data, colWidths=[100, 100, 70, 80, 120])
        route_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(route_table)
        elements.append(Spacer(1, 20))
        
        # ========== NET TRANSFER FLOW ==========
        elements.append(Paragraph("NET TRANSFER FLOW", heading_style))
        elements.append(Spacer(1, 5))
        
        stores = set()
        for route in route_data_dict.values():
            stores.add(route['from'])
            stores.add(route['to'])
        
        flow_data = [['Store', 'Sent Out', 'Received In', 'Net Change']]
        
        for store_name in sorted(stores):
            sent_out_items = 0
            sent_out_value = Decimal('0')
            received_items = 0
            received_value = Decimal('0')
            
            for route in route_data_dict.values():
                if route['from'] == store_name:
                    sent_out_items += route['items']
                    sent_out_value += route['value']
                if route['to'] == store_name:
                    received_items += route['items']
                    received_value += route['value']
            
            net_items = received_items - sent_out_items
            net_text = f"{'+' if net_items > 0 else ''}{net_items:,} pcs ({'Incoming' if net_items > 0 else 'Outgoing' if net_items < 0 else 'Balanced'})"
            
            flow_data.append([
                store_name,
                f"{sent_out_items:,} pcs (UGX {sent_out_value:,.0f})",
                f"{received_items:,} pcs (UGX {received_value:,.0f})",
                net_text
            ])
        
        flow_table = Table(flow_data, colWidths=[100, 150, 150, 150])
        flow_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(flow_table)
        elements.append(Spacer(1, 20))
        
        # ========== DETAILED TRANSFER REQUESTS ==========
        elements.append(Paragraph("DETAILED TRANSFER REQUESTS", heading_style))
        elements.append(Spacer(1, 5))
        
        recent_requests = requests_qs.order_by('-request_date')[:15]
        
        if recent_requests:
            request_data = [
                ['Request ID', 'Date', 'From', 'To', 'Priority', 'Items', 'Value', 'Status', 'Requested By']
            ]
            
            for req in recent_requests:
                priority_display = req.priority.capitalize() if req.priority else 'Normal'
                if req.priority == 'urgent':
                    priority_display = 'Urgent'
                elif req.priority == 'high':
                    priority_display = 'High'
                
                request_data.append([
                    f"REQ-{req.id}",
                    req.request_date.strftime('%Y-%m-%d') if req.request_date else '',
                    req.from_store.name[:10] if req.from_store else 'N/A',
                    req.to_store.name[:10] if req.to_store else 'N/A',
                    priority_display,
                    str(req.total_requested_items),
                    f"UGX {req.total_estimated_value:,.0f}",
                    req.status.capitalize(),
                    req.requested_by.get_full_name()[:8] if req.requested_by else 'N/A'
                ])
            
            request_table = Table(request_data, colWidths=[70, 60, 50, 50, 50, 40, 80, 60, 60])
            request_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (5, 1), (6, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(request_table)
        else:
            elements.append(Paragraph("No transfer requests found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== TRANSFERS BY PRIORITY ==========
        elements.append(Paragraph("TRANSFERS BY PRIORITY", heading_style))
        elements.append(Spacer(1, 5))
        
        priority_counts = {
            'urgent': requests_qs.filter(priority='urgent').count(),
            'high': requests_qs.filter(priority='high').count(),
            'normal': requests_qs.filter(priority='normal').count(),
        }
        
        priority_data = [['Priority', 'Count', 'Items', 'Value', 'Avg Processing Time']]
        
        for priority, count in priority_counts.items():
            if count > 0:
                priority_requests = requests_qs.filter(priority=priority)
                priority_items = sum(req.total_requested_items for req in priority_requests)
                priority_value = sum(req.total_estimated_value for req in priority_requests)
                
                # Calculate avg processing time (simplified)
                if priority == 'urgent':
                    avg_time = '4 hours'
                elif priority == 'high':
                    avg_time = '1 day'
                else:
                    avg_time = '3 days'
                
                priority_data.append([
                    priority.capitalize(),
                    str(count),
                    f"{priority_items:,}",
                    f"UGX {priority_value:,.0f}",
                    avg_time
                ])
        
        priority_data.append(['TOTAL', str(requests_qs.count()), f"{sum(req.total_requested_items for req in requests_qs):,}", f"UGX {sum(req.total_estimated_value for req in requests_qs):,}", ''])
        
        priority_table = Table(priority_data, colWidths=[80, 60, 80, 120, 100])
        priority_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(priority_table)
        elements.append(Spacer(1, 20))
        
        # ========== TRANSFER ITEMS DETAIL ==========
        elements.append(Paragraph("TRANSFER ITEMS DETAIL", heading_style))
        elements.append(Spacer(1, 5))
        
        transfer_items_list = StockTransferItem.objects.filter(
            stock_transfer__in=transfers_qs
        ).select_related('product', 'units', 'stock_transfer', 'stock_transfer__from_store', 'stock_transfer__to_store')[:20]
        
        if transfer_items_list:
            items_data = [
                ['Request ID', 'Product', 'SKU', 'From', 'To', 'Qty', 'Unit', 'Base Qty', 'Unit Cost', 'Line Total']
            ]
            
            total_line_value = Decimal('0')
            for item in transfer_items_list:
                avg_cost = InventoryBatch.objects.filter(
                    product=item.product
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                line_total = item.base_quantity * avg_cost
                total_line_value += line_total
                
                items_data.append([
                    f"REQ-{item.stock_transfer.transfer_request.id if item.stock_transfer and item.stock_transfer.transfer_request else 'N/A'}",
                    item.product.name[:12] + '...' if len(item.product.name) > 12 else item.product.name,
                    item.product.sku or 'N/A',
                    item.stock_transfer.from_store.name[:8] if item.stock_transfer and item.stock_transfer.from_store else 'N/A',
                    item.stock_transfer.to_store.name[:8] if item.stock_transfer and item.stock_transfer.to_store else 'N/A',
                    str(item.quantity),
                    item.units.abbreviation if item.units else 'unit',
                    str(item.base_quantity),
                    f"UGX {avg_cost:,.0f}",
                    f"UGX {line_total:,.0f}"
                ])
            
            items_data.append([
                'TOTAL', '', '', '', '', '', '', f"{total_items}", f"UGX {(total_value/total_items):,.0f}" if total_items > 0 else 'N/A', f"UGX {total_value:,.0f}"
            ])
            
            items_table = Table(items_data, colWidths=[60, 70, 40, 35, 35, 30, 30, 40, 50, 70])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (5, 1), (9, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(items_table)
        else:
            elements.append(Paragraph("No transfer items found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== TOP TRANSFERRED PRODUCTS ==========
        elements.append(Paragraph("TOP TRANSFERRED PRODUCTS", heading_style))
        elements.append(Spacer(1, 5))
        
        # Calculate top products manually
        product_totals = {}
        for item in transfer_items:
            if item.product.id not in product_totals:
                product_totals[item.product.id] = {
                    'name': item.product.name,
                    'sku': item.product.sku,
                    'total_qty': 0,
                    'total_value': Decimal('0'),
                    'transfer_count': 0
                }
            
            product_totals[item.product.id]['total_qty'] += item.base_quantity
            avg_cost = InventoryBatch.objects.filter(
                product=item.product
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            product_totals[item.product.id]['total_value'] += item.base_quantity * avg_cost
            product_totals[item.product.id]['transfer_count'] += 1
        
        # Sort by quantity and get top 10
        top_products_list = sorted(product_totals.values(), key=lambda x: x['total_qty'], reverse=True)[:10]
        
        if top_products_list:
            products_data = [['Rank', 'Product', 'SKU', 'Total Qty', 'Total Value', 'Transfer Count']]
            
            for i, product in enumerate(top_products_list, 1):
                products_data.append([
                    str(i),
                    product['name'][:20] + '...' if len(product['name']) > 20 else product['name'],
                    product['sku'] or 'N/A',
                    f"{product['total_qty']:,}",
                    f"UGX {product['total_value']:,.0f}",
                    str(product['transfer_count'])
                ])
            
            products_table = Table(products_data, colWidths=[40, 120, 60, 80, 120, 80])
            products_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                ('ALIGN', (3, 1), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(products_table)
        else:
            elements.append(Paragraph("No product data available.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== STOCK TRANSFER (Executed) ==========
        elements.append(Paragraph("STOCK TRANSFER (Executed)", heading_style))
        elements.append(Spacer(1, 5))
        
        executed_transfers = transfers_qs.select_related(
            'from_store', 'to_store', 'created_by'
        ).order_by('-transfer_date')[:15]
        
        if executed_transfers:
            exec_data = [
                ['Transfer ID', 'Date', 'From', 'To', 'Items', 'Value', 'Status', 'Completed By']
            ]
            
            for transfer in executed_transfers:
                # Calculate items and value for this transfer
                transfer_items_for_transfer = StockTransferItem.objects.filter(stock_transfer=transfer)
                transfer_items_count = sum(item.base_quantity for item in transfer_items_for_transfer)
                transfer_value = Decimal('0')
                for item in transfer_items_for_transfer:
                    avg_cost = InventoryBatch.objects.filter(
                        product=item.product
                    ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                    transfer_value += item.base_quantity * avg_cost
                
                exec_data.append([
                    f"ST-{transfer.id:04d}",
                    transfer.transfer_date.strftime('%Y-%m-%d') if transfer.transfer_date else '',
                    transfer.from_store.name[:10] if transfer.from_store else 'N/A',
                    transfer.to_store.name[:10] if transfer.to_store else 'N/A',
                    str(transfer_items_count),
                    f"UGX {transfer_value:,.0f}",
                    transfer.status.capitalize().replace('_', ' '),
                    transfer.completed_by or (transfer.created_by.get_full_name()[:10] if transfer.created_by else 'N/A')
                ])
            
            exec_data.append([
                'TOTAL', '', '', '', f"{total_items}", f"UGX {total_value:,.0f}", '', ''
            ])
            
            exec_table = Table(exec_data, colWidths=[70, 65, 60, 60, 50, 100, 70, 70])
            exec_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(exec_table)
        else:
            elements.append(Paragraph("No executed transfers found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== STAFF TRANSFER ACTIVITY ==========
        elements.append(Paragraph("STAFF TRANSFER ACTIVITY", heading_style))
        elements.append(Spacer(1, 5))
        
        # Calculate staff activity manually
        staff_data_dict = {}
        for transfer in transfers_qs:
            if transfer.created_by:
                staff_id = transfer.created_by.id
                if staff_id not in staff_data_dict:
                    staff_data_dict[staff_id] = {
                        'name': transfer.created_by.get_full_name() or transfer.created_by.username,
                        'transfers_created': 0,
                        'transfers_completed': 0,
                        'items_moved': 0,
                        'value_processed': Decimal('0')
                    }
                
                staff_data_dict[staff_id]['transfers_created'] += 1
                if transfer.status == 'completed':
                    staff_data_dict[staff_id]['transfers_completed'] += 1
                
                # Get items for this transfer
                transfer_items_for_transfer = StockTransferItem.objects.filter(stock_transfer=transfer)
                for item in transfer_items_for_transfer:
                    staff_data_dict[staff_id]['items_moved'] += item.base_quantity
                    avg_cost = InventoryBatch.objects.filter(
                        product=item.product
                    ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                    staff_data_dict[staff_id]['value_processed'] += item.base_quantity * avg_cost
        
        staff_data = [['Staff', 'Requests Created', 'Transfers Completed', 'Items Moved', 'Value Processed']]
        
        for staff in staff_data_dict.values():
            staff_data.append([
                staff['name'],
                str(staff['transfers_created']),
                str(staff['transfers_completed']),
                f"{staff['items_moved']:,}",
                f"UGX {staff['value_processed']:,.0f}"
            ])
        
        staff_data.append([
            'TOTAL',
            str(requests_qs.count()),
            str(status_counts['completed']),
            f"{total_items:,}",
            f"UGX {total_value:,.0f}"
        ])
        
        staff_table = Table(staff_data, colWidths=[100, 80, 80, 80, 120])
        staff_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(staff_table)
        elements.append(Spacer(1, 20))
        
        # ========== TRANSFER EFFICIENCY METRICS ==========
        elements.append(Paragraph("TRANSFER EFFICIENCY METRICS", heading_style))
        elements.append(Spacer(1, 5))
        
        approval_rate = (requests_qs.filter(status='approved').count() / requests_qs.count() * 100) if requests_qs.count() > 0 else 0
        fulfillment_rate = (requests_qs.filter(status='fulfilled').count() / requests_qs.count() * 100) if requests_qs.count() > 0 else 0
        
        # Find most active day
        day_counts = {}
        for transfer in transfers_qs:
            day_name = transfer.transfer_date.strftime('%A') if transfer.transfer_date else 'Unknown'
            day_counts[day_name] = day_counts.get(day_name, 0) + 1
        
        most_active_day = max(day_counts.items(), key=lambda x: x[1])[0] if day_counts else 'N/A'
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Transfer Approval Rate', f"{approval_rate:.1f}%"],
            ['Transfer Fulfillment Rate', f"{fulfillment_rate:.1f}%"],
            ['Average Items per Transfer', f"{avg_transfer_size:.0f}"],
            ['Average Value per Transfer', f"UGX {(total_value/total_transfers):,.0f}" if total_transfers > 0 else 'UGX 0'],
            ['Most Active Transfer Day', most_active_day],
            ['Peak Transfer Hour', '10:00 - 11:00 AM'],  # Static for now
        ]
        
        metrics_table = Table(metrics_data, colWidths=[200, 200])
        metrics_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(metrics_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response



# ============================================================================
# STOCK ADJUSTMENT REPORTS VIEWS
# ============================================================================
# Adjustment History, Reason-wise Adjustments, User-wise Adjustments,
# Batch Adjustment Summary
# ============================================================================


@login_required
def stockadj_details(request):
    """
    Stock Adjustment Reports view.
    Matches: reports/stockadj_details.html
    """
    today = timezone.now().date()

    # ── Filters ──────────────────────────────────────────────────────────────
    store_id        = request.GET.get('store_id', '')
    status_param    = request.GET.get('status', '')
    type_param      = request.GET.get('type', '')       # 'increase' | 'decrease'
    reason_param    = request.GET.get('reason', '')
    start_date_raw  = request.GET.get('start_date', '')
    end_date_raw    = request.GET.get('end_date', '')

    start_date = None
    end_date   = None

    if start_date_raw:
        try:
            start_date = date.fromisoformat(start_date_raw)
        except (ValueError, AttributeError):
            pass
    if end_date_raw:
        try:
            end_date = date.fromisoformat(end_date_raw)
        except (ValueError, AttributeError):
            pass

    # ── Base queryset ─────────────────────────────────────────────────────────
    adj_qs = StockAdjustment.objects.select_related(
        'product', 'product__category', 'store', 'created_by'
    )

    if start_date:
        adj_qs = adj_qs.filter(created_at__date__gte=start_date)
    if end_date:
        adj_qs = adj_qs.filter(created_at__date__lte=end_date)
    if store_id:
        adj_qs = adj_qs.filter(store_id=store_id)
    if status_param:
        adj_qs = adj_qs.filter(status__iexact=status_param)
    if type_param == 'increase':
        adj_qs = adj_qs.filter(quantity_change__gt=0)
    elif type_param == 'decrease':
        adj_qs = adj_qs.filter(quantity_change__lt=0)
    if reason_param:
        adj_qs = adj_qs.filter(reason__icontains=reason_param)

    adj_qs = adj_qs.order_by('-created_at')

    # ── KPI metrics ───────────────────────────────────────────────────────────
    total_adjustments = adj_qs.count()

    increases_qs = adj_qs.filter(quantity_change__gt=0)
    decreases_qs = adj_qs.filter(quantity_change__lt=0)

    total_increases         = increases_qs.count()
    total_decreases         = decreases_qs.count()
    total_positive_quantity = increases_qs.aggregate(t=Sum('quantity_change'))['t'] or 0
    total_negative_quantity = decreases_qs.aggregate(t=Sum('quantity_change'))['t'] or 0

    # Net value impact: sum of (quantity_change * unit_cost) across all adjustments
    net_value_impact = adj_qs.aggregate(
        val=Sum(F('quantity_change') * F('unit_cost'))
    )['val'] or Decimal('0')

    net_quantity = adj_qs.aggregate(
        net=Sum('quantity_change')
    )['net'] or 0

    # ── Quick stats ───────────────────────────────────────────────────────────
    active_users     = adj_qs.values('created_by').distinct().count()
    unique_products  = adj_qs.values('product').distinct().count()
    total_batches    = StockAdjustmentItem.objects.filter(stock_adjustment__in=adj_qs).count()
    affected_stores  = adj_qs.values('store').distinct().count()

    approved_count   = adj_qs.filter(status__in=['approved', 'applied']).count()
    approval_rate    = round((approved_count / total_adjustments * 100) if total_adjustments else 0, 1)

    avg_adjustment_value = adj_qs.aggregate(
        avg=Avg(F('quantity_change') * F('unit_cost'))
    )['avg'] or Decimal('0')

    # ── Summary cards ─────────────────────────────────────────────────────────
    reason_summary = adj_qs.values('reason').annotate(
        count=Count('id'),
        net_quantity=Sum('quantity_change'),
    ).order_by('-count')

    store_summary = adj_qs.values('store__name').annotate(
        count=Count('id'),
        total_value=Sum(F('quantity_change') * F('unit_cost')),
    ).order_by('-count')

    user_summary = adj_qs.values(
        'created_by__username',
        'created_by__first_name',
        'created_by__last_name',
    ).annotate(
        count=Count('id'),
        net_quantity=Sum('quantity_change'),
    ).order_by('-count')

    # ── Filter dropdown options ───────────────────────────────────────────────
    stores = StoreLocation.objects.filter(is_active=True).order_by('name')

    context = {
        # Dates (for display in filter button and modal)
        'start_date': start_date,
        'end_date':   end_date,

        # KPI cards
        'total_adjustments':      total_adjustments,
        'total_increases':         total_increases,
        'total_decreases':         total_decreases,
        'total_positive_quantity': total_positive_quantity,
        'total_negative_quantity': total_negative_quantity,
        'net_value_impact':        net_value_impact,
        'net_quantity':            net_quantity,

        # Quick stats row
        'active_users':          active_users,
        'unique_products':        unique_products,
        'total_batches':          total_batches,
        'approval_rate':          approval_rate,
        'avg_adjustment_value':   avg_adjustment_value,
        'affected_stores':        affected_stores,

        # Main table
        'adjustments': adj_qs,

        # Summary cards
        'reason_summary': reason_summary,
        'store_summary':  store_summary,
        'user_summary':   user_summary,

        # Filter state (to restore dropdowns on page reload)
        'selected_store':  store_id,
        'selected_status': status_param,
        'selected_type':   type_param,
        'selected_reason': reason_param,

        # Filter dropdowns
        'stores': stores,
    }

    return render(request, 'reports/stockadj_details.html', context)


@require_GET
@login_required
def adjustment_details_api(request, adjustment_id):
    """API endpoint for adjustment details"""
    try:
        adjustment = StockAdjustment.objects.select_related(
            'store', 'product', 'unit', 'created_by'
        ).get(id=adjustment_id)
        
        data = {
            'id': adjustment.id,
            'reference': adjustment.reference or f'ADJ-{adjustment.id:06d}',
            'created_at': adjustment.created_at.strftime('%Y-%m-%d %H:%M'),
            'store': adjustment.store.name,
            'product_name': adjustment.product.name,
            'sku': adjustment.product.sku,
            'quantity_change': adjustment.quantity_change,
            'unit_cost': str(adjustment.unit_cost or 0),
            'value_impact': str(abs(adjustment.quantity_change * (adjustment.unit_cost or 0))),
            'reason': adjustment.reason,
            'created_by': adjustment.created_by.get_full_name() or adjustment.created_by.username,
            'status': adjustment.get_status_display(),
            'status_color': {
                'approved': 'success',
                'pending': 'warning',
                'applied': 'info',
                'cancelled': 'danger'
            }.get(adjustment.status, 'secondary'),
            'notes': adjustment.note,
        }
        
        return JsonResponse(data)
    except StockAdjustment.DoesNotExist:
        return JsonResponse({'error': 'Adjustment not found'}, status=404)


@require_GET
@login_required
def batch_details_api(request, batch_reference):
    """API endpoint for batch details"""
    try:
        adjustments = StockAdjustment.objects.filter(
            reference=batch_reference
        ).select_related('store', 'product', 'unit', 'created_by')
        
        if not adjustments.exists():
            return JsonResponse({'error': 'Batch not found'}, status=404)
        
        first_adj = adjustments.first()
        total_items = adjustments.count()
        total_quantity = adjustments.aggregate(total=Sum('quantity_change'))['total'] or 0
        total_value = sum(adj.quantity_change * (adj.unit_cost or 0) for adj in adjustments)
        
        items = []
        for adj in adjustments:
            items.append({
                'product_name': adj.product.name,
                'sku': adj.product.sku,
                'quantity_change': adj.quantity_change,
                'unit_cost': str(adj.unit_cost or 0),
                'value_impact': str(abs(adj.quantity_change * (adj.unit_cost or 0))),
            })
        
        data = {
            'reference': batch_reference,
            'date': first_adj.created_at.strftime('%Y-%m-%d'),
            'total_items': total_items,
            'total_quantity': total_quantity,
            'total_value_impact': str(abs(total_value)),
            'store': first_adj.store.name,
            'initiated_by': first_adj.created_by.get_full_name() or first_adj.created_by.username,
            'status': first_adj.get_status_display(),
            'status_color': {
                'approved': 'success',
                'pending': 'warning',
                'applied': 'info',
                'cancelled': 'danger'
            }.get(first_adj.status, 'secondary'),
            'items': items,
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_stockadj_pdf(request):
    """Export comprehensive stock adjustment report as PDF"""
    try:
        # Get filter parameters
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        store_id = request.GET.get('store_id')
        status = request.GET.get('status')
        adjustment_type = request.GET.get('type')
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"adjustment_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Parse dates
        today = timezone.now().date()
        if date_from and date_to:
            try:
                start_date = date.fromisoformat(date_from)
                end_date = date.fromisoformat(date_to)
            except:
                start_date = today.replace(day=1)
                end_date = today
        else:
            start_date = today.replace(day=1)
            end_date = today
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header
        elements.append(Paragraph("STOCK ADJUSTMENT REPORT", title_style))
        elements.append(Paragraph(f"Period: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Base queryset
        adjustments_qs = StockAdjustment.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).select_related('store', 'product', 'unit', 'ui_unit', 'created_by')
        
        if store_id:
            adjustments_qs = adjustments_qs.filter(store_id=store_id)
        if status:
            adjustments_qs = adjustments_qs.filter(status=status)
        if adjustment_type == 'increase':
            adjustments_qs = adjustments_qs.filter(quantity_change__gt=0)
        elif adjustment_type == 'decrease':
            adjustments_qs = adjustments_qs.filter(quantity_change__lt=0)
        
        # Calculate metrics
        total_adjustments = adjustments_qs.count()
        total_positive = adjustments_qs.filter(quantity_change__gt=0).aggregate(total=Sum('quantity_change'))['total'] or 0
        total_negative = abs(adjustments_qs.filter(quantity_change__lt=0).aggregate(total=Sum('quantity_change'))['total'] or 0)
        net_change = total_positive - total_negative
        
        total_value = Decimal('0')
        positive_value = Decimal('0')
        negative_value = Decimal('0')
        
        for adj in adjustments_qs:
            value = adj.quantity_change * (adj.unit_cost or Decimal('0'))
            total_value += value
            if adj.quantity_change > 0:
                positive_value += value
            else:
                negative_value += value
        
        stores_count = adjustments_qs.values('store').distinct().count()
        products_count = adjustments_qs.values('product').distinct().count()
        
        # Status counts
        status_counts = {
            'pending': adjustments_qs.filter(status='pending').count(),
            'approved': adjustments_qs.filter(status='approved').count(),
            'applied': adjustments_qs.filter(status='applied').count(),
            'cancelled': adjustments_qs.filter(status='cancelled').count(),
        }
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Adjustments', f"{total_adjustments}"],
            ['Total Items Adjusted', f"{total_positive + total_negative:,} pcs (base units)"],
            ['Net Inventory Change', f"{net_change:+,} pcs"],
            ['Total Adjustment Value', f"UGX {abs(total_value):,.0f}"],
            ['Stores with Adjustments', f"{stores_count}"],
            ['Products Adjusted', f"{products_count}"],
            ['Pending Approvals', f"{status_counts['pending']}"],
            ['Approved Adjustments', f"{status_counts['approved']}"],
            ['Applied Adjustments', f"{status_counts['applied']}"],
            ['Cancelled', f"{status_counts['cancelled']}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY TYPE ==========
        elements.append(Paragraph("ADJUSTMENTS BY TYPE", heading_style))
        elements.append(Spacer(1, 5))
        
        type_data = [
            ['Adjustment Type', 'Count', 'Quantity Changed', 'Value Impact', '% of Total']
        ]
        
        type_data.append([
            'Positive (Add)',
            str(adjustments_qs.filter(quantity_change__gt=0).count()),
            f"+{total_positive:,}",
            f"UGX +{positive_value:,.0f}",
            f"{(positive_value/abs(total_value)*100):.1f}%" if total_value != 0 else '0%'
        ])
        
        type_data.append([
            'Negative (Remove)',
            str(adjustments_qs.filter(quantity_change__lt=0).count()),
            f"-{total_negative:,}",
            f"UGX -{abs(negative_value):,.0f}",
            f"{(abs(negative_value)/abs(total_value)*100):.1f}%" if total_value != 0 else '0%'
        ])
        
        type_data.append([
            'NET',
            str(total_adjustments),
            f"{net_change:+,}",
            f"UGX {total_value:+,.0f}",
            '100%'
        ])
        
        type_table = Table(type_data, colWidths=[120, 70, 100, 120, 80])
        type_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(type_table)
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY STATUS ==========
        elements.append(Paragraph("ADJUSTMENTS BY STATUS", heading_style))
        elements.append(Spacer(1, 5))
        
        status_items = {}
        status_values = {}
        for status_name in ['pending', 'approved', 'applied', 'cancelled']:
            status_qs = adjustments_qs.filter(status=status_name)
            status_items[status_name] = status_qs.aggregate(total=Sum('quantity_change'))['total'] or 0
            status_values[status_name] = Decimal('0')
            for adj in status_qs:
                status_values[status_name] += adj.quantity_change * (adj.unit_cost or Decimal('0'))
        
        status_data = [['Status', 'Count', 'Items', 'Value', '% of Total']]
        
        for status_name in ['pending', 'approved', 'applied', 'cancelled']:
            count = status_counts[status_name]
            if count > 0:
                percent = (abs(status_values[status_name]) / abs(total_value) * 100) if total_value != 0 else 0
                status_data.append([
                    status_name.capitalize(),
                    str(count),
                    f"{status_items[status_name]:+,}",
                    f"UGX {status_values[status_name]:+,.0f}",
                    f"{percent:.1f}%"
                ])
        
        status_data.append(['TOTAL', str(total_adjustments), f"{net_change:+,}", f"UGX {total_value:+,.0f}", '100%'])
        
        status_table = Table(status_data, colWidths=[100, 70, 100, 120, 80])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY STORE ==========
        elements.append(Paragraph("ADJUSTMENTS BY STORE", heading_style))
        elements.append(Spacer(1, 5))
        
        store_summary = adjustments_qs.values(
            'store__name'
        ).annotate(
            adjustments=Count('id'),
            positive=Sum('quantity_change', filter=Q(quantity_change__gt=0)),
            negative=Sum('quantity_change', filter=Q(quantity_change__lt=0))
        ).order_by('-adjustments')
        
        store_data = [['Store', 'Adjustments', 'Positive', 'Negative', 'Net Change', 'Value Impact']]
        
        for store in store_summary:
            store_name = store['store__name'] or 'Unknown'
            positive = store['positive'] or 0
            negative = abs(store['negative'] or 0)
            net = positive - negative
            
            # Calculate value impact
            store_adj = adjustments_qs.filter(store__name=store_name)
            value_impact = Decimal('0')
            for adj in store_adj:
                value_impact += adj.quantity_change * (adj.unit_cost or Decimal('0'))
            
            store_data.append([
                store_name,
                str(store['adjustments']),
                f"+{positive:,}",
                f"-{negative:,}",
                f"{net:+,}",
                f"UGX {value_impact:+,.0f}"
            ])
        
        store_data.append(['TOTAL', str(total_adjustments), f"+{total_positive:,}", f"-{total_negative:,}", f"{net_change:+,}", f"UGX {total_value:+,.0f}"])
        
        store_table = Table(store_data, colWidths=[100, 70, 70, 70, 80, 120])
        store_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(store_table)
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY REASON ==========
        elements.append(Paragraph("ADJUSTMENTS BY REASON", heading_style))
        elements.append(Spacer(1, 5))
        
        reason_summary = adjustments_qs.exclude(reason__isnull=True).exclude(reason__exact='').values(
            'reason'
        ).annotate(
            count=Count('id'),
            total_qty=Sum('quantity_change'),
            total_value=Sum(F('quantity_change') * F('unit_cost'))
        ).order_by('-total_value')
        
        reason_data = [['Reason', 'Count', 'Quantity', 'Value Impact', '% of Total']]
        
        for reason in reason_summary:
            percent = (abs(reason['total_value']) / abs(total_value) * 100) if total_value != 0 else 0
            reason_data.append([
                reason['reason'][:25] + '...' if len(reason['reason']) > 25 else reason['reason'],
                str(reason['count']),
                f"{reason['total_qty']:+,}",
                f"UGX {reason['total_value']:+,.0f}",
                f"{percent:.1f}%"
            ])
        
        reason_table = Table(reason_data, colWidths=[120, 60, 80, 120, 70])
        reason_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(reason_table)
        elements.append(Spacer(1, 20))
        
        # ========== DETAILED ADJUSTMENTS ==========
        elements.append(Paragraph("DETAILED ADJUSTMENTS", heading_style))
        elements.append(Spacer(1, 5))
        
        recent_adjustments = adjustments_qs.order_by('-created_at')[:15]
        
        if recent_adjustments:
            detail_data = [
                ['Adj ID', 'Date', 'Store', 'Product', 'SKU', 'UI Qty', 'UI Unit', 'Base Qty', 'Unit Cost', 'Value', 'Reason', 'Status']
            ]
            
            for adj in recent_adjustments:
                value = adj.quantity_change * (adj.unit_cost or Decimal('0'))
                
                detail_data.append([
                    adj.reference or f"ADJ-{adj.id}",
                    adj.created_at.strftime('%Y-%m-%d') if adj.created_at else '',
                    adj.store.name[:8] if adj.store else 'N/A',
                    adj.product.name[:12] + '...' if len(adj.product.name) > 12 else adj.product.name,
                    adj.product.sku or 'N/A',
                    f"{adj.ui_quantity_change:+,}" if adj.ui_quantity_change else '0',
                    adj.ui_unit.abbreviation if adj.ui_unit else 'unit',
                    f"{adj.quantity_change:+,}",
                    f"UGX {adj.unit_cost:,.0f}" if adj.unit_cost else 'N/A',
                    f"UGX {value:+,.0f}",
                    adj.reason[:15] + '...' if adj.reason and len(adj.reason) > 15 else (adj.reason or 'N/A'),
                    adj.status.capitalize()
                ])
            
            detail_data.append([
                'TOTAL', '', '', '', '', '', '', f"{net_change:+,}", f"UGX {(total_value/net_change) if net_change != 0 else 0:,.0f}", f"UGX {total_value:+,.0f}", '', ''
            ])
            
            detail_table = Table(detail_data, colWidths=[55, 55, 50, 70, 40, 35, 30, 45, 50, 70, 60, 50])
            detail_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (5, 1), (9, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(detail_table)
        else:
            elements.append(Paragraph("No adjustments found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY PRODUCT ==========
        elements.append(Paragraph("ADJUSTMENTS BY PRODUCT", heading_style))
        elements.append(Spacer(1, 5))
        
        product_summary = adjustments_qs.values(
            'product__name', 'product__sku'
        ).annotate(
            total_adj=Count('id'),
            positive=Sum('quantity_change', filter=Q(quantity_change__gt=0)),
            negative=Sum('quantity_change', filter=Q(quantity_change__lt=0)),
            value_impact=Sum(F('quantity_change') * F('unit_cost'))
        ).order_by('-value_impact')[:15]
        
        if product_summary:
            product_data = [['Product', 'SKU', 'Total Adj', 'Positive', 'Negative', 'Net Change', 'Value Impact']]
            
            for product in product_summary:
                positive = product['positive'] or 0
                negative = abs(product['negative'] or 0)
                net = positive - negative
                
                product_data.append([
                    product['product__name'][:18] + '...' if len(product['product__name']) > 18 else product['product__name'],
                    product['product__sku'] or 'N/A',
                    str(product['total_adj']),
                    f"+{positive:,}",
                    f"-{negative:,}",
                    f"{net:+,}",
                    f"UGX {product['value_impact'] or 0:+,.0f}"
                ])
            
            product_table = Table(product_data, colWidths=[100, 50, 50, 50, 50, 60, 90])
            product_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(product_table)
        else:
            elements.append(Paragraph("No product data available.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== ADJUSTMENTS BY STAFF ==========
        elements.append(Paragraph("ADJUSTMENTS BY STAFF", heading_style))
        elements.append(Spacer(1, 5))
        
        staff_summary = adjustments_qs.values(
            'created_by__id', 'created_by__first_name', 'created_by__last_name', 'created_by__username'
        ).annotate(
            adjustments=Count('id'),
            positive=Count('id', filter=Q(quantity_change__gt=0)),
            negative=Count('id', filter=Q(quantity_change__lt=0)),
            net_qty=Sum('quantity_change'),
            value_impact=Sum(F('quantity_change') * F('unit_cost'))
        ).order_by('-value_impact')
        
        staff_data = [['Staff', 'Adjustments', 'Positive', 'Negative', 'Net Change', 'Value Impact', 'Pending']]
        
        for staff in staff_summary:
            full_name = f"{staff['created_by__first_name']} {staff['created_by__last_name']}".strip()
            if not full_name:
                full_name = staff['created_by__username']
            
            pending_count = adjustments_qs.filter(
                created_by_id=staff['created_by__id'],
                status='pending'
            ).count()
            
            staff_data.append([
                full_name,
                str(staff['adjustments']),
                str(staff['positive']),
                str(staff['negative']),
                f"{staff['net_qty'] or 0:+,}",
                f"UGX {staff['value_impact'] or 0:+,.0f}",
                str(pending_count)
            ])
        
        staff_table = Table(staff_data, colWidths=[100, 60, 50, 50, 70, 100, 50])
        staff_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(staff_table)
        elements.append(Spacer(1, 20))
        
        # ========== QUALITY METRICS ==========
        elements.append(Paragraph("QUALITY METRICS", heading_style))
        elements.append(Spacer(1, 5))
        
        days_in_period = (end_date - start_date).days + 1
        adj_frequency = total_adjustments / days_in_period if days_in_period > 0 else 0
        net_variance_pct = (abs(net_change) / (total_positive + total_negative) * 100) if (total_positive + total_negative) > 0 else 0
        pending_pct = (status_counts['pending'] / total_adjustments * 100) if total_adjustments > 0 else 0
        
        # Get damage/loss rate (simplified)
        damage_qty = abs(adjustments_qs.filter(reason__icontains='damage').aggregate(total=Sum('quantity_change'))['total'] or 0)
        damage_rate = (damage_qty / (total_positive + total_negative) * 100) if (total_positive + total_negative) > 0 else 0
        
        # Get expiry rate
        expiry_qty = abs(adjustments_qs.filter(reason__icontains='expiry').aggregate(total=Sum('quantity_change'))['total'] or 0)
        expiry_rate = (expiry_qty / (total_positive + total_negative) * 100) if (total_positive + total_negative) > 0 else 0
        
        # Get theft rate
        theft_qty = abs(adjustments_qs.filter(reason__icontains='theft').aggregate(total=Sum('quantity_change'))['total'] or 0)
        theft_rate = (theft_qty / (total_positive + total_negative) * 100) if (total_positive + total_negative) > 0 else 0
        
        quality_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Adjustment Frequency', f"{adj_frequency:.1f} per day", '< 2 per day', '✅ Good' if adj_frequency < 2 else '⚠️ Needs Attention'],
            ['Net Variance', f"{net_variance_pct:.1f}%", '< 1%', '✅ Good' if net_variance_pct < 1 else '⚠️ Needs Attention'],
            ['Pending Adjustments', f"{pending_pct:.1f}%", '< 10%', '⚠️ Needs Attention' if pending_pct > 10 else '✅ Good'],
            ['Damage/Loss Rate', f"{damage_rate:.1f}%", '< 1%', '⚠️ Needs Attention' if damage_rate > 1 else '✅ Good'],
            ['Expiry Rate', f"{expiry_rate:.1f}%", '< 0.5%', '🔴 Critical' if expiry_rate > 0.5 else '✅ Good'],
            ['Theft Rate', f"{theft_rate:.1f}%", '< 0.1%', '🔴 Critical' if theft_rate > 0.1 else '⚠️ Needs Attention'],
        ]
        
        quality_table = Table(quality_data, colWidths=[150, 100, 100, 100])
        quality_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(quality_table)
        elements.append(Spacer(1, 20))
        
        # ========== RECOMMENDATIONS ==========
        elements.append(Paragraph("RECOMMENDATIONS", heading_style))
        elements.append(Spacer(1, 5))
        
        recommendations = [
            ['Priority', 'Action', 'Responsible', 'Expected Impact'],
        ]
        
        if damage_rate > 1:
            recommendations.append(['High', 'Investigate high damage rate', 'Store Manager', 'Reduce losses by UGX 200,000'])
        
        if expiry_rate > 0.5:
            recommendations.append(['High', 'Improve expiry tracking for perishables', 'Inventory Controller', 'Prevent monthly losses'])
        
        if pending_pct > 10:
            recommendations.append(['Medium', f'Clear pending adjustments ({status_counts["pending"]} items)', 'All Staff', 'Improve inventory accuracy'])
        
        if theft_rate > 0.1:
            recommendations.append(['High', 'Review security procedures', 'Security Team', 'Reduce theft incidents'])
        
        recommendations.append(['Low', 'Staff training on counting procedures', 'HR/Training', 'Reduce counting errors'])
        
        rec_table = Table(recommendations, colWidths=[80, 200, 120, 150])
        rec_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(rec_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response


# ============================================================================
# FINANCIAL REPORTS VIEWS
# ============================================================================
# Cost of Goods Sold (COGS), Profit Margin Analysis, Revenue vs Cost Report,
# Accounts Receivable
# ============================================================================

@login_required
def financial_details(request):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)

    date_range = request.GET.get('date_range')
    if date_range:
        try:
            s, e = date_range.split(' - ')
            start_date = date.fromisoformat(s)
            end_date = date.fromisoformat(e)
        except ValueError:
            pass

    store_id = request.GET.get('store')
    sales_filter = {'order__store_id': store_id} if store_id else {}
    purchase_filter = {}
    if store_id:
        purchase_filter['order__store_id'] = store_id

    # ---------------------
    # EXPRESSIONS
    # ---------------------
    REVENUE_EXPR = ExpressionWrapper(
        F('sale_price') * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    
    COST_EXPR = ExpressionWrapper(
        F('unit_cost') * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    # ---------------------
    # SALES (Revenue)
    # ---------------------
    sales_items = SalesItem.objects.filter(
        order__sale_date__range=[start_date, end_date],
        **sales_filter
    ).select_related('product', 'product__category', 'order')

    total_revenue = sales_items.aggregate(
        total=Sum(REVENUE_EXPR)
    )['total'] or Decimal('0')

    # ---------------------
    # PROPER COGS CALCULATION
    # Using average cost method for simplicity
    # ---------------------
    total_cogs = Decimal('0')
    product_cogs_map = {}
    
    # First, get all unique products sold
    sold_products = sales_items.values('product_id').distinct()
    
    for sold_product in sold_products:
        product_id = sold_product['product_id']
        
        # Get total quantity sold for this product
        product_sales = sales_items.filter(product_id=product_id).aggregate(
            total_quantity=Sum('quantity'),
            total_revenue=Sum(REVENUE_EXPR)
        )
        
        total_quantity_sold = product_sales['total_quantity'] or 0
        
        if total_quantity_sold > 0:
            # Get average purchase cost for this product
            # Look for purchases before the sale date range (FIFO principle)
            avg_cost_result = PurchaseOrderItem.objects.filter(
                product_id=product_id,
                order__purchase_date__lte=end_date  # Purchases up to the report end date
            ).aggregate(
                avg_cost=Avg('unit_cost')
            )
            
            avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
            
            # Calculate COGS for this product
            product_cogs = avg_cost * total_quantity_sold
            total_cogs += product_cogs
            product_cogs_map[product_id] = {
                'avg_cost': float(avg_cost),
                'quantity_sold': total_quantity_sold,
                'cogs': float(product_cogs)
            }

    # If no proper COGS data, use estimated 60% COGS ratio
    if total_cogs == 0 and total_revenue > 0:
        total_cogs = total_revenue * Decimal('0.60')  # 60% COGS, 40% margin
    
    # Calculate financial metrics
    cogs_ratio = (total_cogs / total_revenue * 100) if total_revenue else Decimal('0')
    gross_profit = total_revenue - total_cogs
    gross_margin = (gross_profit / total_revenue * 100) if total_revenue else Decimal('0')

    # ---------------------
    # MONTHLY TRENDS with PROPER COGS
    # ---------------------
    monthly_trends = []
    
    # Get all months in the date range
    current_month = start_date.replace(day=1)
    while current_month <= end_date.replace(day=1):
        month_start = current_month
        month_end = (current_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Revenue for this month
        month_sales = sales_items.filter(
            order__sale_date__range=[month_start, month_end]
        )
        
        month_revenue = month_sales.aggregate(
            total=Sum(REVENUE_EXPR)
        )['total'] or Decimal('0')
        
        # Calculate COGS for this month using product-level data
        month_cogs = Decimal('0')
        
        if month_revenue > 0:
            # Get unique products sold this month
            month_products = month_sales.values('product_id').distinct()
            
            for product in month_products:
                product_id = product['product_id']
                
                # Get quantity sold this month for this product
                product_month_sales = month_sales.filter(
                    product_id=product_id
                ).aggregate(
                    quantity=Sum('quantity')
                )
                
                quantity_sold = product_month_sales['quantity'] or 0
                
                if quantity_sold > 0:
                    # Use average cost from our product map
                    if product_id in product_cogs_map:
                        avg_cost = Decimal(str(product_cogs_map[product_id]['avg_cost']))
                    else:
                        # Fallback to average purchase cost
                        avg_cost_result = PurchaseOrderItem.objects.filter(
                            product_id=product_id
                        ).aggregate(
                            avg_cost=Avg('unit_cost')
                        )
                        avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                    
                    month_cogs += avg_cost * quantity_sold
        
        # If no proper COGS, use estimated
        if month_cogs == 0 and month_revenue > 0:
            month_cogs = month_revenue * Decimal('0.60')
        
        # Calculate profit and margin
        month_profit = month_revenue - month_cogs
        month_margin = (month_profit / month_revenue * 100) if month_revenue else Decimal('0')
        
        monthly_trends.append({
            'month': current_month.strftime('%Y-%m'),
            'date': current_month.strftime('%b %Y'),
            'revenue': float(month_revenue),
            'cogs': float(month_cogs),
            'profit': float(month_profit),
            'margin': float(month_margin)
        })
        
        # Move to next month
        current_month = (current_month + timedelta(days=32)).replace(day=1)

    # Convert to dictionary for template
    monthly_trends_dict = {item['month']: item for item in monthly_trends}

    # ---------------------
    # CATEGORY PERFORMANCE with PROPER COGS
    # ---------------------
    category_performance = []
    
    # Get all categories from sales
    sales_by_category = sales_items.values(
        'product__category__id', 'product__category__name'
    ).annotate(
        revenue=Sum(REVENUE_EXPR),
        quantity=Sum('quantity')
    ).order_by('-revenue')
    
    for cat_data in sales_by_category:
        category_id = cat_data['product__category__id']
        category_name = cat_data['product__category__name'] or 'Uncategorized'
        revenue = cat_data['revenue'] or Decimal('0')
        quantity = cat_data['quantity'] or 0
        
        # Calculate COGS for this category using products in the category
        category_cogs = Decimal('0')
        
        if revenue > 0:
            # Get products in this category that were sold
            category_products = sales_items.filter(
                product__category_id=category_id
            ).values('product_id').distinct()
            
            for product in category_products:
                product_id = product['product_id']
                
                # Get quantity sold for this product in this category
                product_sales = sales_items.filter(
                    product_id=product_id,
                    product__category_id=category_id
                ).aggregate(
                    quantity=Sum('quantity')
                )
                
                product_quantity = product_sales['quantity'] or 0
                
                if product_quantity > 0:
                    # Use average cost from our product map
                    if product_id in product_cogs_map:
                        avg_cost = Decimal(str(product_cogs_map[product_id]['avg_cost']))
                    else:
                        # Fallback to average purchase cost
                        avg_cost_result = PurchaseOrderItem.objects.filter(
                            product_id=product_id
                        ).aggregate(
                            avg_cost=Avg('unit_cost')
                        )
                        avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                    
                    category_cogs += avg_cost * product_quantity
        
        # If no proper COGS, use estimated
        if category_cogs == 0 and revenue > 0:
            category_cogs = revenue * Decimal('0.60')
        
        profit = revenue - category_cogs
        margin = (profit / revenue * 100) if revenue else Decimal('0')
        
        category_performance.append({
            'category': {
                'id': category_id,
                'name': category_name
            },
            'revenue': float(revenue),
            'cogs': float(category_cogs),
            'profit': float(profit),
            'margin': float(margin),
            'quantity': quantity
        })

    # ---------------------
    # TOP PRODUCTS with PROPER profit margin
    # ---------------------
    top_products = []
    
    # Get top selling products
    top_sales = sales_items.values(
        'product__id', 'product__name', 'product__sku'
    ).annotate(
        revenue=Sum(REVENUE_EXPR),
        quantity=Sum('quantity'),
        avg_price=Avg('sale_price')
    ).order_by('-revenue')[:10]
    
    for product_data in top_sales:
        product_id = product_data['product__id']
        revenue = product_data['revenue'] or Decimal('0')
        quantity = product_data['quantity'] or 0
        avg_price = product_data['avg_price'] or Decimal('0')
        
        # Calculate COGS for this product
        product_cogs = Decimal('0')
        
        if revenue > 0:
            if product_id in product_cogs_map:
                # Use calculated COGS from our map
                product_cogs = Decimal(str(product_cogs_map[product_id]['cogs']))
            else:
                # Calculate using average cost
                avg_cost_result = PurchaseOrderItem.objects.filter(
                    product_id=product_id
                ).aggregate(
                    avg_cost=Avg('unit_cost')
                )
                avg_cost = avg_cost_result['avg_cost'] or Decimal('0')
                product_cogs = avg_cost * quantity
        
        # If still no COGS, use estimated
        if product_cogs == 0 and revenue > 0:
            product_cogs = revenue * Decimal('0.60')
        
        estimated_profit = revenue - product_cogs
        profit_margin = (estimated_profit / revenue * 100) if revenue else Decimal('0')
        
        top_products.append({
            'product__id': product_id,
            'product__name': product_data['product__name'],
            'product__sku': product_data['product__sku'],
            'total_revenue': float(revenue),
            'total_quantity': quantity,
            'avg_price': float(avg_price),
            'estimated_cogs': float(product_cogs),
            'estimated_profit': float(estimated_profit),
            'profit_margin': float(profit_margin)
        })

    # ---------------------
    # NET PROFIT CALCULATION
    # For now, use gross profit as net profit (before operating expenses)
    # ---------------------
    net_profit = gross_profit
    net_profit_margin = gross_margin

    # ---------------------
    # ACCOUNTS RECEIVABLE
    # ---------------------
    accounts_receivable = None
    if store_id:
        store_filter = {'store_id': store_id}
    else:
        store_filter = {}
    
    try:
        accounts_receivable = calculate_accounts_receivable(store_filter)
    except Exception as e:
        print(f"Error calculating accounts receivable: {e}")
        accounts_receivable = {
            'total_ar': Decimal('0'),
            'aging_buckets': {
                'current': {'amount': Decimal('0'), 'count': 0},
                '1_30': {'amount': Decimal('0'), 'count': 0},
                '31_60': {'amount': Decimal('0'), 'count': 0},
                '60_plus': {'amount': Decimal('0'), 'count': 0},
            },
            'invoices': []
        }

    # ---------------------
    # PREPARE CONTEXT
    # ---------------------
    context = {
        'total_revenue': total_revenue,
        'total_cogs': total_cogs,
        'gross_profit': gross_profit,
        'gross_margin': gross_margin,
        'net_profit': net_profit,
        'net_profit_margin': net_profit_margin,
        'cogs_ratio': cogs_ratio,
        
        # Data for tables and charts
        'monthly_trends': monthly_trends_dict,
        'category_performance': category_performance,
        'top_products': top_products,
        
        # Accounts receivable data
        'accounts_receivable': accounts_receivable,
        
        # Filter data
        'stores': StoreLocation.objects.filter(is_active=True),
        'selected_store': store_id,
        'start_date': start_date,
        'end_date': end_date,
        'date_range': f"{start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}",
        'report_id': f"FIN-{timezone.now():%Y%m%d}-{random.randint(1000,9999)}",
        
        # COGS details for the template
        'cogs_details': {
            'cogs_ratio': float(cogs_ratio),
            'total_cogs': float(total_cogs),
            'gross_margin': float(gross_margin)
        }
    }
    
    return render(request, 'reports/financial_details.html', context)

def calculate_accounts_receivable(store_filter):
    """Calculate accounts receivable"""
    # Get sales with balance > 0 (using balance field from your Sales model)
    ar_sales = Sales.objects.filter(
        balance__gt=0,
        **store_filter
    ).select_related('customer', 'store').order_by('sale_date')
    
    total_ar = ar_sales.aggregate(total=Sum('balance'))['total'] or Decimal('0')
    
    # Categorize by aging
    today = timezone.now().date()
    aging_buckets = {
        'current': {'amount': Decimal('0'), 'count': 0},
        '1_30': {'amount': Decimal('0'), 'count': 0},
        '31_60': {'amount': Decimal('0'), 'count': 0},
        '60_plus': {'amount': Decimal('0'), 'count': 0},
    }
    
    invoices = []
    for sale in ar_sales:
        days_old = (today - sale.sale_date).days
        
        if days_old <= 0:
            bucket = 'current'
        elif days_old <= 30:
            bucket = '1_30'
        elif days_old <= 60:
            bucket = '31_60'
        else:
            bucket = '60_plus'
        
        aging_buckets[bucket]['amount'] += sale.balance
        aging_buckets[bucket]['count'] += 1
        
        # Calculate due date (assuming 30 days credit)
        due_date = sale.sale_date + timedelta(days=30)
        
        invoices.append({
            'id': sale.id,
            'receipt_no': sale.receipt_no,
            'customer': sale.customer.name if sale.customer else 'Walk-in',
            'sale_date': sale.sale_date,
            'due_date': due_date,
            'amount': sale.balance,
            'days_old': days_old,
        })
    
    return {
        'total_ar': total_ar,
        'aging_buckets': aging_buckets,
        'invoices': invoices,
    }
    

@login_required
def export_financial_report(request, format):
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=90)

    sales_items = SalesItem.objects.filter(
        order__sale_date__range=[start_date, end_date]
    )

    total_revenue = sales_items.aggregate(
        total=Sum(REVENUE_EXPR)
    )['total'] or Decimal('0')

    total_cogs = PurchaseOrderItem.objects.filter(
        order__purchase_date__range=[start_date, end_date]
    ).aggregate(
        total=Sum(COST_EXPR)
    )['total'] or Decimal('0')

    gross_profit = total_revenue - total_cogs


# ============================================================================
# PRODUCT MASTER REPORTS VIEWS
# ============================================================================
# Product Catalog, Category Analysis, Active/Inactive Products,
# SKU/Barcode Listing, Product Creation History
# ============================================================================

@login_required
def productmaster_details(request):
    """Product Master Reports - REAL production view using EXISTING models"""
    # Get all data using your existing models
    all_products = Product.objects.select_related('category').prefetch_related('inventories').all()

    # Optional date range filter (created_at)
    daterange = request.GET.get('daterange', '').strip()
    start_date = None
    end_date = None
    if daterange:
        try:
            parts = [p.strip() for p in daterange.split(' - ')]
            if len(parts) == 2:
                start_date = date.fromisoformat(parts[0])
                end_date = date.fromisoformat(parts[1])
                all_products = all_products.filter(created_at__date__range=[start_date, end_date])
        except ValueError:
            start_date = None
            end_date = None
    
    # 1. Basic Statistics
    total_products = all_products.count()
    active_products = all_products.filter(is_active=True).count()
    inactive_products = total_products - active_products
    
    categories = Category.objects.annotate(product_count=Count('products')).order_by('-product_count')
    categories_count = categories.count()
    
    # Calculate products with SKU
    total_skus = all_products.filter(sku__isnull=False).exclude(sku='').count()
    
    # 2. Stock analysis
    products_with_stock = sum(1 for p in all_products if p.total_stock > 0)
    out_of_stock_products = total_products - products_with_stock
    total_stock = sum(p.total_stock for p in all_products)
    
    # 3. Percentages
    if total_products > 0:
        active_percentage = (active_products / total_products) * 100
        inactive_percentage = (inactive_products / total_products) * 100
        with_stock_percentage = (products_with_stock / total_products) * 100
        out_of_stock_percentage = (out_of_stock_products / total_products) * 100
    else:
        active_percentage = inactive_percentage = with_stock_percentage = out_of_stock_percentage = 0
    
    # 4. Category analysis
    largest_category = categories.first() if categories else None
    smallest_category = categories.last() if categories else None
    empty_categories = categories.filter(product_count=0).count()
    
    # Calculate low stock categories - FIXED SYNTAX
    low_stock_categories = 0
    for category in categories:
        category_products = category.products.all()
        low_stock_count = 0
        for p in category_products:
            if p.inventories.exists():
                inventory = p.inventories.first()
                if p.total_stock <= inventory.reorder_level:
                    low_stock_count += 1
            else:
                # If no inventory record, consider it out of stock
                low_stock_count += 1
        
        if low_stock_count > 0:
            low_stock_categories += 1
    
    # Average products per category
    avg_products_per_category = categories.aggregate(avg=Avg('product_count'))['avg'] or 0
    
    # 5. Paginate products for catalog tab
    page = request.GET.get('page', 1)
    paginator = Paginator(all_products.order_by('name'), 12)  # 12 products per page
    products_page = paginator.get_page(page)
    
    # 6. Products for status tab
    status_page = request.GET.get('status_page', 1)
    status_paginator = Paginator(all_products.order_by('-is_active', 'name'), 20)
    status_products = status_paginator.get_page(status_page)
    
    # 7. Recent products (for timeline)
    recent_products = all_products.order_by('-created_at')[:10]
    
    # 8. SKU/Barcode data
    sku_products = all_products.filter(sku__isnull=False).exclude(sku='')[:50]
    sku_count = all_products.filter(sku__isnull=False).exclude(sku='').count()
    no_sku_count = total_products - sku_count
    barcode_count = all_products.filter(barcode__isnull=False).exclude(barcode='').count()
    no_barcode_count = total_products - barcode_count
    
    # 9. Creation statistics
    today = timezone.now().date()
    month_start = today.replace(day=1)
    week_start = today - timedelta(days=today.weekday())
    
    products_today = all_products.filter(created_at__date=today).count()
    products_this_week = all_products.filter(created_at__date__gte=week_start).count()
    products_this_month = all_products.filter(created_at__date__gte=month_start).count()
    
    # Calculate creation trend
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)
    products_last_month = all_products.filter(
        created_at__date__gte=last_month_start,
        created_at__date__lt=month_start
    ).count()
    
    if products_last_month > 0:
        creation_trend = ((products_this_month - products_last_month) / products_last_month) * 100
    else:
        creation_trend = 0 if products_this_month == 0 else 100
    
    # Most active category (by product count)
    most_active_category = categories.first()
    
    # Average creation rate per month
    if all_products.exists():
        first_product_date = all_products.order_by('created_at').first().created_at.date()
        months_diff = (today.year - first_product_date.year) * 12 + (today.month - first_product_date.month) + 1
        avg_creation_rate = total_products / max(months_diff, 1)
    else:
        avg_creation_rate = 0
    
    # Most active month
    from django.db.models.functions import TruncMonth
    month_counts = all_products.annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('-count')
    
    most_active_month = month_counts.first()['month'].strftime('%B %Y') if month_counts else 'N/A'
    
    context = {
        'current_date': timezone.now(),
        'start_date': start_date,
        'end_date': end_date,
        'date_range': daterange if start_date and end_date else 'Select Date Range',
        
        # Statistics
        'total_products': total_products,
        'active_products': active_products,
        'inactive_products': inactive_products,
        'categories_count': categories_count,
        'total_skus': total_skus,
        'total_stock': total_stock,
        
        # Stock analysis
        'products_with_stock': products_with_stock,
        'out_of_stock_products': out_of_stock_products,
        
        # Percentages
        'active_percentage': active_percentage,
        'inactive_percentage': inactive_percentage,
        'with_stock_percentage': with_stock_percentage,
        'out_of_stock_percentage': out_of_stock_percentage,
        
        # Category data
        'categories': categories,
        'largest_category': largest_category,
        'smallest_category': smallest_category,
        'empty_categories': empty_categories,
        'low_stock_categories': low_stock_categories,
        'avg_products_per_category': avg_products_per_category,
        'most_active_category': most_active_category,
        
        # Paginated data
        'products': products_page,
        'status_products': status_products,
        'recent_products': recent_products,
        'sku_products': sku_products,
        
        # SKU/Barcode stats
        'sku_count': sku_count,
        'no_sku_count': no_sku_count,
        'barcode_count': barcode_count,
        'no_barcode_count': no_barcode_count,
        
        # Creation stats
        'products_today': products_today,
        'products_this_week': products_this_week,
        'products_this_month': products_this_month,
        'creation_trend': creation_trend,
        'avg_creation_rate': avg_creation_rate,
        'most_active_month': most_active_month,
    }
    
    return render(request, 'reports/productmaster_details.html', context)


@login_required
def export_product_report(request, format):
    """Export product reports in various formats"""
    from django.http import HttpResponse
    import csv
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    import io
    
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Product Name', 'SKU', 'Category', 'Status', 'Stock', 'Price'])
        
        for product in Product.objects.all():
            writer.writerow([
                product.name,
                product.sku,
                product.category.name if product.category else '',
                'Active' if product.is_active else 'Inactive',
                product.total_stock,
                product.default_price or 0
            ])
        
        return response
    
    elif format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="product_report.pdf"'
        
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        p.drawString(100, 750, "Product Master Report")
        p.drawString(100, 730, f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
        p.drawString(100, 710, f"Total Products: {Product.objects.count()}")
        p.showPage()
        p.save()
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
    
    return HttpResponse('Invalid format', status=400)


@login_required
def get_product_catalog_data(request):
    """API endpoint for product catalog data (for AJAX)"""
    import json
    from django.core import serializers
    
    products = Product.objects.all()
    data = serializers.serialize('json', products)
    return HttpResponse(data, content_type='application/json')


@login_required
def get_product_statistics(request):
    """API endpoint for product statistics"""
    from django.http import JsonResponse
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    
    stats = {
        'total_products': Product.objects.count(),
        'active_products': Product.objects.filter(is_active=True).count(),
        'new_this_month': Product.objects.filter(created_at__gte=month_start).count(),
        'out_of_stock': Product.objects.filter(inventories__quantity_in_stock=0).distinct().count(),
        'low_stock': Product.objects.filter(
            inventories__quantity_in_stock__lte=F('inventories__reorder_level')
        ).distinct().count(),
    }
    
    return JsonResponse(stats)


@login_required
def export_productmaster_pdf(request):
    """Export comprehensive product mastery report as PDF"""
    try:
        # Get filter parameters
        category_id = request.GET.get('category')
        brand_filter = request.GET.get('brand')
        status_filter = request.GET.get('status')  # active/inactive/all
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"product_master_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Get current date
        today = timezone.now().date()
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        small_style = ParagraphStyle(
            'SmallStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header
        elements.append(Paragraph("PRODUCT MASTERY REPORT", title_style))
        elements.append(Paragraph(f"As of: {today.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Base queryset
        products_qs = Product.objects.all().select_related('category').prefetch_related('unit_prices', 'unit_prices__unit', 'inventories')
        
        if category_id:
            products_qs = products_qs.filter(category_id=category_id)
        if brand_filter:
            products_qs = products_qs.filter(brand__icontains=brand_filter)
        if status_filter == 'active':
            products_qs = products_qs.filter(is_active=True)
        elif status_filter == 'inactive':
            products_qs = products_qs.filter(is_active=False)
        
        # Calculate metrics
        total_products = products_qs.count()
        active_products = products_qs.filter(is_active=True).count()
        inactive_products = total_products - active_products
        
        categories_count = products_qs.values('category').distinct().count()
        brands_count = products_qs.exclude(brand__isnull=True).exclude(brand__exact='').values('brand').distinct().count()
        units_count = UnitOfMeasure.objects.count()
        stores_count = StoreLocation.objects.filter(is_active=True).count()
        
        # Stock metrics
        products_with_stock = 0
        products_out_of_stock = 0
        products_below_reorder = 0
        total_stock = 0
        total_value = Decimal('0')
        
        for product in products_qs:
            product_total_stock = 0
            for inv in product.inventories.all():
                product_total_stock += inv.quantity_in_stock
                # Get average cost
                avg_cost = InventoryBatch.objects.filter(
                    product=product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                total_value += inv.quantity_in_stock * avg_cost
            
            if product_total_stock > 0:
                products_with_stock += 1
            else:
                products_out_of_stock += 1
            
            # Check if below reorder level in any store
            for inv in product.inventories.all():
                if inv.quantity_in_stock <= inv.reorder_level and inv.quantity_in_stock > 0:
                    products_below_reorder += 1
                    break
            
            total_stock += product_total_stock
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Products', f"{total_products}"],
            ['Active Products', f"{active_products}"],
            ['Inactive Products', f"{inactive_products}"],
            ['Categories', f"{categories_count}"],
            ['Brands', f"{brands_count}"],
            ['Units of Measure', f"{units_count}"],
            ['Stores', f"{stores_count}"],
            ['Products with Stock', f"{products_with_stock}"],
            ['Products Out of Stock', f"{products_out_of_stock}"],
            ['Products Below Reorder Level', f"{products_below_reorder}"],
            ['Total Inventory Value', f"UGX {total_value:,.0f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCT CATEGORY BREAKDOWN ==========
        elements.append(Paragraph("PRODUCT CATEGORY BREAKDOWN", heading_style))
        elements.append(Spacer(1, 5))
        
        categories = Category.objects.all().annotate(
            products_count=Count('products'),
            active_count=Count('products', filter=Q(products__is_active=True)),
            inactive_count=Count('products', filter=Q(products__is_active=False))
        ).order_by('-products_count')
        
        category_data = [['Category', 'Products', 'Active', 'Inactive', '% of Total', 'With Stock', 'Total Value']]
        
        total_category_value = Decimal('0')
        for cat in categories:
            if cat.products_count > 0:
                percent = (cat.products_count / total_products * 100) if total_products > 0 else 0
                
                # Calculate category value
                cat_value = Decimal('0')
                for product in cat.products.all():
                    for inv in product.inventories.all():
                        avg_cost = InventoryBatch.objects.filter(
                            product=product,
                            store=inv.store
                        ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                        cat_value += inv.quantity_in_stock * avg_cost
                
                total_category_value += cat_value
                
                # Count products with stock
                products_with_stock_in_cat = 0
                for product in cat.products.all():
                    if product.total_stock > 0:
                        products_with_stock_in_cat += 1
                
                category_data.append([
                    cat.name[:15] + '...' if len(cat.name) > 15 else cat.name,
                    str(cat.products_count),
                    str(cat.active_count),
                    str(cat.inactive_count),
                    f"{percent:.1f}%",
                    str(products_with_stock_in_cat),
                    f"UGX {cat_value:,.0f}"
                ])
        
        category_data.append([
            'TOTAL', str(total_products), str(active_products), str(inactive_products), 
            '100%', str(products_with_stock), f"UGX {total_category_value:,.0f}"
        ])
        
        category_table = Table(category_data, colWidths=[80, 50, 40, 40, 50, 50, 100])
        category_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(category_table)
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCT STATUS OVERVIEW ==========
        elements.append(Paragraph("PRODUCT STATUS OVERVIEW", heading_style))
        elements.append(Spacer(1, 5))
        
        active_in_stock = 0
        active_below_reorder = 0
        inactive_out_of_stock = 0
        inactive_discontinued = 0
        active_stock_total = 0
        active_stock_value = Decimal('0')
        below_reorder_stock = 0
        below_reorder_value = Decimal('0')
        
        for product in products_qs:
            product_stock = product.total_stock
            product_value = Decimal('0')
            for inv in product.inventories.all():
                avg_cost = InventoryBatch.objects.filter(
                    product=product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                product_value += inv.quantity_in_stock * avg_cost
            
            if product.is_active:
                if product_stock > 0:
                    # Check if below reorder in any store
                    is_below = False
                    for inv in product.inventories.all():
                        if inv.quantity_in_stock <= inv.reorder_level and inv.quantity_in_stock > 0:
                            is_below = True
                            break
                    
                    if is_below:
                        active_below_reorder += 1
                        below_reorder_stock += product_stock
                        below_reorder_value += product_value
                    else:
                        active_in_stock += 1
                        active_stock_total += product_stock
                        active_stock_value += product_value
            else:
                if product_stock > 0:
                    inactive_out_of_stock += 1
                else:
                    inactive_discontinued += 1
        
        status_data = [
            ['Status', 'Count', '% of Total', 'Total Stock', 'Total Value'],
            ['Active (In Stock)', str(active_in_stock), f"{(active_in_stock/total_products*100):.1f}%" if total_products > 0 else '0%', 
             f"{active_stock_total:,}", f"UGX {active_stock_value:,.0f}"],
            ['Active (Below Reorder)', str(active_below_reorder), f"{(active_below_reorder/total_products*100):.1f}%" if total_products > 0 else '0%',
             f"{below_reorder_stock:,}", f"UGX {below_reorder_value:,.0f}"],
            ['Inactive (Out of Stock)', str(inactive_out_of_stock), f"{(inactive_out_of_stock/total_products*100):.1f}%" if total_products > 0 else '0%',
             '0', 'UGX 0'],
            ['Inactive (Discontinued)', str(inactive_discontinued), f"{(inactive_discontinued/total_products*100):.1f}%" if total_products > 0 else '0%',
             '0', 'UGX 0'],
            ['TOTAL', str(total_products), '100%', f"{total_stock:,}", f"UGX {total_value:,.0f}"]
        ]
        
        status_table = Table(status_data, colWidths=[120, 50, 70, 100, 120])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCT LIST WITH DETAILS ==========
        elements.append(Paragraph("PRODUCT LIST WITH DETAILS", heading_style))
        elements.append(Spacer(1, 5))
        
        product_data = [
            ['SKU', 'Product Name', 'Brand', 'Category', 'Base Unit', 'Default Price', 'Total Stock', 'Status', 'Last Updated']
        ]
        
        for product in products_qs.order_by('sku')[:25]:
            base_unit_price = product.unit_prices.filter(conversion_factor=1).first()
            base_unit = base_unit_price.unit.abbreviation if base_unit_price else 'N/A'
            default_price = base_unit_price.price if base_unit_price else 0
            
            total_stock_for_product = product.total_stock
            
            # Determine status
            if not product.is_active:
                status = 'Inactive'
            elif total_stock_for_product == 0:
                status = 'Out of Stock'
            else:
                # Check if below reorder
                is_below = False
                for inv in product.inventories.all():
                    if inv.quantity_in_stock <= inv.reorder_level and inv.quantity_in_stock > 0:
                        is_below = True
                        break
                status = 'Below Reorder' if is_below else 'Active'
            
            product_data.append([
                product.sku or 'N/A',
                product.name[:20] + '...' if len(product.name) > 20 else product.name,
                product.brand or 'N/A',
                product.category.name[:10] if product.category else 'N/A',
                base_unit,
                f"UGX {default_price:,.0f}",
                f"{total_stock_for_product:,}",
                status,
                product.created_at.strftime('%Y-%m-%d') if product.created_at else 'N/A'
            ])
        
        product_table = Table(product_data, colWidths=[55, 70, 45, 45, 35, 55, 50, 50, 55])
        product_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (5, 1), (6, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(product_table)
        elements.append(Paragraph(f"* Showing first 25 of {total_products} products", small_style))
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCTS BY BRAND ==========
        elements.append(Paragraph("PRODUCTS BY BRAND", heading_style))
        elements.append(Spacer(1, 5))
        
        brand_dict = {}
        for product in products_qs:
            if product.brand:
                brand = product.brand
                if brand not in brand_dict:
                    brand_dict[brand] = {
                        'products': 0,
                        'active': 0,
                        'stock': 0,
                        'value': Decimal('0'),
                        'top_product': product.name
                    }
                
                brand_dict[brand]['products'] += 1
                if product.is_active:
                    brand_dict[brand]['active'] += 1
                
                for inv in product.inventories.all():
                    brand_dict[brand]['stock'] += inv.quantity_in_stock
                    avg_cost = InventoryBatch.objects.filter(
                        product=product,
                        store=inv.store
                    ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                    brand_dict[brand]['value'] += inv.quantity_in_stock * avg_cost
        
        brand_data = [['Brand', 'Products', 'Active', 'Top Product', 'Total Stock', 'Total Value']]
        
        other_brand_products = 0
        other_brand_active = 0
        other_brand_stock = 0
        other_brand_value = Decimal('0')
        
        sorted_brands = sorted(brand_dict.items(), key=lambda x: x[1]['products'], reverse=True)
        
        for brand, data in sorted_brands[:9]:  # Top 9 brands
            brand_data.append([
                brand[:15] + '...' if len(brand) > 15 else brand,
                str(data['products']),
                str(data['active']),
                data['top_product'][:12] + '...' if len(data['top_product']) > 12 else data['top_product'],
                f"{data['stock']:,}",
                f"UGX {data['value']:,.0f}"
            ])
        
        # Calculate "Other Brands" total
        for brand, data in sorted_brands[9:]:
            other_brand_products += data['products']
            other_brand_active += data['active']
            other_brand_stock += data['stock']
            other_brand_value += data['value']
        
        if other_brand_products > 0:
            brand_data.append([
                'Other Brands',
                str(other_brand_products),
                str(other_brand_active),
                'Various',
                f"{other_brand_stock:,}",
                f"UGX {other_brand_value:,.0f}"
            ])
        
        brand_data.append([
            'TOTAL',
            str(total_products),
            str(active_products),
            '',
            f"{total_stock:,}",
            f"UGX {total_value:,.0f}"
        ])
        
        brand_table = Table(brand_data, colWidths=[80, 50, 40, 80, 70, 100])
        brand_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(brand_table)
        elements.append(Spacer(1, 20))
        
        # ========== UNITS OF MEASURE ==========
        elements.append(Paragraph("UNITS OF MEASURE", heading_style))
        elements.append(Spacer(1, 5))
        
        units = UnitOfMeasure.objects.all()
        unit_data = [['Unit Name', 'Abbreviation', 'Products Using', 'Is Base Unit For', 'Example Products']]
        
        for unit in units:
            products_using = ProductUnitPrice.objects.filter(unit=unit).values('product').distinct().count()
            base_unit_count = ProductUnitPrice.objects.filter(unit=unit, conversion_factor=1).count()
            
            # Get example products
            example_pups = ProductUnitPrice.objects.filter(unit=unit).select_related('product')[:3]
            examples = ', '.join([pup.product.name[:10] for pup in example_pups if pup.product])
            
            unit_data.append([
                unit.name,
                unit.abbreviation,
                str(products_using),
                str(base_unit_count),
                examples
            ])
        
        unit_table = Table(unit_data, colWidths=[80, 60, 60, 60, 120])
        unit_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(unit_table)
        elements.append(Spacer(1, 20))
        
        # ========== CONVERSION FACTORS BY PRODUCT ==========
        elements.append(Paragraph("CONVERSION FACTORS BY PRODUCT", heading_style))
        elements.append(Spacer(1, 5))
        
        conversion_data = [['Product', 'Base Unit', 'Other Unit', 'Conversion Factor', 'Price per Other Unit']]
        
        pup_count = 0
        for pup in ProductUnitPrice.objects.filter(conversion_factor__gt=1).select_related('product', 'unit')[:20]:
            base_unit = ProductUnitPrice.objects.filter(product=pup.product, conversion_factor=1).first()
            base_unit_name = base_unit.unit.abbreviation if base_unit else 'unit'
            
            conversion_data.append([
                pup.product.name[:15] + '...' if len(pup.product.name) > 15 else pup.product.name,
                base_unit_name,
                pup.unit.abbreviation,
                f"{pup.conversion_factor:.0f}",
                f"UGX {pup.price:,.0f}"
            ])
            pup_count += 1
        
        conversion_table = Table(conversion_data, colWidths=[100, 60, 60, 60, 100])
        conversion_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (3, 0), (4, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(conversion_table)
        elements.append(Spacer(1, 20))
        
        # ========== TOP 10 PRODUCTS BY STOCK VALUE ==========
        elements.append(Paragraph("TOP 10 PRODUCTS BY STOCK VALUE", heading_style))
        elements.append(Spacer(1, 5))
        
        product_values = []
        for product in products_qs:
            product_value = Decimal('0')
            for inv in product.inventories.all():
                avg_cost = InventoryBatch.objects.filter(
                    product=product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                product_value += inv.quantity_in_stock * avg_cost
            
            if product_value > 0:
                product_values.append({
                    'product': product,
                    'value': product_value,
                    'stock': product.total_stock
                })
        
        product_values.sort(key=lambda x: x['value'], reverse=True)
        
        top_products_data = [['Rank', 'Product', 'SKU', 'Brand', 'Category', 'Stock Qty', 'Unit Price', 'Total Value']]
        
        for i, item in enumerate(product_values[:10], 1):
            default_price = 0
            base_unit_price = item['product'].unit_prices.filter(conversion_factor=1).first()
            if base_unit_price:
                default_price = base_unit_price.price
            
            top_products_data.append([
                str(i),
                item['product'].name[:20] + '...' if len(item['product'].name) > 20 else item['product'].name,
                item['product'].sku or 'N/A',
                item['product'].brand or 'N/A',
                item['product'].category.name[:10] if item['product'].category else 'N/A',
                f"{item['stock']:,}",
                f"UGX {default_price:,.0f}",
                f"UGX {item['value']:,.0f}"
            ])
        
        top_table = Table(top_products_data, colWidths=[30, 80, 50, 50, 50, 50, 60, 90])
        top_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (5, 1), (7, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(top_table)
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCTS WITHOUT UNIT PRICES ==========
        elements.append(Paragraph("PRODUCTS WITHOUT UNIT PRICES", heading_style))
        elements.append(Spacer(1, 5))
        
        products_without_pricing = Product.objects.filter(unit_prices__isnull=True).order_by('-created_at')[:15]
        
        if products_without_pricing:
            no_pricing_data = [['SKU', 'Product Name', 'Brand', 'Category', 'Created Date', 'Action Required']]
            
            for product in products_without_pricing:
                no_pricing_data.append([
                    product.sku or 'N/A',
                    product.name[:20] + '...' if len(product.name) > 20 else product.name,
                    product.brand or 'N/A',
                    product.category.name[:10] if product.category else 'N/A',
                    product.created_at.strftime('%Y-%m-%d') if product.created_at else 'N/A',
                    'Set Unit Prices'
                ])
            
            no_pricing_data.append(['TOTAL', f"{products_without_pricing.count()} Products", '', '', '', ''])
            
            no_pricing_table = Table(no_pricing_data, colWidths=[60, 90, 50, 60, 70, 80])
            no_pricing_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(no_pricing_table)
        else:
            elements.append(Paragraph("All products have unit prices set.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== INACTIVE PRODUCTS ==========
        elements.append(Paragraph("INACTIVE PRODUCTS", heading_style))
        elements.append(Spacer(1, 5))
        
        inactive_products = Product.objects.filter(is_active=False).order_by('-created_at')[:15]
        
        if inactive_products:
            inactive_data = [['SKU', 'Product Name', 'Brand', 'Category', 'Last Stock Date', 'Reason']]
            
            for product in inactive_products:
                # Get last stock movement date
                last_movement = StockMovement.objects.filter(product=product).order_by('-timestamp').first()
                last_stock_date = last_movement.timestamp.strftime('%Y-%m-%d') if last_movement else 'Never'
                
                inactive_data.append([
                    product.sku or 'N/A',
                    product.name[:20] + '...' if len(product.name) > 20 else product.name,
                    product.brand or 'N/A',
                    product.category.name[:10] if product.category else 'N/A',
                    last_stock_date,
                    'Discontinued / Low Sales'
                ])
            
            inactive_data.append(['TOTAL', f"{inactive_products.count()} Products", '', '', '', ''])
            
            inactive_table = Table(inactive_data, colWidths=[60, 90, 50, 60, 70, 70])
            inactive_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(inactive_table)
        else:
            elements.append(Paragraph("No inactive products found.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== PRODUCTS ADDED (Last 30 Days) ==========
        elements.append(Paragraph("PRODUCTS ADDED (Last 30 Days)", heading_style))
        elements.append(Spacer(1, 5))
        
        thirty_days_ago = today - timedelta(days=30)
        recent_products = Product.objects.filter(created_at__date__gte=thirty_days_ago).order_by('-created_at')[:15]
        
        if recent_products:
            recent_data = [['Date Added', 'SKU', 'Product Name', 'Brand', 'Category', 'Created By']]
            
            for product in recent_products:
                # Try to get creator from audit trail - simplified
                created_by = 'System'
                
                recent_data.append([
                    product.created_at.strftime('%Y-%m-%d') if product.created_at else 'N/A',
                    product.sku or 'N/A',
                    product.name[:20] + '...' if len(product.name) > 20 else product.name,
                    product.brand or 'N/A',
                    product.category.name[:10] if product.category else 'N/A',
                    created_by
                ])
            
            recent_data.append(['TOTAL', f"{recent_products.count()} Products", '', '', '', ''])
            
            recent_table = Table(recent_data, colWidths=[70, 60, 90, 50, 60, 70])
            recent_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(recent_table)
        else:
            elements.append(Paragraph("No products added in the last 30 days.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== RECOMMENDATIONS ==========
        elements.append(Paragraph("PRODUCT MASTERY RECOMMENDATIONS", heading_style))
        elements.append(Spacer(1, 5))
        
        recommendations = [
            ['Priority', 'Action', 'Products Affected', 'Responsible'],
        ]
        
        if products_without_pricing.exists():
            recommendations.append(['High', f'Set unit prices for products without pricing', str(products_without_pricing.count()), 'Pricing Team'])
        
        if inactive_products.exists():
            recommendations.append(['High', f'Review inactive products for removal/disposal', str(inactive_products.count()), 'Product Manager'])
        
        # Count slow movers (simplified - products with low turnover)
        slow_movers = 108  # Placeholder - would need actual sales data
        recommendations.append(['Medium', f'Investigate slow-moving products', str(slow_movers), 'Sales Team'])
        
        recommendations.append(['Medium', 'Review conversion factors for accuracy', 'All', 'Inventory Team'])
        recommendations.append(['Low', 'Update product categories and brands', 'Various', 'Product Team'])
        
        rec_table = Table(recommendations, colWidths=[80, 200, 80, 100])
        rec_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(rec_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response


# ============================================================================
# REORDER & LOW STOCK REPORTS VIEWS
# ============================================================================
# Below Reorder Level Alerts, Out-of-stock Report, Store-specific Low Stock,
# Automatic Reorder Suggestions
# ============================================================================

@login_required
def reorder_details(request):
    """Reorder & Low Stock Reports view"""
    
    # Get all active stores
    stores = StoreLocation.objects.filter(is_active=True)
    
    # Get current date for calculations
    today = timezone.now().date()
    
    # 1. BELOW REORDER LEVEL ITEMS
    below_reorder_items = []
    for inventory in Inventory.objects.filter(
        quantity_in_stock__gt=0  # Only items with some stock
    ).select_related('product', 'store'):
        
        # Calculate stock percentage relative to reorder level
        if inventory.reorder_level > 0:
            stock_percentage = (inventory.quantity_in_stock / inventory.reorder_level) * 100
        else:
            stock_percentage = 0
            
        # Determine priority
        if inventory.quantity_in_stock == 0:
            priority = 'critical'
        elif stock_percentage <= 50:
            priority = 'critical'
        elif stock_percentage <= 75:
            priority = 'high'
        elif stock_percentage <= 90:
            priority = 'medium'
        else:
            priority = 'low'
        
        # Calculate average daily sales (last 30 days)
        thirty_days_ago = today - timedelta(days=30)
        total_sales = SalesItem.objects.filter(
            product=inventory.product,
            order__sale_date__gte=thirty_days_ago,
            order__store=inventory.store
        ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        avg_daily_sales = total_sales / 30 if total_sales > 0 else 0
        
        # Calculate days until stockout
        days_until_stockout = inventory.quantity_in_stock / avg_daily_sales if avg_daily_sales > 0 else 999
        
        # Get days below reorder level (simplified - would need historical tracking)
        days_below_reorder = 0  # You would need to track this historically
        
        if inventory.quantity_in_stock <= inventory.reorder_level:
            below_reorder_items.append({
                'inventory': inventory,
                'stock_percentage': stock_percentage,
                'priority': priority,
                'avg_daily_sales': round(avg_daily_sales, 1),
                'days_until_stockout': round(days_until_stockout, 1),
                'days_below_reorder': days_below_reorder,
                'stock_value': inventory.quantity_in_stock * inventory.product.default_price
            })
    
    # Sort by priority (critical first)
    below_reorder_items.sort(key=lambda x: {
        'critical': 0, 'high': 1, 'medium': 2, 'low': 3
    }[x['priority']])
    
    # 2. OUT OF STOCK ITEMS
    out_of_stock_items = []
    for inventory in Inventory.objects.filter(
        quantity_in_stock=0
    ).select_related('product', 'store'):
        
        # Get last sale date
        last_sale = SalesItem.objects.filter(
            product=inventory.product,
            order__store=inventory.store
        ).order_by('-order__sale_date').first()
        
        # Count backorders/requests (would need a backorder model)
        backorders = 0
        
        # Get supplier info (from purchase orders)
        last_purchase = PurchaseOrderItem.objects.filter(
            product=inventory.product
        ).order_by('-order__purchase_date').first()
        
        supplier = last_purchase.order.supplier if last_purchase else None
        
        # Calculate days out of stock
        days_out = 0  # You would need to track when it went out of stock
        
        # Determine urgency based on sales velocity
        thirty_days_ago = today - timedelta(days=30)
        sales_last_month = SalesItem.objects.filter(
            product=inventory.product,
            order__sale_date__gte=thirty_days_ago,
            order__store=inventory.store
        ).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        
        if sales_last_month > 20:
            urgency = 'critical'
        elif sales_last_month > 10:
            urgency = 'high'
        elif sales_last_month > 0:
            urgency = 'medium'
        else:
            urgency = 'low'
        
        out_of_stock_items.append({
            'inventory': inventory,
            'last_sale': last_sale.order.sale_date if last_sale else None,
            'backorders': backorders,
            'supplier': supplier,
            'days_out': days_out,
            'urgency': urgency,
            'sales_last_month': sales_last_month
        })
    
    # Sort by urgency
    out_of_stock_items.sort(key=lambda x: {
        'critical': 0, 'high': 1, 'medium': 2, 'low': 3
    }[x['urgency']])
    
    # 3. STORE-SPECIFIC LOW STOCK
    store_specific_data = {}
    for store in stores:
        low_stock_in_store = []
        
        for inventory in Inventory.objects.filter(
            store=store,
            quantity_in_stock__gt=0,  # Has some stock
            quantity_in_stock__lte=F('reorder_level') * 2  # Below 2x reorder level
        ).select_related('product'):
            
            # Get stock in other stores
            other_store_stock = []
            for other_store in stores.exclude(id=store.id):
                try:
                    other_inv = Inventory.objects.get(
                        product=inventory.product,
                        store=other_store
                    )
                    if other_inv.quantity_in_stock > inventory.reorder_level * 1.5:  # Has excess stock
                        other_store_stock.append({
                            'store': other_store,
                            'quantity': other_inv.quantity_in_stock
                        })
                except Inventory.DoesNotExist:
                    continue
            
            # Determine transfer recommendation
            if other_store_stock:
                # Can transfer from other stores
                best_source = max(other_store_stock, key=lambda x: x['quantity'])
                
                # Get the reorder level for the source store
                try:
                    source_inventory = Inventory.objects.get(
                        product=inventory.product,
                        store=best_source['store']
                    )
                    source_reorder_level = source_inventory.reorder_level
                except Inventory.DoesNotExist:
                    source_reorder_level = 10  # Default
                
                # Calculate suggested units
                suggested_units = min(
                    inventory.reorder_level - inventory.quantity_in_stock,
                    best_source['quantity'] - source_reorder_level
                )
                suggested_units = max(suggested_units, 1)  # At least 1 unit
                
                recommendation = {
                    'type': 'transfer',
                    'from_store': best_source['store'],
                    'suggested_units': suggested_units,
                    'lead_time': '1-2 days'
                }
            else:
                # Need to reorder from supplier
                recommendation = {
                    'type': 'reorder',
                    'lead_time': '4-5 days'
                }
            
            low_stock_in_store.append({
                'inventory': inventory,
                'other_store_stock': other_store_stock,
                'recommendation': recommendation
            })
        
        store_specific_data[store] = low_stock_in_store
    
    # 4. SUMMARY STATISTICS
    total_out_of_stock = Inventory.objects.filter(quantity_in_stock=0).count()
    total_below_reorder = len(below_reorder_items)
    
    # Calculate total value at risk
    total_value_at_risk = sum(
        item['stock_value'] for item in below_reorder_items
    ) + sum(
        item['inventory'].reorder_level * item['inventory'].product.default_price
        for item in below_reorder_items if item['inventory'].quantity_in_stock == 0
    )
    
    # Store-wise counts
    store_stats = {}
    for store in stores:
        store_stats[store.name] = {
            'low_stock': Inventory.objects.filter(
                store=store,
                quantity_in_stock__gt=0,
                quantity_in_stock__lte=F('reorder_level')
            ).count(),
            'out_of_stock': Inventory.objects.filter(
                store=store,
                quantity_in_stock=0
            ).count()
        }
    
    context = {
        'below_reorder_items': below_reorder_items[:50],  # Limit to 50 items
        'out_of_stock_items': out_of_stock_items[:50],
        'store_specific_data': store_specific_data,
        'stores': stores,
        'total_out_of_stock': total_out_of_stock,
        'total_below_reorder': total_below_reorder,
        'total_value_at_risk': total_value_at_risk,
        'store_stats': store_stats,
        'report_id': f"STOCK-ALERT-{today.strftime('%Y%m%d')}-001",
        'report_date': today,
        'urgent_count': len([item for item in below_reorder_items if item['priority'] == 'critical']),
        'warning_count': len([item for item in below_reorder_items if item['priority'] == 'high']),
        'monitor_count': len([item for item in below_reorder_items if item['priority'] in ['medium', 'low']]),
    }
    
    return render(request, 'reports/reorder_details.html', context)


@login_required
def export_reorder_pdf(request):
    """Export comprehensive reorder level report as PDF"""
    try:
        # Get filter parameters
        store_id = request.GET.get('store')
        priority = request.GET.get('priority')  # critical/warning/at_reorder/all
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"reorder_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Get current date
        today = timezone.now().date()
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles (same as previous)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        small_style = ParagraphStyle(
            'SmallStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header
        elements.append(Paragraph("REORDER LEVEL REPORT", title_style))
        elements.append(Paragraph(f"As of: {today.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Base queryset
        inventories_qs = Inventory.objects.select_related(
            'product', 'product__category', 'store'
        ).filter(product__is_active=True)
        
        if store_id:
            inventories_qs = inventories_qs.filter(store_id=store_id)
        
        # Calculate metrics
        total_products = inventories_qs.values('product').distinct().count()
        products_with_reorder = inventories_qs.exclude(reorder_level=0).values('product').distinct().count()
        
        # Categorize by reorder status
        critical_items = []  # Out of stock
        warning_items = []   # Below reorder level
        at_reorder_items = [] # At or near reorder level
        adequate_items = []   # Above reorder level
        no_reorder_items = [] # No reorder level set
        
        total_critical_stock = 0
        total_warning_stock = 0
        total_at_reorder_stock = 0
        total_adequate_stock = 0
        
        total_critical_value = Decimal('0')
        total_warning_value = Decimal('0')
        total_at_reorder_value = Decimal('0')
        total_reorder_qty_needed = 0
        total_reorder_value_needed = Decimal('0')
        
        for inv in inventories_qs:
            if inv.quantity_in_stock == 0:
                critical_items.append(inv)
                total_critical_stock += inv.quantity_in_stock
                
                # Get cost
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                
                # For out of stock, reorder value is reorder_level * cost
                reorder_needed = inv.reorder_level
                total_reorder_qty_needed += reorder_needed
                reorder_value = reorder_needed * avg_cost
                total_reorder_value_needed += reorder_value
                total_critical_value += reorder_value
                
            elif inv.quantity_in_stock <= inv.reorder_level:
                warning_items.append(inv)
                total_warning_stock += inv.quantity_in_stock
                
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                
                shortage = inv.reorder_level - inv.quantity_in_stock
                total_reorder_qty_needed += shortage
                reorder_value = shortage * avg_cost
                total_reorder_value_needed += reorder_value
                total_warning_value += inv.quantity_in_stock * avg_cost
                
            elif inv.quantity_in_stock <= inv.reorder_level * 1.2:  # Within 20% of reorder level
                at_reorder_items.append(inv)
                total_at_reorder_stock += inv.quantity_in_stock
                
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                total_at_reorder_value += inv.quantity_in_stock * avg_cost
            else:
                adequate_items.append(inv)
                total_adequate_stock += inv.quantity_in_stock
        
        # Products without reorder levels
        products_without_reorder = Product.objects.filter(
            is_active=True,
            inventories__isnull=False
        ).exclude(
            inventories__reorder_level__gt=0
        ).distinct()
        
        for product in products_without_reorder:
            for inv in product.inventories.all():
                if inv.reorder_level == 0:
                    no_reorder_items.append(inv)
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Products Monitored', f"{total_products}"],
            ['Products with Reorder Levels', f"{products_with_reorder}"],
            ['Products Below Reorder Level', f"{len(warning_items)}"],
            ['Products Out of Stock', f"{len(critical_items)}"],
            ['Total Reorder Quantity Needed', f"{total_reorder_qty_needed} units"],
            ['Estimated Reorder Value', f"UGX {total_reorder_value_needed:,.0f}"],
            ['Stores with Below Reorder Items', f"{inventories_qs.values('store').distinct().count() if warning_items else 0}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== REORDER STATUS SUMMARY ==========
        elements.append(Paragraph("REORDER STATUS SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        status_data = [
            ['Status', 'Count', '% of Total', 'Total Stock', 'Reorder Value', 'Priority'],
            ['Critical (Out of Stock)', str(len(critical_items)), f"{(len(critical_items)/total_products*100):.1f}%" if total_products > 0 else '0%', 
             '0', f"UGX {total_critical_value:,.0f}", '🔴 Immediate'],
            ['Warning (Below Reorder)', str(len(warning_items)), f"{(len(warning_items)/total_products*100):.1f}%" if total_products > 0 else '0%',
             f"{total_warning_stock:,}", f"UGX {total_reorder_value_needed - total_critical_value:,.0f}", '⚠️ This Week'],
            ['At Reorder Point', str(len(at_reorder_items)), f"{(len(at_reorder_items)/total_products*100):.1f}%" if total_products > 0 else '0%',
             f"{total_at_reorder_stock:,}", 'UGX 0', '📅 Next Week'],
            ['Adequate (Above Reorder)', str(len(adequate_items)), f"{(len(adequate_items)/total_products*100):.1f}%" if total_products > 0 else '0%',
             f"{total_adequate_stock:,}", 'UGX 0', '✅ Good'],
            ['Not Set', str(len(no_reorder_items)), f"{(len(no_reorder_items)/total_products*100):.1f}%" if total_products > 0 else '0%',
             '0', 'UGX 0', '📝 Setup Required'],
            ['TOTAL', str(total_products), '100%', f"{total_critical_stock + total_warning_stock + total_at_reorder_stock + total_adequate_stock:,}", 
             f"UGX {total_reorder_value_needed:,.0f}", ''],
        ]
        
        status_table = Table(status_data, colWidths=[120, 40, 50, 70, 90, 80])
        status_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(status_table)
        elements.append(Spacer(1, 20))
        
        # ========== CRITICAL ITEMS - OUT OF STOCK ==========
        if critical_items and (priority == 'critical' or not priority):
            elements.append(Paragraph("CRITICAL ITEMS - OUT OF STOCK", heading_style))
            elements.append(Spacer(1, 5))
            
            critical_data = [['Product', 'SKU', 'Brand', 'Category', 'Store', 'Last Received', 'Days Out', 'Lost Revenue']]
            
            total_lost_revenue = Decimal('0')
            for inv in critical_items[:15]:
                last_batch = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).order_by('-received_date').first()
                
                last_received = last_batch.received_date.date() if last_batch else 'Never'
                days_out = (today - last_received).days if last_batch and last_received != 'Never' else 'N/A'
                
                # Estimate lost revenue (simplified)
                default_price = 0
                base_unit_price = inv.product.unit_prices.filter(conversion_factor=1).first()
                if base_unit_price:
                    default_price = base_unit_price.price
                
                lost_revenue = inv.reorder_level * default_price
                total_lost_revenue += lost_revenue
                
                critical_data.append([
                    inv.product.name[:15] + '...' if len(inv.product.name) > 15 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.product.brand or 'N/A',
                    inv.product.category.name[:10] if inv.product.category else 'N/A',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(last_received) if last_received != 'Never' else 'Never',
                    str(days_out) if days_out != 'N/A' else 'N/A',
                    f"UGX {lost_revenue:,.0f}"
                ])
            
            critical_data.append(['TOTAL', '', '', '', '', '', '', f"UGX {total_lost_revenue:,.0f}"])
            
            critical_table = Table(critical_data, colWidths=[70, 45, 40, 45, 40, 60, 35, 70])
            critical_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (6, 1), (7, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(critical_table)
            elements.append(Spacer(1, 20))
        
        # ========== ITEMS BELOW REORDER LEVEL ==========
        if warning_items and (priority == 'warning' or not priority):
            elements.append(Paragraph("ITEMS BELOW REORDER LEVEL", heading_style))
            elements.append(Spacer(1, 5))
            
            warning_data = [['Product', 'SKU', 'Store', 'Current Stock', 'Reorder Level', 'Shortage', 'Reorder Qty', 'Unit Cost', 'Est Cost', 'Brand']]
            
            total_est_cost = Decimal('0')
            for inv in warning_items[:20]:
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                
                shortage = inv.reorder_level - inv.quantity_in_stock
                reorder_qty = inv.reorder_level  # Reorder to full level
                est_cost = reorder_qty * avg_cost
                total_est_cost += est_cost
                
                warning_data.append([
                    inv.product.name[:12] + '...' if len(inv.product.name) > 12 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(inv.quantity_in_stock),
                    str(inv.reorder_level),
                    str(shortage),
                    str(reorder_qty),
                    f"UGX {avg_cost:,.0f}",
                    f"UGX {est_cost:,.0f}",
                    inv.product.brand or 'N/A'
                ])
            
            warning_data.append(['TOTAL', '', '', '', '', '', '', '', f"UGX {total_est_cost:,.0f}", ''])
            
            warning_table = Table(warning_data, colWidths=[60, 40, 35, 30, 30, 25, 30, 45, 60, 45])
            warning_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (3, 1), (8, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(warning_table)
            elements.append(Spacer(1, 20))
        
        # ========== ITEMS AT REORDER POINT ==========
        if at_reorder_items and (priority == 'at_reorder' or not priority):
            elements.append(Paragraph("ITEMS AT REORDER POINT", heading_style))
            elements.append(Spacer(1, 5))
            
            at_data = [['Product', 'SKU', 'Store', 'Current Stock', 'Reorder Level', 'Reorder Qty', 'Est Cost', 'Lead Time', 'Supplier']]
            
            total_at_cost = Decimal('0')
            for inv in at_reorder_items[:15]:
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=inv.store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                
                reorder_qty = inv.reorder_level
                est_cost = reorder_qty * avg_cost
                total_at_cost += est_cost
                
                # Get supplier from recent purchase
                last_po_item = PurchaseOrderItem.objects.filter(
                    product=inv.product
                ).select_related('order__supplier').order_by('-order__purchase_date').first()
                
                supplier = last_po_item.order.supplier.name if last_po_item and last_po_item.order.supplier else 'N/A'
                lead_time = '3 days'  # Placeholder
                
                at_data.append([
                    inv.product.name[:12] + '...' if len(inv.product.name) > 12 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(inv.quantity_in_stock),
                    str(inv.reorder_level),
                    str(reorder_qty),
                    f"UGX {est_cost:,.0f}",
                    lead_time,
                    supplier[:10] if supplier != 'N/A' else 'N/A'
                ])
            
            at_data.append(['TOTAL', '', '', '', '', '', f"UGX {total_at_cost:,.0f}", '', ''])
            
            at_table = Table(at_data, colWidths=[60, 40, 35, 30, 30, 30, 60, 40, 50])
            at_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (3, 1), (6, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(at_table)
            elements.append(Spacer(1, 20))
        
        # ========== PRODUCTS WITHOUT REORDER LEVELS ==========
        if no_reorder_items:
            elements.append(Paragraph("PRODUCTS WITHOUT REORDER LEVELS", heading_style))
            elements.append(Spacer(1, 5))
            
            no_reorder_data = [['Product', 'SKU', 'Brand', 'Category', 'Store', 'Current Stock', 'Recommended Level', 'Priority']]
            
            for inv in no_reorder_items[:15]:
                # Calculate recommended level based on sales (simplified)
                recommended = max(10, inv.quantity_in_stock // 3)
                priority = 'Medium' if inv.quantity_in_stock > 50 else 'Low'
                
                no_reorder_data.append([
                    inv.product.name[:15] + '...' if len(inv.product.name) > 15 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.product.brand or 'N/A',
                    inv.product.category.name[:10] if inv.product.category else 'N/A',
                    inv.store.name[:8] if inv.store else 'N/A',
                    str(inv.quantity_in_stock),
                    str(recommended),
                    priority
                ])
            
            no_reorder_data.append(['TOTAL', f"{len(no_reorder_items)} Products" '', '', '', '', '', '', ''])
            
            no_reorder_table = Table(no_reorder_data, colWidths=[70, 45, 40, 45, 40, 40, 50, 50])
            no_reorder_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(no_reorder_table)
            elements.append(Spacer(1, 20))
        
        # ========== REORDER VALUE BY STORE ==========
        elements.append(Paragraph("REORDER VALUE BY STORE", heading_style))
        elements.append(Spacer(1, 5))
        
        store_reorder = {}
        for inv in critical_items + warning_items:
            store_name = inv.store.name if inv.store else 'Unknown'
            if store_name not in store_reorder:
                store_reorder[store_name] = {
                    'below_count': 0,
                    'reorder_qty': 0,
                    'reorder_value': Decimal('0')
                }
            
            avg_cost = InventoryBatch.objects.filter(
                product=inv.product,
                store=inv.store
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            
            if inv.quantity_in_stock == 0:
                needed_qty = inv.reorder_level
            else:
                needed_qty = inv.reorder_level - inv.quantity_in_stock
            
            store_reorder[store_name]['below_count'] += 1
            store_reorder[store_name]['reorder_qty'] += needed_qty
            store_reorder[store_name]['reorder_value'] += needed_qty * avg_cost
        
        store_reorder_data = [['Store', 'Products Below Reorder', 'Reorder Quantity', 'Reorder Value', 'Priority Items']]
        
        for store_name, data in store_reorder.items():
            store_reorder_data.append([
                store_name[:15] + '...' if len(store_name) > 15 else store_name,
                str(data['below_count']),
                f"{data['reorder_qty']}",
                f"UGX {data['reorder_value']:,.0f}",
                'Rice, Water, Oil'  # Placeholder
            ])
        
        store_reorder_data.append(['TOTAL', str(len(critical_items) + len(warning_items)), str(total_reorder_qty_needed), f"UGX {total_reorder_value_needed:,.0f}", ''])
        
        store_reorder_table = Table(store_reorder_data, colWidths=[80, 70, 70, 100, 100])
        store_reorder_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(store_reorder_table)
        elements.append(Spacer(1, 20))
        
        # ========== RECOMMENDED PURCHASE ORDER BY SUPPLIER ==========
        elements.append(Paragraph("RECOMMENDED PURCHASE ORDER BY SUPPLIER", heading_style))
        elements.append(Spacer(1, 5))
        
        # Group by supplier from recent purchase orders
        supplier_orders = {}
        for inv in critical_items + warning_items:
            last_po_item = PurchaseOrderItem.objects.filter(
                product=inv.product
            ).select_related('order__supplier').order_by('-order__purchase_date').first()
            
            supplier_name = 'Unknown'
            if last_po_item and last_po_item.order.supplier:
                supplier_name = last_po_item.order.supplier.name
            
            if supplier_name not in supplier_orders:
                supplier_orders[supplier_name] = {
                    'items': [],
                    'total_qty': 0,
                    'total_value': Decimal('0')
                }
            
            avg_cost = InventoryBatch.objects.filter(
                product=inv.product,
                store=inv.store
            ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
            
            if inv.quantity_in_stock == 0:
                needed_qty = inv.reorder_level
            else:
                needed_qty = inv.reorder_level - inv.quantity_in_stock
            
            supplier_orders[supplier_name]['items'].append({
                'product': inv.product,
                'store': inv.store,
                'qty': needed_qty,
                'cost': avg_cost
            })
            supplier_orders[supplier_name]['total_qty'] += needed_qty
            supplier_orders[supplier_name]['total_value'] += needed_qty * avg_cost
        
        po_data = [['Supplier', 'Product', 'SKU', 'Store', 'Reorder Qty', 'Unit', 'Unit Cost', 'Est Cost']]
        
        total_po_value = Decimal('0')
        for supplier_name, data in supplier_orders.items():
            for item in data['items'][:3]:  # Limit per supplier
                base_unit = 'N/A'
                base_unit_price = item['product'].unit_prices.filter(conversion_factor=1).first()
                if base_unit_price:
                    base_unit = base_unit_price.unit.abbreviation
                
                po_data.append([
                    supplier_name[:12] + '...' if len(supplier_name) > 12 else supplier_name,
                    item['product'].name[:12] + '...' if len(item['product'].name) > 12 else item['product'].name,
                    item['product'].sku or 'N/A',
                    item['store'].name[:8] if item['store'] else 'N/A',
                    str(item['qty']),
                    base_unit,
                    f"UGX {item['cost']:,.0f}",
                    f"UGX {item['qty'] * item['cost']:,.0f}"
                ])
                total_po_value += item['qty'] * item['cost']
        
        po_data.append(['TOTAL', '', '', '', '', '', '', f"UGX {total_po_value:,.0f}"])
        
        po_table = Table(po_data, colWidths=[70, 70, 45, 40, 35, 30, 50, 70])
        po_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(po_table)
        elements.append(Spacer(1, 20))
        
        # ========== REORDER RECOMMENDATIONS ==========
        elements.append(Paragraph("REORDER RECOMMENDATIONS", heading_style))
        elements.append(Spacer(1, 5))
        
        recommendations = [
            ['Priority', 'Action', 'Products', 'Responsible', 'Timeline'],
        ]
        
        if critical_items:
            recommendations.append(['Immediate', f'Place emergency orders for out-of-stock items', str(len(critical_items)), 'Procurement', 'Today'])
        
        if warning_items:
            recommendations.append(['High', f'Order items below reorder level', str(len(warning_items)), 'Procurement', 'This Week'])
        
        if no_reorder_items:
            recommendations.append(['Medium', f'Set reorder levels for {len(no_reorder_items)} products without', 'Inventory Team', 'This Month'])
        
        if at_reorder_items:
            recommendations.append(['Medium', f'Review reorder levels for items at threshold', str(len(at_reorder_items)), 'Inventory Team', 'This Month'])
        
        recommendations.append(['Low', 'Analyze reorder levels by seasonality', 'All', 'Planning Team', 'Next Month'])
        
        rec_table = Table(recommendations, colWidths=[80, 200, 50, 80, 60])
        rec_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(rec_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response


# ============================================================================
# STORE LOCATION REPORTS VIEWS
# ============================================================================
# Store Capacity/Utilization, Store Activity Analysis, Branch-wise Store Comparison,
# Default Store Analysis
# ============================================================================

def decimal_to_float(obj):
    """Convert Decimal objects to float for JSON serialization"""
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError(f"Object of type {obj.__class__.__name__} is not JSON serializable")


@login_required
def stocklocation_details(request):
    """
    Stock Location Performance Dashboard view.
    Matches: reports/stocklocation_details.html
    """
    today = timezone.now().date()

    # ── Filters ──────────────────────────────────────────────────────────────
    branch_id        = request.GET.get('branch', '')
    status_param     = request.GET.get('status', '')      # 'active' | 'inactive'
    performance_param = request.GET.get('performance', '') # 'high' | 'medium' | 'low'
    start_date_raw   = request.GET.get('start_date', '')
    end_date_raw     = request.GET.get('end_date', '')

    start_date = None
    end_date   = None

    if start_date_raw:
        try:
            start_date = date.fromisoformat(start_date_raw)
        except (ValueError, AttributeError):
            pass
    if end_date_raw:
        try:
            end_date = date.fromisoformat(end_date_raw)
        except (ValueError, AttributeError):
            pass

    # Display dates — fall back to sensible defaults for the template header
    display_start = start_date or date(today.year, today.month, 1)
    display_end   = end_date or today

    # ── Store queryset ────────────────────────────────────────────────────────
    stores_qs = StoreLocation.objects.select_related('branch')
    if branch_id:
        stores_qs = stores_qs.filter(branch_id=branch_id)
    if status_param == 'active':
        stores_qs = stores_qs.filter(is_active=True)
    elif status_param == 'inactive':
        stores_qs = stores_qs.filter(is_active=False)

    # ── Sales queryset for the period ─────────────────────────────────────────
    sales_qs = Sales.objects.filter(is_cancelled=False)
    if start_date:
        sales_qs = sales_qs.filter(sale_date__gte=start_date)
    if end_date:
        sales_qs = sales_qs.filter(sale_date__lte=end_date)
    if branch_id:
        sales_qs = sales_qs.filter(store__branch_id=branch_id)

    # Previous period for trend calculation
    if start_date and end_date:
        period_days = (end_date - start_date).days or 1
        prev_start  = start_date - timedelta(days=period_days)
        prev_end    = start_date - timedelta(days=1)
    else:
        prev_start = date(today.year, today.month, 1) - timedelta(days=30)
        prev_end   = date(today.year, today.month, 1) - timedelta(days=1)

    prev_sales_qs = Sales.objects.filter(
        is_cancelled=False,
        sale_date__range=[prev_start, prev_end]
    )

    # ── KPI metrics ───────────────────────────────────────────────────────────
    total_stores   = stores_qs.count()
    active_stores  = stores_qs.filter(is_active=True).count()
    inactive_stores = stores_qs.filter(is_active=False).count()

    total_sales = sales_qs.filter(
        store__in=stores_qs
    ).aggregate(t=Sum('total_amount'))['t'] or 0

    total_transactions = sales_qs.filter(store__in=stores_qs).count()

    # Inventory totals
    inv_qs = Inventory.objects.filter(store__in=stores_qs)
    total_stock_units = inv_qs.aggregate(t=Sum('quantity_in_stock'))['t'] or 0
    total_skus        = inv_qs.values('product').distinct().count()

    # Low stock across all stores
    total_low_stock = inv_qs.filter(
        quantity_in_stock__gt=0,
        quantity_in_stock__lte=F('reorder_level')
    ).count()

    # Stock value
    total_stock_value = InventoryBatch.objects.filter(
        store__in=stores_qs,
        remaining_quantity__gt=0
    ).aggregate(
        val=Sum(F('remaining_quantity') * F('unit_cost'))
    )['val'] or Decimal('0')

    avg_transaction_total = (
        Decimal(str(total_sales)) / total_transactions
    ) if total_transactions else Decimal('0')

    # ── Per-store performance data ─────────────────────────────────────────────
    store_performance_data = []
    efficiency_list = []
    growth_list     = []

    for store in stores_qs:
        # Sales this period
        store_sales_qs = sales_qs.filter(store=store)
        monthly_sales  = store_sales_qs.aggregate(t=Sum('total_amount'))['t'] or 0
        transactions   = store_sales_qs.count()
        avg_transaction = (monthly_sales / transactions) if transactions else 0

        # Previous period sales for growth
        prev_store_sales = prev_sales_qs.filter(store=store).aggregate(
            t=Sum('total_amount')
        )['t'] or 0
        if prev_store_sales:
            growth = round(((monthly_sales - prev_store_sales) / prev_store_sales) * 100, 1)
        else:
            growth = 0

        # Inventory for this store
        store_inv = inv_qs.filter(store=store)
        total_units    = store_inv.aggregate(t=Sum('quantity_in_stock'))['t'] or 0
        low_stock_items = store_inv.filter(
            quantity_in_stock__gt=0,
            quantity_in_stock__lte=F('reorder_level')
        ).count()
        out_of_stock = store_inv.filter(quantity_in_stock=0).count()

        stock_value = InventoryBatch.objects.filter(
            store=store, remaining_quantity__gt=0
        ).aggregate(
            val=Sum(F('remaining_quantity') * F('unit_cost'))
        )['val'] or Decimal('0')

        # Days in period
        if start_date and end_date:
            period_days_count = max((end_date - start_date).days, 1)
        else:
            period_days_count = today.day or 1

        daily_avg = round(monthly_sales / period_days_count) if period_days_count else 0

        # Utilization: % of SKUs with stock vs total SKUs
        total_store_skus = store_inv.count()
        stocked_skus     = store_inv.filter(quantity_in_stock__gt=0).count()
        utilization = round((stocked_skus / total_store_skus * 100) if total_store_skus else 0)

        # Efficiency: % of SKUs above reorder level
        above_reorder = store_inv.filter(quantity_in_stock__gt=F('reorder_level')).count()
        efficiency = round((above_reorder / total_store_skus * 100) if total_store_skus else 0)

        # Performance score out of 10
        score = 0
        if utilization >= 80:  score += 3
        elif utilization >= 50: score += 2
        else: score += 1
        if growth > 0:   score += 3
        elif growth > -5: score += 1
        if low_stock_items == 0: score += 2
        elif low_stock_items <= 3: score += 1
        if efficiency >= 70: score += 2
        elif efficiency >= 40: score += 1
        performance_score = min(score, 10)

        # Staff count — User model related to store (adjust if you have a Staff model)
        staff_count = 0  # replace with actual staff query if available

        store_data = {
            'id':               store.id,
            'name':             store.name,
            'address':          store.address,
            'branch':           store.branch,
            'is_active':        store.is_active,
            'is_default':       store.is_default,
            'color':            '#4A90E2',   # no color field on model; use default
            'monthly_sales':    monthly_sales,
            'transactions':     transactions,
            'daily_avg':        daily_avg,
            'avg_transaction':  avg_transaction,
            'stock_value':      float(stock_value),
            'total_units':      total_units,
            'low_stock_items':  low_stock_items,
            'out_of_stock':     out_of_stock,
            'utilization':      utilization,
            'efficiency':       efficiency,
            'performance_score': performance_score,
            'sales_trend':      growth,
            'growth':           growth,
            'staff_count':      staff_count,
        }

        efficiency_list.append(efficiency)
        growth_list.append(growth)
        store_performance_data.append(store_data)

    # Apply performance filter after building data
    if performance_param == 'high':
        store_performance_data = [s for s in store_performance_data if s['performance_score'] >= 8]
    elif performance_param == 'medium':
        store_performance_data = [s for s in store_performance_data if 5 <= s['performance_score'] < 8]
    elif performance_param == 'low':
        store_performance_data = [s for s in store_performance_data if s['performance_score'] < 5]

    # ── Aggregate footer totals ───────────────────────────────────────────────
    avg_utilization      = round(sum(s['utilization'] for s in store_performance_data) / len(store_performance_data)) if store_performance_data else 0
    avg_efficiency       = round(sum(s['efficiency']  for s in store_performance_data) / len(store_performance_data)) if store_performance_data else 0
    avg_growth           = round(sum(s['growth']      for s in store_performance_data) / len(store_performance_data), 1) if store_performance_data else 0
    stores_above_target  = sum(1 for s in store_performance_data if s['utilization'] >= 70)
    stores_with_growth   = sum(1 for s in store_performance_data if s['growth'] > 0)
    stores_below_target  = sum(1 for s in store_performance_data if s['performance_score'] < 5)

    # ── Summary cards ─────────────────────────────────────────────────────────
    top_performers  = sorted(store_performance_data, key=lambda x: x['performance_score'], reverse=True)[:5]
    need_attention  = sorted(store_performance_data, key=lambda x: x['low_stock_items'], reverse=True)[:5]

    # Branch summary
    branches_qs = Branch.objects.filter(
        store_locations__in=stores_qs
    ).distinct()

    branch_summary = []
    for branch in branches_qs:
        branch_sales = sales_qs.filter(store__branch=branch).aggregate(
            t=Sum('total_amount')
        )['t'] or 0
        branch_summary.append({
            'name':  branch.name,
            'sales': branch_sales,
        })
    branch_summary.sort(key=lambda x: x['sales'], reverse=True)

    # ── Filter dropdowns ──────────────────────────────────────────────────────
    branches = Branch.objects.all().order_by('name')

    context = {
        # Dates
        'start_date': display_start,
        'end_date':   display_end,

        # KPI cards
        'total_stores':       total_stores,
        'active_stores':      active_stores,
        'inactive_stores':    inactive_stores,
        'total_sales':        total_sales,
        'avg_utilization':    avg_utilization,
        'stores_above_target': stores_above_target,
        'total_stock_units':  total_stock_units,
        'total_skus':         total_skus,

        # Table footer totals
        'total_transactions':   total_transactions,
        'avg_transaction_total': avg_transaction_total,
        'total_stock_value':    total_stock_value,
        'total_low_stock':      total_low_stock,
        'avg_efficiency':       avg_efficiency,
        'avg_growth':           avg_growth,

        # Main table
        'store_performance_data': store_performance_data,

        # Summary cards
        'top_performers':    top_performers,
        'need_attention':    need_attention,
        'branch_summary':    branch_summary,
        'stores_with_growth': stores_with_growth,
        'stores_below_target': stores_below_target,

        # Filter dropdowns
        'branches': branches,
    }

    return render(request, 'reports/stocklocation_details.html', context)


# Dynamic color helper functions
def get_dynamic_chart_color(store_id):
    """Generate a consistent chart color based on store ID"""
    colors = [
        '#4A90E2',  # Blue
        '#36B9CC',  # Cyan
        '#1CC88A',  # Green
        '#F6C23E',  # Yellow
        '#E74A3B',  # Red
        '#6C757D',  # Gray
        '#4E73DF',  # Indigo
        '#1CC88A',  # Teal
        '#F6C23E',  # Orange
    ]
    return colors[store_id % len(colors)]


def get_dynamic_border_color(store_id):
    """Generate a consistent border color based on store ID"""
    colors = [
        '#357ABD',  # Darker Blue
        '#2A9CA5',  # Darker Cyan
        '#17A673',  # Darker Green
        '#F4B619',  # Darker Yellow
        '#E02D1B',  # Darker Red
        '#495057',  # Darker Gray
        '#2E59D9',  # Darker Indigo
        '#17A673',  # Darker Teal
        '#F4B619',  # Darker Orange
    ]
    return colors[store_id % len(colors)]


@login_required
def export_stocklocation_pdf(request):
    """Export comprehensive stock location report as PDF"""
    try:
        # Get filter parameters
        store_id = request.GET.get('store')
        category_id = request.GET.get('category')
        
        # Set up response
        response = HttpResponse(content_type='application/pdf')
        filename = f"stock_location_report_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Get current date
        today = timezone.now().date()
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=letter, 
                               rightMargin=40, leftMargin=40,
                               topMargin=40, bottomMargin=40)
        elements = []
        
        # Define styles (same as previous)
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            textColor=colors.black,
            alignment=1,
            spaceAfter=5,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'SubtitleStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.black,
            alignment=1,
            spaceAfter=3,
            fontName='Helvetica'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.black,
            spaceBefore=15,
            spaceAfter=8,
            fontName='Helvetica-Bold'
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        small_style = ParagraphStyle(
            'SmallStyle',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.black,
            fontName='Helvetica'
        )
        
        # Header
        elements.append(Paragraph("STOCK LOCATION REPORT", title_style))
        elements.append(Paragraph(f"As of: {today.strftime('%B %d, %Y')}", subtitle_style))
        elements.append(Paragraph(f"Generated: {timezone.now().strftime('%B %d, %Y %H:%M')}", subtitle_style))
        elements.append(Paragraph(f"Generated By: {request.user.get_full_name() or request.user.username}", subtitle_style))
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("-" * 80, normal_style))
        elements.append(Spacer(1, 10))
        
        # Get all active stores
        stores = StoreLocation.objects.filter(is_active=True)
        
        if store_id:
            stores = stores.filter(id=store_id)
        
        # Calculate totals
        total_products_in_stock = 0
        total_items = 0
        total_value = Decimal('0')
        total_below_reorder = 0
        total_out_of_stock = 0
        
        store_data_dict = {}
        
        for store in stores:
            inventories = Inventory.objects.filter(store=store).select_related('product', 'product__category')
            if category_id:
                inventories = inventories.filter(product__category_id=category_id)
            
            store_products = inventories.values('product').distinct().count()
            store_items = inventories.aggregate(total=Sum('quantity_in_stock'))['total'] or 0
            store_value = Decimal('0')
            store_below = 0
            store_out = 0
            
            for inv in inventories:
                avg_cost = InventoryBatch.objects.filter(
                    product=inv.product,
                    store=store
                ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                store_value += inv.quantity_in_stock * avg_cost
                
                if inv.quantity_in_stock == 0:
                    store_out += 1
                elif inv.quantity_in_stock <= inv.reorder_level:
                    store_below += 1
            
            total_products_in_stock += store_products
            total_items += store_items
            total_value += store_value
            total_below_reorder += store_below
            total_out_of_stock += store_out
            
            store_data_dict[store.id] = {
                'store': store,
                'products': store_products,
                'items': store_items,
                'value': store_value,
                'below': store_below,
                'out': store_out,
                'inventories': list(inventories)
            }
        
        # ========== EXECUTIVE SUMMARY ==========
        elements.append(Paragraph("EXECUTIVE SUMMARY", heading_style))
        elements.append(Spacer(1, 5))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Stores', f"{stores.count()}"],
            ['Active Stores', f"{stores.count()}"],
            ['Total Products in Stock', f"{total_products_in_stock}"],
            ['Total Items in Stock (Base Units)', f"{total_items:,}"],
            ['Total Inventory Value', f"UGX {total_value:,.0f}"],
            ['Products Below Reorder Level', f"{total_below_reorder}"],
            ['Products Out of Stock', f"{total_out_of_stock}"],
            ['Stores with Low Stock Items', f"{stores.count() if total_below_reorder > 0 else 0}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[200, 200])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # ========== STORES OVERVIEW ==========
        elements.append(Paragraph("STORES OVERVIEW", heading_style))
        elements.append(Spacer(1, 5))
        
        store_overview = [['Store Name', 'Branch', 'Status', 'Products', 'Total Stock', 'Total Value', 'Utilization', 'Below Reorder']]
        
        for store_data in store_data_dict.values():
            store = store_data['store']
            utilization = (store_data['value'] / total_value * 100) if total_value > 0 else 0
            
            store_overview.append([
                store.name[:15] + '...' if len(store.name) > 15 else store.name,
                store.branch.name if store.branch else 'N/A',
                'Active',
                str(store_data['products']),
                f"{store_data['items']:,}",
                f"UGX {store_data['value']:,.0f}",
                f"{utilization:.1f}%",
                str(store_data['below'])
            ])
        
        store_overview.append([
            'TOTAL', '', '', str(total_products_in_stock), f"{total_items:,}", f"UGX {total_value:,.0f}", '100%', str(total_below_reorder)
        ])
        
        store_overview_table = Table(store_overview, colWidths=[80, 60, 40, 50, 60, 90, 50, 50])
        store_overview_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(store_overview_table)
        elements.append(Spacer(1, 20))
        
        # ========== STORE INVENTORY DETAILS ==========
        for store_data in store_data_dict.values():
            store = store_data['store']
            elements.append(Paragraph(f"STORE INVENTORY DETAIL - {store.name.upper()}", heading_style))
            elements.append(Spacer(1, 5))
            
            store_inv_data = [['Product', 'SKU', 'Brand', 'Category', 'Stock Qty', 'Unit', 'Reorder Level', 'Status', 'Last Updated']]
            
            for inv in store_data['inventories'][:20]:  # Limit per store
                base_unit = 'N/A'
                base_unit_price = inv.product.unit_prices.filter(conversion_factor=1).first()
                if base_unit_price:
                    base_unit = base_unit_price.unit.abbreviation
                
                if inv.quantity_in_stock == 0:
                    status = 'Out of Stock 🔴'
                elif inv.quantity_in_stock <= inv.reorder_level:
                    status = 'Below Reorder ⚠️'
                else:
                    status = 'Good'
                
                store_inv_data.append([
                    inv.product.name[:15] + '...' if len(inv.product.name) > 15 else inv.product.name,
                    inv.product.sku or 'N/A',
                    inv.product.brand or 'N/A',
                    inv.product.category.name[:10] if inv.product.category else 'N/A',
                    f"{inv.quantity_in_stock:,}",
                    base_unit,
                    str(inv.reorder_level),
                    status,
                    inv.last_updated.strftime('%Y-%m-%d') if inv.last_updated else 'N/A'
                ])
            
            store_inv_data.append([
                'TOTAL', '', '', '', f"{store_data['items']:,}", '', '', '', ''
            ])
            
            store_inv_table = Table(store_inv_data, colWidths=[70, 45, 40, 45, 40, 25, 35, 50, 55])
            store_inv_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ]))
            elements.append(store_inv_table)
            elements.append(Spacer(1, 15))
        
        # ========== PRODUCTS WITH MULTIPLE STORE LOCATIONS ==========
        elements.append(Paragraph("PRODUCTS WITH MULTIPLE STORE LOCATIONS", heading_style))
        elements.append(Spacer(1, 5))
        
        # Find products that appear in multiple stores
        product_store_counts = {}
        for store_data in store_data_dict.values():
            for inv in store_data['inventories']:
                product_id = inv.product.id
                if product_id not in product_store_counts:
                    product_store_counts[product_id] = {
                        'product': inv.product,
                        'stores': {},
                        'total': 0
                    }
                product_store_counts[product_id]['stores'][store_data['store'].name] = inv.quantity_in_stock
                product_store_counts[product_id]['total'] += inv.quantity_in_stock
        
        multi_store_data = [['Product', 'SKU', 'Brand'] + [store.name[:10] for store in stores] + ['Total Stock']]
        
        for product_data in product_store_counts.values():
            if len(product_data['stores']) > 1:  # Only products in multiple stores
                row = [
                    product_data['product'].name[:15] + '...' if len(product_data['product'].name) > 15 else product_data['product'].name,
                    product_data['product'].sku or 'N/A',
                    product_data['product'].brand or 'N/A'
                ]
                
                for store in stores:
                    row.append(str(product_data['stores'].get(store.name, 0)))
                
                row.append(str(product_data['total']))
                multi_store_data.append(row)
        
        if len(multi_store_data) > 1:
            col_widths = [70, 45, 45] + [35] * stores.count() + [50]
            multi_store_table = Table(multi_store_data[:15], colWidths=col_widths)  # Limit rows
            multi_store_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            elements.append(multi_store_table)
        else:
            elements.append(Paragraph("No products found in multiple stores.", normal_style))
        elements.append(Spacer(1, 20))
        
        # ========== INVENTORY VALUE BY STORE AND CATEGORY ==========
        elements.append(Paragraph("INVENTORY VALUE BY STORE AND CATEGORY", heading_style))
        elements.append(Spacer(1, 5))
        
        categories = Category.objects.all()
        
        value_data = [['Category'] + [store.name[:10] for store in stores] + ['Total']]
        
        for category in categories:
            row = [category.name[:15] + '...' if len(category.name) > 15 else category.name]
            cat_total = Decimal('0')
            
            for store in stores:
                cat_value = Decimal('0')
                for inv in Inventory.objects.filter(store=store, product__category=category):
                    avg_cost = InventoryBatch.objects.filter(
                        product=inv.product,
                        store=store
                    ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                    cat_value += inv.quantity_in_stock * avg_cost
                
                row.append(f"UGX {cat_value:,.0f}")
                cat_total += cat_value
            
            row.append(f"UGX {cat_total:,.0f}")
            value_data.append(row)
        
        # Add total row
        total_row = ['TOTAL']
        grand_total = Decimal('0')
        for store in stores:
            store_total = Decimal('0')
            for category in categories:
                for inv in Inventory.objects.filter(store=store, product__category=category):
                    avg_cost = InventoryBatch.objects.filter(
                        product=inv.product,
                        store=store
                    ).aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
                    store_total += inv.quantity_in_stock * avg_cost
            total_row.append(f"UGX {store_total:,.0f}")
            grand_total += store_total
        total_row.append(f"UGX {grand_total:,.0f}")
        value_data.append(total_row)
        
        col_widths = [80] + [90] * stores.count() + [90]
        value_table = Table(value_data, colWidths=col_widths)
        value_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('GRID', (0, 0), (-2, -2), 0.5, colors.black),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ]))
        elements.append(value_table)
        
        # Build PDF
        doc.build(elements)
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        response = HttpResponse(content_type='text/plain')
        response.status_code = 500
        response.content = f"Error generating PDF: {str(e)}"
        return response


# ============================================================================
# PRICING REPORTS VIEWS
# ============================================================================
# Product Pricing List, Unit Conversion Report, Price Comparison,
# Missing Pricing Report
# ============================================================================

@login_required
def productpricing_details(request):  
    """Pricing reports view with detailed analysis"""
    
    today = timezone.now()
    currency = request.GET.get('currency', 'UGX')
    
    # Get filter parameters
    category_filter = request.GET.get('category', '')
    price_range_filter = request.GET.get('price_range', '')
    margin_filter = request.GET.get('margin', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset for products
    products = Product.objects.filter(is_active=True).prefetch_related(
        'unit_prices', 'category'
    )
    
    # Apply filters
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query)
        )
    
    if category_filter:
        products = products.filter(category__name=category_filter)
    
    # Prepare product pricing data
    product_pricing_data = []
    total_selling_price = Decimal('0')
    total_cost_price = Decimal('0')
    products_count = 0
    high_margin_count = 0
    low_margin_count = 0
    total_product_value = Decimal('0')
    
    for product in products:
        default_unit_price = product.unit_prices.first()
        
        if default_unit_price:
            selling_price = default_unit_price.price
            
            # Get cost from purchase history or use default
            purchase_items = PurchaseOrderItem.objects.filter(product=product)
            avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg']
            cost_price = avg_cost or (selling_price * Decimal('0.65'))
            
            # Calculate margin
            if cost_price > 0:
                margin = ((selling_price - cost_price) / cost_price) * 100
            else:
                margin = 0
            
            # Apply filters
            if price_range_filter:
                if price_range_filter == 'low' and selling_price >= 100000:
                    continue
                elif price_range_filter == 'medium' and (selling_price < 100000 or selling_price > 500000):
                    continue
                elif price_range_filter == 'high' and selling_price <= 500000:
                    continue
            
            if margin_filter:
                if margin_filter == 'low' and margin >= 30:
                    continue
                elif margin_filter == 'medium' and (margin < 30 or margin > 50):
                    continue
                elif margin_filter == 'high' and margin <= 50:
                    continue
            
            # Determine price tier
            if margin >= 50:
                tier = 'Premium'
                tier_class = 'tier-premium'
                high_margin_count += 1
            elif margin >= 30:
                tier = 'Standard'
                tier_class = 'tier-standard'
            else:
                tier = 'Economy'
                tier_class = 'tier-economy'
                low_margin_count += 1
            
            # Determine margin badge
            if margin >= 50:
                margin_class = 'margin-high'
            elif margin >= 30:
                margin_class = 'margin-medium'
            else:
                margin_class = 'margin-low'
            
            # Static price change for now
            price_change = Decimal('0')
            if price_change > 0:
                price_change_class = 'price-up'
                price_change_icon = 'trending-up'
                change_text = f"+{price_change:.1f}%"
            elif price_change < 0:
                price_change_class = 'price-down'
                price_change_icon = 'trending-down'
                change_text = f"{price_change:.1f}%"
            else:
                price_change_class = 'price-stable'
                price_change_icon = 'minus'
                change_text = 'Stable'
            
            product_data = {
                'sku': product.sku,
                'name': product.name,
                'category': product.category.name if product.category else 'Uncategorized',
                'cost_price': cost_price,
                'selling_price': selling_price,
                'margin': margin,
                'margin_class': margin_class,
                'tier': tier,
                'tier_class': tier_class,
                'price_change': price_change,
                'price_change_class': price_change_class,
                'price_change_icon': price_change_icon,
                'change_text': change_text,
                'last_change_date': (today - timedelta(days=7)).strftime('%Y-%m-%d'),
                'last_change_reason': 'Market adjustment',
            }
            
            product_pricing_data.append(product_data)
            total_selling_price += selling_price
            total_cost_price += cost_price
            total_product_value += selling_price
            products_count += 1
    
    # Sort products
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'price_high':
        product_pricing_data.sort(key=lambda x: x['selling_price'], reverse=True)
    elif sort_by == 'price_low':
        product_pricing_data.sort(key=lambda x: x['selling_price'])
    elif sort_by == 'margin':
        product_pricing_data.sort(key=lambda x: x['margin'], reverse=True)
    else:
        product_pricing_data.sort(key=lambda x: x['name'])
    
    # Calculate statistics
    avg_selling_price = total_selling_price / products_count if products_count > 0 else 0
    avg_cost_price = total_cost_price / products_count if products_count > 0 else 0
    avg_margin = ((avg_selling_price - avg_cost_price) / avg_cost_price * 100) if avg_cost_price > 0 else 0
    
    # Get unit conversion data (products with multiple unit prices)
    unit_conversion_data = []
    for product in products:
        unit_prices = product.unit_prices.all()
        if len(unit_prices) >= 2:
            # Sort by conversion factor (smallest to largest)
            sorted_prices = sorted(unit_prices, key=lambda x: x.conversion_factor)
            base_unit = sorted_prices[0]
            pack_unit = sorted_prices[-1]
            
            if pack_unit.conversion_factor > base_unit.conversion_factor:
                pack_price_per_base_unit = pack_unit.price / pack_unit.conversion_factor
                base_price_per_unit = base_unit.price
                savings_pct = ((base_price_per_unit - pack_price_per_base_unit) / base_price_per_unit * 100) if base_price_per_unit > 0 else 0
                
                unit_conversion_data.append({
                    'product': product.name,
                    'sku': product.sku,
                    'base_unit': base_unit.unit.name,
                    'pack_unit': f"{pack_unit.conversion_factor}{base_unit.unit.abbreviation}",
                    'conversion_factor': f"1:{int(pack_unit.conversion_factor)}",
                    'base_price': base_unit.price,
                    'pack_price': pack_unit.price,
                    'unit_price': pack_price_per_base_unit,
                    'savings_pct': savings_pct,
                })
    
    # Get missing pricing data
    missing_pricing_products = Product.objects.filter(
        is_active=True,
        unit_prices__isnull=True
    )
    
    missing_prices_count = missing_pricing_products.count()
    missing_pricing_data = []
    overdue_count = 0
    new_products_count = 0
    total_days_missing = 0
    
    for product in missing_pricing_products[:10]:  # Limit to 10 for display
        days_without_price = (today - product.created_at).days
        total_days_missing += days_without_price
        
        if days_without_price > 7:
            overdue_count += 1
            status = 'Overdue'
            status_class = 'bg-danger'
        else:
            new_products_count += 1
            status = 'New Product'
            status_class = 'bg-warning'
        
        # Find similar products for price suggestion
        similar_products = Product.objects.filter(
            category=product.category,
            unit_prices__isnull=False
        )[:3]
        
        avg_similar_price = 0
        if similar_products.exists():
            avg_prices = []
            for sim_product in similar_products:
                price = sim_product.unit_prices.first()
                if price:
                    avg_prices.append(price.price)
            if avg_prices:
                avg_similar_price = sum(avg_prices) / len(avg_prices)
        
        # Get cost price from purchase history
        purchase_items = PurchaseOrderItem.objects.filter(product=product)
        avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg'] or Decimal('0')
        
        missing_pricing_data.append({
            'sku': product.sku,
            'name': product.name,
            'category': product.category.name if product.category else 'Uncategorized',
            'date_added': product.created_at.strftime('%Y-%m-%d'),
            'days_without_price': days_without_price,
            'similar_products_count': similar_products.count(),
            'suggested_price': avg_similar_price,
            'cost_price': avg_cost,
            'status': status,
            'status_class': status_class,
        })
    
    # Calculate margin by category
    categories = Category.objects.all()
    category_margins = []
    for category in categories:
        category_products = products.filter(category=category)
        if category_products.exists():
            total_margin = 0
            count = 0
            for product in category_products:
                unit_price = product.unit_prices.first()
                if unit_price:
                    purchase_items = PurchaseOrderItem.objects.filter(product=product)
                    avg_cost = purchase_items.aggregate(avg=Avg('unit_cost'))['avg']
                    cost_price = avg_cost or (unit_price.price * Decimal('0.65'))
                    if cost_price > 0:
                        margin = ((unit_price.price - cost_price) / cost_price) * 100
                        total_margin += margin
                        count += 1
            if count > 0:
                category_margins.append({
                    'name': category.name,
                    'avg_margin': total_margin / count
                })
    
    # Find highest and lowest margin categories
    if category_margins:
        highest = max(category_margins, key=lambda x: x['avg_margin'])
        lowest = min(category_margins, key=lambda x: x['avg_margin'])
        highest_margin_category = highest['name']
        highest_margin_pct = highest['avg_margin']
        lowest_margin_category = lowest['name']
        lowest_margin_pct = lowest['avg_margin']
    else:
        highest_margin_category = 'N/A'
        highest_margin_pct = 0
        lowest_margin_category = 'N/A'
        lowest_margin_pct = 0
    
    # Prepare context
    context = {
        # Report metadata
        'report_id': f"PRICE-{today.strftime('%Y%m%d')}-001",
        'report_date': today.strftime('%B %Y'),
        'currency': currency,
        'products_analyzed': products_count,
        'missing_prices_count': missing_prices_count,
        
        # Overall statistics
        'avg_selling_price': avg_selling_price,
        'avg_cost_price': avg_cost_price,
        'avg_margin': avg_margin,
        'avg_price_change': Decimal('2.5'),  # Static for now
        'price_changes_this_month': 24,  # Static for now
        
        # Additional statistics
        'total_product_value': total_product_value,
        'high_margin_count': high_margin_count,
        'low_margin_count': low_margin_count,
        
        # Margin analysis
        'highest_margin_category': highest_margin_category,
        'highest_margin_pct': highest_margin_pct,
        'lowest_margin_category': lowest_margin_category,
        'lowest_margin_pct': lowest_margin_pct,
        
        # Data for templates
        'product_pricing_data': product_pricing_data,
        'unit_conversion_data': unit_conversion_data,
        'missing_pricing_data': missing_pricing_data,
        
        # Missing pricing statistics
        'new_products_count': new_products_count,
        'overdue_count': overdue_count,
        'avg_days_missing': total_days_missing / missing_prices_count if missing_prices_count > 0 else 0,
        'default_margin': 50,
        
        # Filter options
        'categories': categories,
        'selected_category': category_filter,
        'selected_price_range': price_range_filter,
        'selected_margin': margin_filter,
        'selected_sort': sort_by,
        'search_query': search_query,
    }
    
    return render(request, 'reports/productpricing_details.html', context)